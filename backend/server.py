from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Request, Body
from fastapi.responses import HTMLResponse, Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import re
import logging
import asyncio
import base64
import json
import certifi
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
import resend
import branding
import invoice_logo
import reminders as reminders_module
import prod_sync as prod_sync_module
import calendar_feed as calendar_feed_module
import geocoder as geocoder_module

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env', override=False)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
# Use TLS with certifi CA bundle for MongoDB Atlas (mongodb+srv://)
if mongo_url.startswith("mongodb+srv://") or "mongodb.net" in mongo_url:
    client = AsyncIOMotorClient(mongo_url, tls=True, tlsCAFile=certifi.where())
else:
    client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Resend email config
resend.api_key = os.environ.get('RESEND_API_KEY', '')
NOTIFY_EMAIL = os.environ.get('NOTIFY_EMAIL', '')


def inject_branding(html: str, recipient_email: str = "") -> str:
    """Insert the branded business header into any outgoing HTML email.

    - If the template already contains a <body> tag, the header is inserted
      right after it and the footer right before </body>.
    - Otherwise, the header/footer are prepended/appended.
    - Idempotent: if the template already contains the header signature OR
      already embeds the business logo near the top, it's left alone
      (prevents double-branding on emails that used branding.wrap_email).
    - Includes a CASL-compliant "Se désabonner" link in the footer.
    """
    if not html:
        return html
    # Build unsubscribe URL (per-recipient if known)
    app_url = os.environ.get("APP_URL", "").rstrip("/")
    if recipient_email and app_url:
        from urllib.parse import quote
        unsub_url = f"{app_url}/api/unsubscribe?email={quote(recipient_email)}"
    elif app_url:
        unsub_url = f"{app_url}/api/unsubscribe"
    else:
        unsub_url = ""

    # Already branded? Either our tagged header, OR the logo data URL appears
    # near the top (wrap_email already inserted it)
    top_slice = html[:4000]
    if 'data-gexia-header="1"' in html:
        return html
    if 'data:image/jpeg;base64' in top_slice and branding.LOGO_BASE64[:30] in top_slice:
        return html

    header = branding.build_email_header_html()
    footer = branding.build_email_footer_html(unsub_url)
    header_tagged = header.replace('<div ', '<div data-gexia-header="1" ', 1)
    if "<body" in html:
        if "</body>" in html:
            html = html.replace("</body>", footer + "</body>", 1)
        html = re.sub(r"(<body[^>]*>)", r"\1" + header_tagged, html, count=1)
        return html
    return branding.wrap_email(html, unsubscribe_url=unsub_url)

app = FastAPI()
api_router = APIRouter(prefix="/api")


def _fmt_date_fr(d: str) -> str:
    """Format an ISO date (YYYY-MM-DD) as a French long date, e.g. '25 avril 2026'."""
    try:
        dt = datetime.fromisoformat((d or "").strip())
        months_fr = ["janvier", "février", "mars", "avril", "mai", "juin",
                     "juillet", "août", "septembre", "octobre", "novembre", "décembre"]
        return f"{dt.day} {months_fr[dt.month - 1]} {dt.year}"
    except Exception:
        return d or ""


# --- Unsubscribe (CASL/CAN-SPAM compliance) ---

@api_router.get("/unsubscribe", response_class=HTMLResponse)
async def unsubscribe_page(email: str = ""):
    """Show a confirmation page for unsubscribing.
    Email is pre-filled from query param if available."""
    safe_email = (email or "").strip().lower()
    is_unsubbed = False
    if safe_email:
        existing = await db.unsubscribes.find_one({"email": safe_email})
        is_unsubbed = bool(existing)

    if is_unsubbed:
        body = f"""
        <div style="background:#FEF3C7;border-left:4px solid #F59E0B;padding:16px;border-radius:8px">
          <h2 style="margin:0 0 8px 0;color:#92400E">Déjà désabonné</h2>
          <p style="margin:0;color:#78350F">L'adresse <strong>{safe_email}</strong> est déjà désabonnée. Vous ne recevrez plus de courriels promotionnels.</p>
        </div>
        <p style="margin-top:20px;font-size:13px;color:#6B7280">Vous pouvez fermer cette page.</p>
        """
    else:
        body = f"""
        <h2 style="color:#1E5BA8;margin-top:0">Se désabonner</h2>
        <p>Confirmer le désabonnement de l'adresse :</p>
        <form method="POST" action="/api/unsubscribe">
          <input type="email" name="email" value="{safe_email}" required
            style="width:100%;padding:12px;border:1px solid #D1D5DB;border-radius:6px;font-size:14px;box-sizing:border-box;margin-bottom:12px" />
          <button type="submit" style="width:100%;background:#DC2626;color:white;border:none;padding:14px;border-radius:8px;font-size:15px;font-weight:600;cursor:pointer">Confirmer le désabonnement</button>
        </form>
        <p style="margin-top:16px;font-size:12px;color:#6B7280">Vous ne recevrez plus de courriels promotionnels (campagnes saisonnières) de Lavage de Vitres Bois-Franc. Les courriels de service comme les confirmations de RDV continueront d'arriver.</p>
        """
    return branding.wrap_email(body, subtitle="Préférences de courriel")


@api_router.post("/unsubscribe", response_class=HTMLResponse)
async def unsubscribe_submit(request: Request):
    """Process unsubscribe request from the form."""
    form = await request.form()
    email = (form.get("email") or "").strip().lower()
    if not email or "@" not in email:
        return HTMLResponse(branding.wrap_email(
            "<p style='color:#DC2626'>Adresse courriel invalide.</p>",
            subtitle="Erreur"
        ), status_code=400)
    # Idempotent insert
    await db.unsubscribes.update_one(
        {"email": email},
        {"$set": {"email": email, "unsubscribed_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    body = f"""
    <div style="background:#D1FAE5;border-left:4px solid #10B981;padding:16px;border-radius:8px;text-align:center">
      <h2 style="margin:0 0 8px 0;color:#065F46">✅ Désabonnement confirmé</h2>
      <p style="margin:0;color:#047857">L'adresse <strong>{email}</strong> ne recevra plus de courriels promotionnels.</p>
    </div>
    <p style="margin-top:20px;font-size:13px;color:#6B7280;text-align:center">Vous pouvez fermer cette page. Si c'était une erreur, écrivez-nous à {branding.BUSINESS_EMAIL}.</p>
    """
    return branding.wrap_email(body, subtitle="Désabonnement réussi")


async def is_email_unsubscribed(email: str) -> bool:
    """Returns True if the given email has unsubscribed from marketing."""
    if not email:
        return False
    e = email.strip().lower()
    doc = await db.unsubscribes.find_one({"email": e})
    return bool(doc)


# --- Models ---

class AppointmentCreate(BaseModel):
    title: str
    client_name: str
    client_email: Optional[str] = ""
    client_phone: Optional[str] = ""
    client_address: Optional[str] = ""
    date: str  # YYYY-MM-DD
    time_slot: str  # HH:MM
    duration_minutes: int = 30
    price: Optional[float] = 0.0
    notes: Optional[str] = ""
    status: str = "upcoming"  # upcoming, completed, cancelled

class AppointmentUpdate(BaseModel):
    title: Optional[str] = None
    client_name: Optional[str] = None
    client_email: Optional[str] = None
    client_phone: Optional[str] = None
    client_address: Optional[str] = None
    date: Optional[str] = None
    time_slot: Optional[str] = None
    duration_minutes: Optional[int] = None
    price: Optional[float] = None
    notes: Optional[str] = None
    status: Optional[str] = None
    client_photo: Optional[str] = None
    # Non-stored flag: when true and date/time changed, send a reschedule email to client
    notify_client: Optional[bool] = None

class AppointmentResponse(BaseModel):
    id: str
    title: str
    client_name: str
    client_email: str = ""
    client_phone: str = ""
    client_address: str = ""
    date: str
    time_slot: str
    duration_minutes: int
    price: float = 0.0
    notes: str
    status: str
    created_at: str
    assigned_to: Optional[str] = None
    assigned_id: Optional[str] = None
    assigned_color: Optional[str] = None
    client_photo: Optional[str] = None
    client_id: Optional[str] = None
    archived_at: Optional[str] = None
    # Payment tracking (set by /encaisser endpoint)
    paid_at: Optional[str] = None
    paid_amount: Optional[float] = None
    paid_method: Optional[str] = None
    revenue_id: Optional[str] = None
    # Client-side response tracking (from public portal page /client-confirm)
    client_confirmed: Optional[bool] = None
    client_confirmed_at: Optional[str] = None
    client_requested_alternative: Optional[bool] = None
    client_alt_requested_at: Optional[str] = None
    client_suggested_date: Optional[str] = None
    client_suggested_time: Optional[str] = None
    client_suggested_note: Optional[str] = None
    # Up to 3 tentatively-proposed alternative slots while waiting for the client's reply.
    # Each entry: {"date": "YYYY-MM-DD", "time_slot": "HH:MM", "duration_minutes": int}
    proposed_alternatives: Optional[List[Dict[str, Any]]] = None

# --- Proposed alternatives model ---

class ProposedAlternativesUpdate(BaseModel):
    alternatives: List[Dict[str, Any]]  # [{date, time_slot, duration_minutes}, ...] up to 3

# --- Request Models ---

class RequestCreate(BaseModel):
    customer_name: str
    customer_email: str
    customer_phone: Optional[str] = ""
    customer_address: Optional[str] = ""
    preferred_date: str  # YYYY-MM-DD
    preferred_time: str  # HH:MM
    message: Optional[str] = ""
    request_type: Optional[str] = "rdv"  # 'rdv' or 'est'

class RequestSuggest(BaseModel):
    suggested_date: str
    suggested_time: str
    note: Optional[str] = ""

class RequestEstimate(BaseModel):
    price: float
    note: Optional[str] = ""
    valid_until: Optional[str] = None  # ISO date, optional

class RequestResponse(BaseModel):
    id: str
    customer_name: str
    customer_email: str
    customer_phone: str
    customer_address: str
    preferred_date: str
    preferred_time: str
    message: str
    status: str  # pending, accepted, alternative_offered, declined
    suggested_date: Optional[str] = None
    suggested_time: Optional[str] = None
    suggested_note: Optional[str] = None
    created_at: str
    request_type: Optional[str] = "rdv"
    client_id: Optional[str] = None
    quoted_price: Optional[float] = None
    quote_note: Optional[str] = None
    quote_valid_until: Optional[str] = None
    quoted_at: Optional[str] = None


# --- Voice Transcription ---

@api_router.post("/transcribe")
async def transcribe_audio(file: UploadFile = File(...)):
    """Transcribe audio to text using OpenAI Whisper"""
    import tempfile
    from emergentintegrations.llm.openai import OpenAISpeechToText

    api_key = os.environ.get('EMERGENT_LLM_KEY', '')
    if not api_key:
        raise HTTPException(status_code=400, detail="API key not configured")

    try:
        # Save uploaded file to temp
        suffix = ".m4a"
        if file.filename:
            suffix = "." + file.filename.split(".")[-1]
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name

        stt = OpenAISpeechToText(api_key=api_key)
        with open(tmp_path, "rb") as audio_file:
            response = await stt.transcribe(
                file=audio_file,
                model="whisper-1",
                language="fr",
                response_format="json",
                prompt="Ceci est une prise de rendez-vous. Le client donne son nom, téléphone, adresse et détails du service.",
            )

        os.unlink(tmp_path)
        return {"text": response.text}
    except Exception as e:
        logger.error(f"Transcription error: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur de transcription: {str(e)}")


# --- Booking Page (for customers) ---

@api_router.get("/booking", response_class=HTMLResponse)
async def booking_page():
    """Public booking page for customers to request appointments"""
    html_path = ROOT_DIR / "booking.html"
    content = html_path.read_text()
    # Inject the backend API URL so form submits to Gexia360 even when embedded on external sites
    app_url = os.environ.get("APP_URL", "").rstrip("/")
    content = content.replace("window.location.origin", f'"{app_url}"')
    return HTMLResponse(content=content)


@api_router.get("/client-signup", response_class=HTMLResponse)
async def client_signup_page():
    """Public signup form - embed on website to collect clients into the database"""
    html_path = ROOT_DIR / "client-signup.html"
    content = html_path.read_text()
    app_url = os.environ.get("APP_URL", "").rstrip("/")
    if app_url:
        content = content.replace("window.location.origin", f'"{app_url}"')
    return HTMLResponse(content=content)


@api_router.get("/booking-qr")
async def booking_qr_code():
    """Generate a QR code pointing to the booking page"""
    import qrcode
    from fastapi.responses import Response
    from io import BytesIO

    app_url = os.environ.get("APP_URL", "").rstrip("/")
    booking_url = f"{app_url}/api/booking"

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=12,
        border=4,
    )
    qr.add_data(booking_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return Response(content=buf.getvalue(), media_type="image/png", headers={
        "Content-Disposition": 'inline; filename="brightcalendar-qr.png"'
    })


@api_router.get("/company-logo")
async def company_logo():
    """Serve the company logo (JPEG) for use in campaigns and branding."""
    from fastapi.responses import FileResponse
    logo_path = ROOT_DIR / "assets" / "company-logo.jpeg"
    if not logo_path.exists():
        raise HTTPException(status_code=404, detail="Logo not found")
    return FileResponse(str(logo_path), media_type="image/jpeg", headers={"Cache-Control": "public, max-age=3600"})


@api_router.get("/business-plan")
async def business_plan():
    """Serve the Gexia360 commercialization business plan (Word document)."""
    from fastapi.responses import FileResponse
    doc_path = ROOT_DIR / "assets" / "Gexia360_Plan_Commercialisation.docx"
    if not doc_path.exists():
        raise HTTPException(status_code=404, detail="Business plan not found")
    return FileResponse(
        str(doc_path),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        filename="Gexia360_Plan_Commercialisation.docx",
        headers={"Content-Disposition": 'attachment; filename="Gexia360_Plan_Commercialisation.docx"'}
    )


@api_router.get("/brochure")
async def brochure():
    """Serve the Gexia360 commercial brochure (Word document, for prospects)."""
    from fastapi.responses import FileResponse
    doc_path = ROOT_DIR / "assets" / "Gexia360_Brochure_Commerciale.docx"
    if not doc_path.exists():
        raise HTTPException(status_code=404, detail="Brochure not found")
    return FileResponse(
        str(doc_path),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        filename="Gexia360_Brochure_Commerciale.docx",
        headers={"Content-Disposition": 'attachment; filename="Gexia360_Brochure_Commerciale.docx"'}
    )


@api_router.get("/business-plan-pdf")
async def business_plan_pdf():
    """Serve the Gexia360 commercialization business plan (PDF)."""
    from fastapi.responses import FileResponse
    doc_path = ROOT_DIR / "assets" / "Gexia360_Plan_Commercialisation.pdf"
    if not doc_path.exists():
        raise HTTPException(status_code=404, detail="Business plan PDF not found")
    return FileResponse(
        str(doc_path),
        media_type="application/pdf",
        filename="Gexia360_Plan_Commercialisation.pdf",
        headers={"Content-Disposition": 'inline; filename="Gexia360_Plan_Commercialisation.pdf"'}
    )


@api_router.get("/brochure-pdf")
async def brochure_pdf():
    """Serve the Gexia360 commercial brochure (PDF, for prospects)."""
    from fastapi.responses import FileResponse
    doc_path = ROOT_DIR / "assets" / "Gexia360_Brochure_Commerciale.pdf"
    if not doc_path.exists():
        raise HTTPException(status_code=404, detail="Brochure PDF not found")
    return FileResponse(
        str(doc_path),
        media_type="application/pdf",
        filename="Gexia360_Brochure_Commerciale.pdf",
        headers={"Content-Disposition": 'inline; filename="Gexia360_Brochure_Commerciale.pdf"'}
    )


@api_router.get("/booking-qr-card")
async def booking_qr_card(request: Request, format: str = "png"):
    """Generate a complete branded QR card — ready to use on a website. Supports ?format=jpeg or ?format=png (default)."""
    from PIL import Image, ImageDraw, ImageFont
    import qrcode
    import io

    # Auto-detect base URL from the incoming request (most reliable)
    forwarded_host = request.headers.get("x-forwarded-host") or request.headers.get("host", "")
    scheme = request.headers.get("x-forwarded-proto", "https")
    if forwarded_host and "preview" in forwarded_host or "emergent" in forwarded_host:
        app_url = f"{scheme}://{forwarded_host}"
    else:
        app_url = os.environ.get("APP_URL", "").rstrip("/") or "https://booking-hub-406.preview.emergentagent.com"
    booking_url = f"{app_url}/api/booking"

    # Card dimensions (portrait, web-friendly)
    W, H = 800, 1100

    # Gradient background — cyan top, deeper cyan bottom
    img = Image.new("RGB", (W, H), "#0891B2")
    draw = ImageDraw.Draw(img)
    top_color = (8, 145, 178)        # #0891B2
    bot_color = (6, 95, 125)         # darker cyan
    for y in range(H):
        ratio = y / H
        r = int(top_color[0] * (1 - ratio) + bot_color[0] * ratio)
        g = int(top_color[1] * (1 - ratio) + bot_color[1] * ratio)
        b = int(top_color[2] * (1 - ratio) + bot_color[2] * ratio)
        draw.line([(0, y), (W, y)], fill=(r, g, b))

    # Try to find a nice font
    def _font(size: int, bold: bool = False):
        candidates = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            "/System/Library/Fonts/Helvetica.ttc",
        ]
        for path in candidates:
            try:
                return ImageFont.truetype(path, size)
            except Exception:
                continue
        return ImageFont.load_default()

    f_tagline = _font(22, bold=True)
    f_title = _font(42, bold=True)
    f_subtitle = _font(22, bold=True)
    f_hint = _font(18)
    f_contact = _font(20, bold=True)
    f_contact_sm = _font(17)
    f_footer = _font(14)

    def _text_w(text, font):
        bbox = draw.textbbox((0, 0), text, font=font)
        return bbox[2] - bbox[0]

    # --- Tagline ---
    tagline = "LAVAGE DE VITRES BOIS-FRANC"
    tw = _text_w(tagline, f_tagline)
    draw.text(((W - tw) / 2, 40), tagline, fill="white", font=f_tagline)

    # --- Title ---
    title = "Prenez rendez-vous"
    tw = _text_w(title, f_title)
    draw.text(((W - tw) / 2, 80), title, fill="white", font=f_title)

    # Small calendar emoji (colored circle indicator)
    # (skip emoji for PIL compatibility)

    # --- QR Code ---
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=10, border=2)
    qr.add_data(booking_url)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    QR_SIZE = 480
    qr_img = qr_img.resize((QR_SIZE, QR_SIZE), Image.Resampling.LANCZOS)

    # White rounded panel under QR
    PANEL_PAD = 30
    panel_x = (W - QR_SIZE - PANEL_PAD * 2) // 2
    panel_y = 180
    panel_w = QR_SIZE + PANEL_PAD * 2
    panel_h = QR_SIZE + PANEL_PAD * 2
    draw.rounded_rectangle(
        [panel_x, panel_y, panel_x + panel_w, panel_y + panel_h],
        radius=24, fill="white"
    )
    img.paste(qr_img, (panel_x + PANEL_PAD, panel_y + PANEL_PAD))

    # --- Scan instructions ---
    scan_y = panel_y + panel_h + 30
    scan_label = "Scannez avec votre téléphone"
    tw = _text_w(scan_label, f_subtitle)
    draw.text(((W - tw) / 2, scan_y), scan_label, fill="white", font=f_subtitle)

    hint = "Ouvrez l'appareil photo et pointez-le vers le QR"
    tw = _text_w(hint, f_hint)
    draw.text(((W - tw) / 2, scan_y + 40), hint, fill=(220, 240, 250), font=f_hint)

    # --- Contact panel (white rounded at bottom) ---
    contact_y = scan_y + 100
    contact_h = 160
    contact_pad = 40
    draw.rounded_rectangle(
        [contact_pad, contact_y, W - contact_pad, contact_y + contact_h],
        radius=16, fill="white"
    )

    phone_text = "☎  514-570-9802"
    mail_text = "✉  lavagedevitreboisfranc@live.com"
    web_text = "🌐  Lavagedevitre.org"

    # Render each line centered
    line_y = contact_y + 20
    for line, font in [(phone_text, f_contact), (mail_text, f_contact_sm), (web_text, f_contact_sm)]:
        tw = _text_w(line, font)
        draw.text(((W - tw) / 2, line_y), line, fill=(8, 145, 178), font=font)
        line_y += 40

    # --- Footer URL ---
    footer_y = H - 40
    footer = booking_url
    tw = _text_w(footer, f_footer)
    if tw < W - 40:
        draw.text(((W - tw) / 2, footer_y), footer, fill=(200, 230, 240), font=f_footer)

    # Output PNG or JPEG
    fmt = (format or "png").lower()
    buf = io.BytesIO()
    if fmt in ("jpg", "jpeg"):
        # JPEG doesn't support alpha, so ensure RGB
        img.convert("RGB").save(buf, format="JPEG", quality=92, optimize=True)
        media = "image/jpeg"
        filename = "qr-card-brightcalendar.jpg"
    else:
        img.save(buf, format="PNG", quality=95)
        media = "image/png"
        filename = "qr-card-brightcalendar.png"
    buf.seek(0)
    return Response(content=buf.getvalue(), media_type=media, headers={
        "Content-Disposition": f"inline; filename={filename}",
        "Cache-Control": "public, max-age=300",
    })


@api_router.get("/booking-qr-page", response_class=HTMLResponse)
async def booking_qr_page():
    """A printable page with the QR code + instructions"""
    app_url = os.environ.get("APP_URL", "").rstrip("/")
    qr_url = f"{app_url}/api/booking-qr"
    booking_url = f"{app_url}/api/booking"
    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>QR Code — Prendre rendez-vous</title>
<style>
*{{box-sizing:border-box;}}body{{font-family:-apple-system,'Segoe UI',sans-serif;max-width:600px;margin:40px auto;padding:30px;text-align:center;color:#1F2937;}}
.card{{background:linear-gradient(135deg,#0891B2 0%,#06B6D4 100%);color:#FFF;padding:40px 30px;border-radius:24px;box-shadow:0 10px 40px rgba(0,0,0,0.1);}}
h1{{margin:0 0 8px 0;font-size:28px;}}
.tagline{{font-size:14px;opacity:0.9;margin-bottom:24px;text-transform:uppercase;letter-spacing:2px;}}
.qr-wrap{{background:#FFF;padding:20px;border-radius:20px;display:inline-block;margin:16px 0;}}
.qr-wrap img{{display:block;width:260px;height:260px;}}
.cta{{font-size:22px;font-weight:800;margin:20px 0 6px 0;}}
.hint{{font-size:13px;opacity:0.9;}}
.contact{{background:#FFF;color:#0A0A0A;border-radius:16px;padding:20px;margin-top:24px;font-size:14px;line-height:1.8;}}
.contact strong{{color:#0891B2;}}
.url-fallback{{margin-top:20px;font-size:11px;color:#737373;word-break:break-all;}}
@media print{{body{{margin:0;padding:20px;}}.card{{box-shadow:none;-webkit-print-color-adjust:exact;print-color-adjust:exact;}}}}
</style></head><body>
<div class="card">
  <div class="tagline">Lavage de Vitres Bois-Franc</div>
  <h1>📅 Prenez rendez-vous</h1>
  <div class="qr-wrap"><img src="{qr_url}" alt="QR Code"/></div>
  <div class="cta">Scannez avec votre téléphone</div>
  <div class="hint">Ouvrez l'appareil photo et pointez-le vers le QR code</div>
</div>
<div class="contact">
  <strong>☎ 514-570-9802</strong><br>
  ✉ lavagedevitreboisfranc@live.com<br>
  🌐 Lavagedevitre.org
</div>
<div class="url-fallback">Ou visitez: {booking_url}</div>
</body></html>"""
    return HTMLResponse(content=html)

# --- Routes ---

@api_router.get("/")
async def root():
    return {"message": "Appointment Manager API"}


# ============================================================
# AUTO-LINK HELPER — Phase 4: Link appointments/requests to stored clients
# ============================================================

async def _auto_link_client(name: str, email: str = "", phone: str = "",
                             address: str = "", source: str = "appointment") -> Optional[str]:
    """Find existing client by email/phone, or create a new one. Returns client_id."""
    import re as _re
    email_n = (email or "").strip().lower()
    phone_n = _re.sub(r"\D", "", phone or "")
    name_t = (name or "").strip()

    existing = None
    if email_n:
        existing = await db.clients.find_one({"email_norm": email_n})
    if not existing and phone_n:
        existing = await db.clients.find_one({"phone_norm": phone_n})

    now = datetime.now(timezone.utc).isoformat()

    if existing:
        # Enrich non-empty missing fields
        update_fields = {"updated_at": now}
        if phone and not existing.get("phone"):
            update_fields["phone"] = phone.strip()
            update_fields["phone_norm"] = phone_n
        if email and not existing.get("email"):
            update_fields["email"] = email.strip()
            update_fields["email_norm"] = email_n
        if address and not existing.get("address"):
            update_fields["address"] = address.strip()
        tags = existing.get("tags", []) or []
        if source and source not in tags:
            tags.append(source)
            update_fields["tags"] = tags
        if len(update_fields) > 1:
            await db.clients.update_one({"id": existing["id"]}, {"$set": update_fields})
        return existing["id"]

    # Only create if we have at least a name or email
    if not name_t and not email_n:
        return None

    new_id = str(uuid.uuid4())
    await db.clients.insert_one({
        "id": new_id,
        "name": name_t or email_n or "(sans nom)",
        "email": (email or "").strip(),
        "email_norm": email_n,
        "phone": (phone or "").strip(),
        "phone_norm": phone_n,
        "address": (address or "").strip(),
        "notes": "",
        "tags": [source] if source else ["auto"],
        "created_at": now,
        "updated_at": now,
    })
    logger.info(f"Auto-created client '{name_t}' from {source}")
    return new_id


@api_router.post("/appointments", response_model=AppointmentResponse)
async def create_appointment(data: AppointmentCreate):
    # Auto-link to stored clients DB
    client_id = await _auto_link_client(
        name=data.client_name,
        email=data.client_email or "",
        phone=data.client_phone or "",
        address=data.client_address or "",
        source="appointment",
    )
    appointment = {
        "id": str(uuid.uuid4()),
        "title": data.title,
        "client_name": data.client_name,
        "client_email": data.client_email or "",
        "client_phone": data.client_phone or "",
        "client_address": data.client_address or "",
        "date": data.date,
        "time_slot": data.time_slot,
        "duration_minutes": data.duration_minutes,
        "price": data.price or 0.0,
        "notes": data.notes or "",
        "status": data.status,
        "client_id": client_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.appointments.insert_one(appointment)
    return AppointmentResponse(**{k: v for k, v in appointment.items() if k != "_id"})

@api_router.get("/appointments", response_model=List[AppointmentResponse])
async def get_appointments(
    date: Optional[str] = None,
    status: Optional[str] = None,
    include_archived: bool = False,
):
    query = {}
    if date:
        query["date"] = date
    if status:
        query["status"] = status
    elif not include_archived:
        # By default, exclude archived appointments from the main calendar
        query["status"] = {"$ne": "archived"}
    appointments = await db.appointments.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    return [AppointmentResponse(**a) for a in appointments]

@api_router.get("/appointments/{appointment_id}", response_model=AppointmentResponse)
async def get_appointment(appointment_id: str):
    appointment = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not appointment:
        raise HTTPException(status_code=404, detail="Appointment not found")
    return AppointmentResponse(**appointment)

@api_router.put("/appointments/{appointment_id}", response_model=AppointmentResponse)
async def update_appointment(appointment_id: str, data: AppointmentUpdate):
    raw = data.dict()
    notify_client = bool(raw.pop("notify_client", False))
    update_data = {k: v for k, v in raw.items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")

    # Read the current appointment first so we can detect date/time changes
    current = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not current:
        raise HTTPException(status_code=404, detail="Appointment not found")

    old_date = current.get("date") or ""
    old_time = (current.get("time_slot") or "")[:5]
    new_date = update_data.get("date", old_date) or old_date
    new_time = (update_data.get("time_slot", old_time) or old_time)[:5]
    is_reschedule = (new_date != old_date) or (new_time != old_time)

    result = await db.appointments.update_one({"id": appointment_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Appointment not found")
    appointment = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})

    # Auto-create a Revenue entry when an appointment transitions to "completed"
    # Idempotent: skip if appointment already has a revenue_id, or if no price.
    auto_revenue_created = False
    auto_revenue_id = None
    auto_revenue_amount = 0.0
    new_status = update_data.get("status")
    old_status = current.get("status")
    just_completed = (new_status == "completed" and old_status != "completed")
    if just_completed and not current.get("revenue_id"):
        price = float(appointment.get("price") or 0)
        if price > 0:
            try:
                rev_id = str(uuid.uuid4())
                now_iso = datetime.now(timezone.utc).isoformat()
                rev_doc = {
                    "id": rev_id,
                    "amount": price,
                    "category": "Lavage de vitres",
                    "date": appointment.get("date") or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                    "description": f"RDV complété — {appointment.get('title', 'Service')}",
                    "client_name": appointment.get("client_name", "") or "",
                    "payment_method": "cash",  # Default — user can edit later
                    "appointment_id": appointment_id,
                    "created_at": now_iso,
                    "updated_at": now_iso,
                    "auto_created": True,
                }
                await db.revenues.insert_one(rev_doc)
                # Link revenue back to appointment so we never double-create
                await db.appointments.update_one(
                    {"id": appointment_id},
                    {"$set": {"revenue_id": rev_id}},
                )
                appointment["revenue_id"] = rev_id
                auto_revenue_created = True
                auto_revenue_id = rev_id
                auto_revenue_amount = price
                logger.info(f"Auto-created revenue {rev_id[:8]} (${price}) for completed appointment {appointment_id[:8]}")
            except Exception as e:
                logger.error(f"Failed to auto-create revenue for appointment {appointment_id}: {e}")

    # If this update is a real date/time reschedule, send a branded email to the client
    if is_reschedule and notify_client:
        client_email = (appointment.get("client_email") or "").strip()
        if client_email and resend.api_key:
            try:
                client_name = appointment.get("client_name", "")
                duration = int(appointment.get("duration_minutes") or 60)
                address = appointment.get("client_address", "")
                fmt_date_fr = _fmt_date_fr(new_date)
                old_fmt = _fmt_date_fr(old_date)

                html = f"""<!DOCTYPE html>
<html><body style="margin:0;padding:0;background:#F1F5F9;font-family:'Segoe UI',Arial,sans-serif;">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#F1F5F9;padding:24px 12px;">
    <tr><td align="center">
      <table role="presentation" width="600" cellpadding="0" cellspacing="0" style="max-width:600px;background:#FFFFFF;border-radius:14px;overflow:hidden;box-shadow:0 6px 24px rgba(15,23,42,0.08);">
        <tr><td style="background:linear-gradient(135deg,#0891B2 0%,#0E7490 100%);padding:24px 28px;">
          <h1 style="margin:0;color:#FFFFFF;font-size:22px;font-weight:700;">📅 Rendez-vous reprogrammé</h1>
          <p style="margin:6px 0 0 0;color:#E0F2FE;font-size:14px;">Votre rendez-vous a été déplacé</p>
        </td></tr>
        <tr><td style="padding:28px;">
          <p style="margin:0 0 16px 0;font-size:15px;color:#0F172A;">Bonjour <strong>{client_name}</strong>,</p>
          <p style="margin:0 0 18px 0;font-size:14px;color:#334155;line-height:1.6;">
            Nous vous informons que votre rendez-vous a été <strong>reprogrammé</strong>.
            Voici les nouveaux détails&nbsp;:
          </p>

          <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="margin:0 0 18px 0;">
            <tr><td style="background:#FEE2E2;border-left:4px solid #DC2626;padding:14px 16px;border-radius:8px;">
              <div style="font-size:11px;font-weight:700;color:#991B1B;text-transform:uppercase;letter-spacing:0.5px;">Ancien créneau</div>
              <div style="font-size:15px;color:#7F1D1D;margin-top:4px;text-decoration:line-through;">{old_fmt} à {old_time}</div>
            </td></tr>
            <tr><td style="height:10px;"></td></tr>
            <tr><td style="background:#D1FAE5;border-left:4px solid #059669;padding:14px 16px;border-radius:8px;">
              <div style="font-size:11px;font-weight:700;color:#065F46;text-transform:uppercase;letter-spacing:0.5px;">✅ Nouveau créneau</div>
              <div style="font-size:18px;font-weight:700;color:#064E3B;margin-top:4px;">{fmt_date_fr}</div>
              <div style="font-size:15px;color:#065F46;margin-top:2px;">⏰ {new_time} ({duration} min)</div>
              {f'<div style="font-size:13px;color:#065F46;margin-top:6px;">📍 {address}</div>' if address else ''}
            </td></tr>
          </table>

          <p style="margin:0 0 6px 0;font-size:13px;color:#475569;line-height:1.6;">
            Si ce nouveau créneau ne vous convient pas, n'hésitez pas à nous contacter pour en convenir d'un autre.
          </p>
          <p style="margin:0;font-size:13px;color:#475569;line-height:1.6;">
            Merci de votre compréhension et au plaisir de vous voir bientôt&nbsp;!
          </p>
        </td></tr>
        <tr><td style="padding:16px 28px;background:#0F172A;text-align:center;">
          <p style="margin:0;font-size:12px;color:#94A3B8;">Merci de votre confiance!</p>
        </td></tr>
      </table>
    </td></tr>
  </table>
</body></html>"""

                from_addr = os.environ.get("RESEND_FROM") or "onboarding@resend.dev"
                await asyncio.to_thread(
                    resend.Emails.send,
                    {
                        "from": from_addr,
                        "to": [client_email],
                        "reply_to": NOTIFY_EMAIL if NOTIFY_EMAIL else None,
                        "subject": f"Rendez-vous reprogrammé — {fmt_date_fr} à {new_time}",
                        "html": inject_branding(html, recipient_email=client_email),
                    },
                )
                logger.info(f"Reschedule email sent to client {client_email}")
            except Exception as e:
                logger.error(f"Failed to send reschedule email to {client_email}: {e}")

    return AppointmentResponse(**appointment)

@api_router.delete("/appointments/{appointment_id}")
async def delete_appointment(appointment_id: str):
    """Soft-delete: marks the appointment as archived instead of removing it.
    Use DELETE /appointments/{id}/permanent to hard-delete (cannot be restored)."""
    result = await db.appointments.update_one(
        {"id": appointment_id},
        {"$set": {"status": "archived", "archived_at": datetime.utcnow().isoformat()}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Appointment not found")
    # Tombstone so prod_sync never re-imports this archived item back
    await prod_sync_module.add_tombstone(db, "appointment", appointment_id)
    return {"message": "Appointment archived", "archived": True}

@api_router.post("/appointments/{appointment_id}/restore", response_model=AppointmentResponse)
async def restore_appointment(appointment_id: str):
    """Restore a previously archived appointment back to 'upcoming' status."""
    appointment = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not appointment:
        raise HTTPException(status_code=404, detail="Appointment not found")
    if appointment.get("status") != "archived":
        return AppointmentResponse(**appointment)
    await db.appointments.update_one(
        {"id": appointment_id},
        {"$set": {"status": "upcoming"}, "$unset": {"archived_at": ""}},
    )
    appointment = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    return AppointmentResponse(**appointment)

@api_router.delete("/appointments/{appointment_id}/permanent")
async def permanent_delete_appointment(appointment_id: str):
    """Hard-delete an appointment permanently. Use with caution — cannot be undone."""
    result = await db.appointments.delete_one({"id": appointment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Appointment not found")
    # Tombstone so prod_sync doesn't re-import this id from production
    await prod_sync_module.add_tombstone(db, "appointment", appointment_id)
    return {"message": "Appointment permanently deleted"}

# --- Request Routes (public + admin) ---

@api_router.post("/requests", response_model=RequestResponse)
async def create_request(data: RequestCreate):
    """Public endpoint: customers submit appointment requests from website"""
    # Auto-link to stored clients DB (source based on request_type)
    source = "booking-est" if (data.request_type or "rdv") == "est" else "booking-rdv"
    client_id = await _auto_link_client(
        name=data.customer_name,
        email=data.customer_email or "",
        phone=data.customer_phone or "",
        address=data.customer_address or "",
        source=source,
    )
    request_doc = {
        "id": str(uuid.uuid4()),
        "customer_name": data.customer_name,
        "customer_email": data.customer_email,
        "customer_phone": data.customer_phone or "",
        "customer_address": data.customer_address or "",
        "preferred_date": data.preferred_date,
        "preferred_time": data.preferred_time,
        "message": data.message or "",
        "status": "pending",
        "suggested_date": None,
        "suggested_time": None,
        "suggested_note": None,
        "request_type": data.request_type or "rdv",
        "client_id": client_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.appointment_requests.insert_one(request_doc)

    # Send branded email notification with rich layout + CTA button
    if NOTIFY_EMAIL and resend.api_key:
        try:
            app_url = os.environ.get("APP_URL", "").rstrip("/")
            request_url = f"{app_url}/request-detail?id={request_doc['id']}" if app_url else ""
            req_type = request_doc.get("request_type") or "rdv"
            type_label = "Demande d'estimation" if req_type == "est" else "Demande de rendez-vous"
            type_icon = "💰" if req_type == "est" else "📅"
            type_color = "#7C3AED" if req_type == "est" else "#0891B2"

            # Build long French date
            long_date = _fmt_date_fr(request_doc['preferred_date'])

            # Phone clickable for tel: + sms:
            phone_clean = (request_doc.get('customer_phone') or '').replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
            phone_block = f"""
                <div style="display:flex;align-items:center;gap:8px;margin-top:8px;">
                  <a href="tel:{phone_clean}" style="background:#10B981;color:#FFFFFF;padding:8px 14px;border-radius:8px;text-decoration:none;font-size:13px;font-weight:700;">📞 Appeler</a>
                  <a href="sms:{phone_clean}" style="background:#3B82F6;color:#FFFFFF;padding:8px 14px;border-radius:8px;text-decoration:none;font-size:13px;font-weight:700;">💬 SMS</a>
                </div>
            """ if phone_clean else ""

            # Email clickable
            email_block = f'<a href="mailto:{request_doc["customer_email"]}" style="color:#0891B2;text-decoration:none;">{request_doc["customer_email"]}</a>' if request_doc.get('customer_email') else '—'

            # Map link
            address = request_doc.get('customer_address') or ''
            map_link = ""
            if address:
                map_url = f"https://maps.apple.com/?q={address.replace(' ', '+')}"
                map_link = f' &nbsp; <a href="{map_url}" style="color:#0891B2;text-decoration:none;font-size:12px;">🗺️ Carte</a>'

            # Optional message block
            message = (request_doc.get('message') or '').strip()
            message_block = ""
            if message:
                safe_msg = message.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('\n', '<br>')
                message_block = f"""
                <div style="background:#FEF3C7;border-left:4px solid #F59E0B;padding:12px 14px;border-radius:8px;margin:14px 0;">
                  <div style="font-size:11px;font-weight:700;color:#92400E;text-transform:uppercase;letter-spacing:0.5px;">💬 Message du client</div>
                  <div style="font-size:14px;color:#78350F;margin-top:4px;line-height:1.5;">{safe_msg}</div>
                </div>
                """

            cta_button = ""
            if request_url:
                cta_button = f"""
                <table role="presentation" cellspacing="0" cellpadding="0" style="margin:24px auto 8px auto;">
                  <tr><td style="background:{type_color};border-radius:10px;text-align:center;">
                    <a href="{request_url}" style="display:inline-block;padding:14px 28px;color:#FFFFFF;font-size:15px;font-weight:700;text-decoration:none;border-radius:10px;">
                      ✅ Voir et accepter
                    </a>
                  </td></tr>
                </table>
                <p style="text-align:center;margin:0;color:#94A3B8;font-size:11px;">
                  Cliquez pour ouvrir la demande dans l'app Gexia360
                </p>
                """

            body = f"""
<div style="text-align:center;margin-bottom:8px;">
  <span style="display:inline-block;background:{type_color};color:#FFFFFF;padding:6px 14px;border-radius:20px;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;">{type_icon} {type_label}</span>
</div>

<h2 style="margin:8px 0 4px 0;color:#0F172A;text-align:center;font-size:20px;">Nouvelle demande reçue!</h2>
<p style="margin:0 0 18px 0;color:#64748B;font-size:13px;text-align:center;">
  Reçue le {datetime.now(ZoneInfo('America/Toronto')).strftime('%d/%m/%Y à %H:%M')}
</p>

<div style="background:#F8FAFC;border-radius:10px;padding:16px;margin:0 0 8px 0;">
  <div style="font-size:11px;font-weight:700;color:#64748B;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:4px;">👤 Client</div>
  <div style="font-size:18px;font-weight:700;color:#0F172A;">{request_doc['customer_name']}</div>
  <div style="font-size:13px;color:#475569;margin-top:4px;">✉ {email_block}</div>
  <div style="font-size:13px;color:#475569;margin-top:4px;">📞 {request_doc.get('customer_phone') or '—'}</div>
  {phone_block}
</div>

<div style="background:#F8FAFC;border-radius:10px;padding:16px;margin:10px 0 0 0;">
  <div style="font-size:11px;font-weight:700;color:#64748B;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:4px;">📍 Adresse</div>
  <div style="font-size:14px;color:#0F172A;">{address or '—'}{map_link}</div>
</div>

<div style="background:{type_color}1A;border-left:4px solid {type_color};padding:14px 16px;border-radius:8px;margin:14px 0;">
  <div style="font-size:11px;font-weight:700;color:{type_color};text-transform:uppercase;letter-spacing:0.5px;">🗓 Créneau souhaité</div>
  <div style="font-size:17px;font-weight:700;color:#0F172A;margin-top:4px;">{long_date}</div>
  <div style="font-size:14px;color:#475569;margin-top:2px;">⏰ {request_doc['preferred_time']}</div>
</div>

{message_block}

{cta_button}

<hr style="border:none;border-top:1px solid #E5E5E5;margin:24px 0 12px 0;">
<p style="margin:0;color:#94A3B8;font-size:11px;text-align:center;">
  Vous pouvez aussi accepter, proposer un autre horaire ou refuser depuis l'écran <strong>Demandes</strong> dans l'app.
</p>
""".strip()

            await asyncio.to_thread(resend.Emails.send, {
                "from": os.environ.get("RESEND_FROM") or "onboarding@resend.dev",
                "to": [NOTIFY_EMAIL],
                "subject": f"{type_icon} {type_label} — {request_doc['customer_name']} — {long_date} {request_doc['preferred_time']}",
                "html": inject_branding(body),
            })
            logger.info(f"Email notification sent to {NOTIFY_EMAIL} (request {request_doc['id'][:8]}, CTA: {bool(request_url)})")
        except Exception as e:
            logger.error(f"Failed to send email notification: {e}")

    return RequestResponse(**{k: v for k, v in request_doc.items() if k != "_id"})

@api_router.get("/requests", response_model=List[RequestResponse])
async def get_requests(status: Optional[str] = None):
    query = {}
    if status:
        query["status"] = status
    requests = await db.appointment_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [RequestResponse(**r) for r in requests]

@api_router.get("/requests/{request_id}", response_model=RequestResponse)
async def get_request(request_id: str):
    req = await db.appointment_requests.find_one({"id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    return RequestResponse(**req)

class AcceptRequest(BaseModel):
    price: Optional[float] = 0.0
    date: Optional[str] = None  # override preferred_date (YYYY-MM-DD)
    time: Optional[str] = None  # override preferred_time (HH:MM)
    duration_minutes: Optional[int] = None

@api_router.put("/requests/{request_id}/accept")
async def accept_request(request_id: str, data: AcceptRequest = AcceptRequest()):
    """Accept a request: creates a confirmed appointment and marks request as accepted.
    Optionally override date/time/duration (used when the user picks a different slot from
    the day's availability view)."""
    req = await db.appointment_requests.find_one({"id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    if req["status"] == "accepted":
        raise HTTPException(status_code=400, detail="Request already accepted")

    final_date = data.date or req["preferred_date"]
    final_time = data.time or req["preferred_time"]
    final_duration = data.duration_minutes if data.duration_minutes is not None else 30

    # Create appointment from request — preserve client_id link from the request
    appointment = {
        "id": str(uuid.uuid4()),
        "title": f"Meeting with {req['customer_name']}",
        "client_name": req["customer_name"],
        "client_email": req.get("customer_email", ""),
        "client_phone": req.get("customer_phone", ""),
        "client_address": req.get("customer_address", ""),
        "date": final_date,
        "time_slot": final_time,
        "duration_minutes": final_duration,
        "price": data.price or 0.0,
        "notes": req.get("message", ""),
        "status": "upcoming",
        "client_id": req.get("client_id"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.appointments.insert_one(appointment)

    # Mark request as accepted
    await db.appointment_requests.update_one(
        {"id": request_id},
        {"$set": {"status": "accepted"}}
    )

    # --- Send branded confirmation email to client (BCC owner) ---
    confirmation_email_sent = False
    confirmation_email_error = None
    confirmation_email_bcc = ""
    client_email = (req.get("customer_email") or "").strip()
    if client_email and resend.api_key:
        try:
            client_name = req.get("customer_name", "")
            address = req.get("customer_address", "") or ""
            phone = req.get("customer_phone", "") or ""
            long_date = _fmt_date_fr(final_date)
            address_block = f'<div style="font-size:13px;color:#065F46;margin-top:6px;">📍 {address}</div>' if address else ''
            body = f"""
<h2 style="margin:0 0 8px 0;color:#0F172A;">✅ Rendez-vous confirmé</h2>
<p style="margin:0 0 14px 0;color:#475569;font-size:14px;line-height:1.6;">
  Bonjour <strong>{client_name}</strong>,
</p>
<p style="margin:0 0 14px 0;color:#334155;font-size:14px;line-height:1.6;">
  Bonne nouvelle&nbsp;! Votre demande de rendez-vous a été <strong>acceptée</strong>.
  Voici les détails confirmés&nbsp;:
</p>
<div style="background:#D1FAE5;border-left:4px solid #059669;padding:14px 16px;border-radius:8px;margin:16px 0;">
  <div style="font-size:11px;font-weight:700;color:#065F46;text-transform:uppercase;letter-spacing:0.5px;">📅 Détails</div>
  <div style="font-size:17px;font-weight:700;color:#064E3B;margin-top:6px;">{long_date}</div>
  <div style="font-size:14px;color:#065F46;margin-top:2px;">⏰ {final_time} ({final_duration} min)</div>
  {address_block}
</div>
<p style="margin:14px 0 0 0;color:#475569;font-size:13px;line-height:1.6;">
  Vous recevrez un rappel automatique <strong>la veille</strong> de votre rendez-vous.
</p>
<p style="margin:8px 0 0 0;color:#475569;font-size:13px;line-height:1.6;">
  Si vous devez modifier ou annuler, répondez à ce courriel ou appelez-nous au <strong>514-570-9802</strong>.
</p>
<p style="margin:8px 0 0 0;color:#475569;font-size:13px;line-height:1.6;">
  Au plaisir de vous voir bientôt&nbsp;!
</p>
""".strip()
            payload = {
                "from": os.environ.get("RESEND_FROM") or "onboarding@resend.dev",
                "to": [client_email],
                "subject": f"✅ Rendez-vous confirmé — {long_date} à {final_time}",
                "html": inject_branding(body, recipient_email=client_email),
            }
            owner_email = (NOTIFY_EMAIL or "").strip()
            if owner_email and owner_email.lower() != client_email.lower():
                payload["bcc"] = [owner_email]
                confirmation_email_bcc = owner_email
            await asyncio.to_thread(resend.Emails.send, payload)
            confirmation_email_sent = True
            logger.info(f"Acceptance email sent to {client_email} (BCC: {confirmation_email_bcc or 'none'})")
        except Exception as e:
            confirmation_email_error = str(e)[:200]
            logger.error(f"Failed to send acceptance email: {e}")

    # Build response dict (no Pydantic model — we want to include the _notification meta)
    response_body = {k: v for k, v in appointment.items() if k != "_id"}
    response_body["_notification"] = {
        "email_sent": confirmation_email_sent,
        "email_to": client_email,
        "email_bcc": confirmation_email_bcc,
        "email_error": confirmation_email_error,
    }
    return response_body

@api_router.put("/requests/{request_id}/suggest", response_model=RequestResponse)
async def suggest_alternative(request_id: str, data: RequestSuggest):
    """Suggest an alternative date/time for a request.
    Also sends an email to the CLIENT with the new proposal."""
    req = await db.appointment_requests.find_one({"id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")

    await db.appointment_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "alternative_offered",
            "suggested_date": data.suggested_date,
            "suggested_time": data.suggested_time,
            "suggested_note": data.note or "",
        }}
    )

    # Send email to CLIENT with the proposed alternative
    client_email = req.get("customer_email", "").strip()
    if client_email and resend.api_key:
        try:
            client_name = req.get("customer_name", "")
            original_date = req.get("preferred_date", "")
            original_time = req.get("preferred_time", "")
            new_date = data.suggested_date
            new_time = data.suggested_time
            note = (data.note or "").strip()

            # Format dates nicely
            def _fmt_date(d: str) -> str:
                try:
                    dt = datetime.fromisoformat(d)
                    months_fr = ["janvier", "février", "mars", "avril", "mai", "juin",
                                 "juillet", "août", "septembre", "octobre", "novembre", "décembre"]
                    return f"{dt.day} {months_fr[dt.month - 1]} {dt.year}"
                except Exception:
                    return d

            note_block = ""
            if note:
                safe_note = note.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('\n', '<br>')
                note_block = f"""
                <div style="background:#FEF3C7;border-left:4px solid #F59E0B;padding:14px 16px;border-radius:8px;margin:20px 0;">
                    <p style="margin:0 0 4px 0;color:#92400E;font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;">💬 Message de Lavage de Vitres Bois-Franc</p>
                    <p style="margin:0;color:#78350F;font-size:15px;line-height:1.6;">{safe_note}</p>
                </div>
                """

            # Build email
            html = f"""<!DOCTYPE html>
<html lang="fr">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#F3F4F6;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#F3F4F6;padding:20px 0;">
    <tr><td align="center">
      <table role="presentation" width="600" cellpadding="0" cellspacing="0" style="max-width:600px;width:100%;background:#FFFFFF;border-radius:16px;overflow:hidden;box-shadow:0 6px 28px rgba(0,0,0,0.08);">
        <tr><td style="height:6px;background:linear-gradient(90deg,#F59E0B,#FBBF24,#FCD34D);"></td></tr>
        <tr><td style="padding:32px 28px 8px 28px;">
          <p style="margin:0;font-size:28px;">📅</p>
          <h1 style="margin:8px 0 4px 0;color:#111827;font-size:22px;">Nouvelle proposition de rendez-vous</h1>
          <p style="margin:0;color:#6B7280;font-size:14px;">Bonjour {client_name},</p>
        </td></tr>
        <tr><td style="padding:12px 28px 0 28px;">
          <p style="margin:0 0 16px 0;color:#374151;font-size:15px;line-height:1.6;">
            Merci pour votre demande de rendez-vous. La date et l'heure que vous avez proposées ne sont malheureusement pas disponibles.<br><br>
            <strong>Je vous propose la nouvelle date suivante :</strong>
          </p>

          <!-- New proposed slot — highlighted -->
          <div style="background:linear-gradient(135deg,#ECFDF5,#D1FAE5);border:2px solid #10B981;border-radius:12px;padding:20px;text-align:center;margin:8px 0 20px 0;">
            <p style="margin:0 0 6px 0;color:#047857;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:1px;">✅ Nouvelle proposition</p>
            <p style="margin:0;color:#065F46;font-size:22px;font-weight:800;">{_fmt_date(new_date)}</p>
            <p style="margin:4px 0 0 0;color:#065F46;font-size:18px;font-weight:600;">à {new_time}</p>
          </div>

          <!-- Original request — muted -->
          <div style="background:#F9FAFB;border:1px solid #E5E7EB;border-radius:8px;padding:12px 16px;margin:0 0 20px 0;">
            <p style="margin:0;color:#6B7280;font-size:12px;">
              <span style="text-decoration:line-through;">Demande initiale : {_fmt_date(original_date)} à {original_time}</span>
            </p>
          </div>

          {note_block}

          <p style="margin:20px 0 8px 0;color:#374151;font-size:15px;line-height:1.6;">
            <strong>Est-ce que cette nouvelle date vous convient ?</strong><br>
            Merci de me confirmer par retour de courriel ou par téléphone.
          </p>
        </td></tr>

        <!-- Contact card -->
        <tr><td style="padding:0 28px 24px 28px;">
          <div style="background:#F3F4F6;border-radius:12px;padding:16px;text-align:center;">
            <p style="margin:0 0 8px 0;color:#111827;font-size:15px;font-weight:700;">Lavage de Vitres Bois-Franc</p>
            <p style="margin:0 0 4px 0;">
              <a href="tel:+15145709802" style="color:#0891B2;text-decoration:none;font-size:15px;font-weight:600;">📞 514-570-9802</a>
            </p>
            <p style="margin:0;">
              <a href="https://Lavagedevitre.org" style="color:#0891B2;text-decoration:none;font-size:15px;font-weight:600;">🌐 Lavagedevitre.org</a>
            </p>
          </div>
        </td></tr>

        <!-- Footer -->
        <tr><td style="padding:16px 28px;background:#0F172A;text-align:center;">
          <p style="margin:0;font-size:12px;color:#94A3B8;">Merci de votre confiance!</p>
        </td></tr>
      </table>
    </td></tr>
  </table>
</body></html>"""

            from_addr = os.environ.get("RESEND_FROM") or "onboarding@resend.dev"
            await asyncio.to_thread(
                resend.Emails.send,
                {
                    "from": from_addr,
                    "to": [client_email],
                    "reply_to": NOTIFY_EMAIL if NOTIFY_EMAIL else None,
                    "subject": f"Nouvelle proposition de rendez-vous — {_fmt_date(new_date)} à {new_time}",
                    "html": inject_branding(html),
                },
            )
            logger.info(f"Alternative offer email sent to client {client_email}")
        except Exception as e:
            logger.error(f"Failed to send alternative-offer email to {client_email}: {e}")

    updated = await db.appointment_requests.find_one({"id": request_id}, {"_id": 0})
    return RequestResponse(**updated)

@api_router.put("/requests/{request_id}/send-estimate", response_model=RequestResponse)
async def send_estimate(request_id: str, data: RequestEstimate):
    """Send a quote/estimate to the client with a price.
    Marks the request as 'estimate_sent' and emails a beautifully formatted quote to the client."""
    req = await db.appointment_requests.find_one({"id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")

    now_iso = datetime.now(timezone.utc).isoformat()
    await db.appointment_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "estimate_sent",
            "quoted_price": float(data.price),
            "quote_note": data.note or "",
            "quote_valid_until": data.valid_until or "",
            "quoted_at": now_iso,
        }}
    )

    # Send email to client with the quote
    client_email = (req.get("customer_email") or "").strip()
    sent = False
    if client_email and resend.api_key:
        try:
            client_name = req.get("customer_name", "")
            original_date = req.get("preferred_date", "")
            original_time = req.get("preferred_time", "")
            address = req.get("customer_address", "")
            note = (data.note or "").strip()
            price_str = f"{float(data.price):,.2f} $".replace(",", " ")

            def _fmt_date(d: str) -> str:
                try:
                    dt = datetime.fromisoformat(d)
                    months_fr = ["janvier", "février", "mars", "avril", "mai", "juin",
                                 "juillet", "août", "septembre", "octobre", "novembre", "décembre"]
                    return f"{dt.day} {months_fr[dt.month - 1]} {dt.year}"
                except Exception:
                    return d

            note_block = ""
            if note:
                safe_note = note.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('\n', '<br>')
                note_block = f"""
                <div style="background:#F0F9FF;border-left:4px solid #0891B2;padding:14px 16px;border-radius:8px;margin:20px 0;">
                    <p style="margin:0 0 4px 0;color:#075985;font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;">💬 Note de Lavage de Vitres Bois-Franc</p>
                    <p style="margin:0;color:#0C4A6E;font-size:15px;line-height:1.6;">{safe_note}</p>
                </div>
                """

            valid_block = ""
            if data.valid_until:
                valid_block = f"""
                <p style="margin:8px 0 0 0;color:#6B7280;font-size:12px;text-align:center;">
                    Valide jusqu'au {_fmt_date(data.valid_until)}
                </p>
                """

            html = f"""<!DOCTYPE html>
<html lang="fr">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#F3F4F6;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#F3F4F6;padding:20px 0;">
    <tr><td align="center">
      <table role="presentation" width="600" cellpadding="0" cellspacing="0" style="max-width:600px;width:100%;background:#FFFFFF;border-radius:16px;overflow:hidden;box-shadow:0 6px 28px rgba(0,0,0,0.08);">
        <tr><td style="height:6px;background:linear-gradient(90deg,#0891B2,#06B6D4,#22D3EE);"></td></tr>
        <tr><td style="padding:32px 28px 8px 28px;">
          <p style="margin:0;font-size:28px;">💰</p>
          <h1 style="margin:8px 0 4px 0;color:#111827;font-size:22px;">Voici votre estimation</h1>
          <p style="margin:0;color:#6B7280;font-size:14px;">Bonjour {client_name},</p>
        </td></tr>
        <tr><td style="padding:12px 28px 0 28px;">
          <p style="margin:0 0 20px 0;color:#374151;font-size:15px;line-height:1.6;">
            Merci pour votre demande d'estimation pour le lavage de vitres. Voici mon prix proposé :
          </p>

          <!-- Price — big highlight -->
          <div style="background:linear-gradient(135deg,#ECFDF5,#D1FAE5);border:2px solid #10B981;border-radius:12px;padding:24px;text-align:center;margin:8px 0 20px 0;">
            <p style="margin:0 0 6px 0;color:#047857;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:1px;">Prix estimé</p>
            <p style="margin:0;color:#065F46;font-size:36px;font-weight:800;">{price_str}</p>
            {valid_block}
          </div>

          <!-- Service details -->
          <div style="background:#F9FAFB;border:1px solid #E5E7EB;border-radius:8px;padding:14px 16px;margin:0 0 16px 0;">
            <p style="margin:0 0 8px 0;color:#374151;font-size:13px;font-weight:700;">📋 Détails de la demande</p>
            <p style="margin:0 0 4px 0;color:#4B5563;font-size:13px;">📅 Date souhaitée : {_fmt_date(original_date)} à {original_time}</p>
            {f'<p style="margin:0;color:#4B5563;font-size:13px;">📍 Adresse : {address}</p>' if address else ''}
          </div>

          {note_block}

          <p style="margin:20px 0 8px 0;color:#374151;font-size:15px;line-height:1.6;">
            <strong>Cette estimation vous convient-elle ?</strong><br>
            Répondez-moi par courriel ou téléphone pour confirmer le rendez-vous.
          </p>
        </td></tr>

        <!-- Contact card -->
        <tr><td style="padding:0 28px 24px 28px;">
          <div style="background:#F3F4F6;border-radius:12px;padding:16px;text-align:center;">
            <p style="margin:0 0 8px 0;color:#111827;font-size:15px;font-weight:700;">Lavage de Vitres Bois-Franc</p>
            <p style="margin:0 0 4px 0;">
              <a href="tel:+15145709802" style="color:#0891B2;text-decoration:none;font-size:15px;font-weight:600;">📞 514-570-9802</a>
            </p>
            <p style="margin:0;">
              <a href="https://Lavagedevitre.org" style="color:#0891B2;text-decoration:none;font-size:15px;font-weight:600;">🌐 Lavagedevitre.org</a>
            </p>
          </div>
        </td></tr>

        <tr><td style="padding:16px 28px;background:#0F172A;text-align:center;">
          <p style="margin:0;font-size:12px;color:#94A3B8;">Merci de votre confiance!</p>
        </td></tr>
      </table>
    </td></tr>
  </table>
</body></html>"""

            from_addr = os.environ.get("RESEND_FROM") or "onboarding@resend.dev"
            await asyncio.to_thread(
                resend.Emails.send,
                {
                    "from": from_addr,
                    "to": [client_email],
                    "reply_to": NOTIFY_EMAIL if NOTIFY_EMAIL else None,
                    "subject": f"Votre estimation — {price_str}",
                    "html": inject_branding(html),
                },
            )
            sent = True
            logger.info(f"Estimate email sent to client {client_email} — price {price_str}")
        except Exception as e:
            logger.error(f"Failed to send estimate email to {client_email}: {e}")

    updated = await db.appointment_requests.find_one({"id": request_id}, {"_id": 0})
    result = RequestResponse(**updated)
    # Return the email-send status in a side-channel for the frontend
    return result


@api_router.delete("/requests/{request_id}")
async def decline_request(request_id: str):
    """Decline a request — SOFT delete: marks as declined but keeps in DB so it can be recovered."""
    result = await db.appointment_requests.update_one(
        {"id": request_id},
        {"$set": {"status": "declined", "declined_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Request not found")
    # Tombstone (declined items are also kept out of prod_sync re-import)
    await prod_sync_module.add_tombstone(db, "request", request_id)
    return {"message": "Request declined (soft delete — can be restored)"}


@api_router.put("/requests/{request_id}/restore")
async def restore_request(request_id: str):
    """Restore a previously declined request → back to pending."""
    result = await db.appointment_requests.update_one(
        {"id": request_id},
        {"$set": {"status": "pending"}, "$unset": {"declined_at": ""}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Request not found")
    # Remove tombstone (restored items can sync again if needed)
    try:
        await db.tombstones.delete_one({"kind": "request", "id": request_id})
    except Exception:
        pass
    return {"message": "Request restored to pending"}


@api_router.delete("/requests/{request_id}/permanent")
async def delete_request_permanent(request_id: str):
    """Permanently delete a request (only available from archive)."""
    result = await db.appointment_requests.delete_one({"id": request_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Request not found")
    # Tombstone so prod_sync doesn't re-import the deleted request
    await prod_sync_module.add_tombstone(db, "request", request_id)
    return {"message": "Request permanently deleted"}

@api_router.get("/requests/count/pending")
async def get_pending_count():
    count = await db.appointment_requests.count_documents({"status": "pending"})
    return {"count": count}


# --- Statistics & Client History ---

@api_router.get("/stats")
async def get_stats():
    """Dashboard statistics"""
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")
    
    # Current month range
    month_start = now.strftime("%Y-%m-01")
    month_end = now.strftime("%Y-%m-31")
    
    # Total appointments
    total_appointments = await db.appointments.count_documents({})
    
    # This month appointments
    month_appointments = await db.appointments.count_documents({
        "date": {"$gte": month_start, "$lte": month_end}
    })
    
    # Today appointments
    today_appointments = await db.appointments.count_documents({"date": today})
    
    # Total revenue (all time)
    pipeline_revenue = [
        {"$group": {"_id": None, "total": {"$sum": "$price"}}}
    ]
    revenue_result = await db.appointments.aggregate(pipeline_revenue).to_list(1)
    total_revenue = revenue_result[0]["total"] if revenue_result else 0
    
    # Month revenue
    pipeline_month_rev = [
        {"$match": {"date": {"$gte": month_start, "$lte": month_end}}},
        {"$group": {"_id": None, "total": {"$sum": "$price"}}}
    ]
    month_rev_result = await db.appointments.aggregate(pipeline_month_rev).to_list(1)
    month_revenue = month_rev_result[0]["total"] if month_rev_result else 0
    
    # Pending requests
    pending_requests = await db.appointment_requests.count_documents({"status": "pending"})
    total_requests = await db.appointment_requests.count_documents({})
    accepted_requests = await db.appointment_requests.count_documents({"status": "accepted"})
    acceptance_rate = round((accepted_requests / total_requests * 100), 1) if total_requests > 0 else 0
    
    # Top clients (by appointment count)
    pipeline_clients = [
        {"$group": {"_id": "$client_name", "count": {"$sum": 1}, "total_spent": {"$sum": "$price"}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    top_clients = await db.appointments.aggregate(pipeline_clients).to_list(10)
    
    # Completed vs upcoming
    completed = await db.appointments.count_documents({"status": "completed"})
    upcoming = await db.appointments.count_documents({"status": "upcoming"})
    cancelled = await db.appointments.count_documents({"status": "cancelled"})
    
    return {
        "total_appointments": total_appointments,
        "month_appointments": month_appointments,
        "today_appointments": today_appointments,
        "total_revenue": total_revenue,
        "month_revenue": month_revenue,
        "pending_requests": pending_requests,
        "acceptance_rate": acceptance_rate,
        "completed": completed,
        "upcoming": upcoming,
        "cancelled": cancelled,
        "top_clients": [
            {"name": c["_id"], "count": c["count"], "total_spent": c.get("total_spent", 0)}
            for c in top_clients if c["_id"]
        ],
    }

@api_router.get("/clients")
async def get_clients():
    """List all unique clients with stats"""
    pipeline = [
        {"$group": {
            "_id": "$client_name",
            "count": {"$sum": 1},
            "total_spent": {"$sum": "$price"},
            "last_visit": {"$max": "$date"},
            "email": {"$first": "$client_email"},
            "phone": {"$first": "$client_phone"},
            "address": {"$first": "$client_address"},
        }},
        {"$sort": {"last_visit": -1}},
    ]
    clients = await db.appointments.aggregate(pipeline).to_list(500)
    return [
        {
            "name": c["_id"],
            "count": c["count"],
            "total_spent": c.get("total_spent", 0),
            "last_visit": c.get("last_visit", ""),
            "email": c.get("email", ""),
            "phone": c.get("phone", ""),
            "address": c.get("address", ""),
        }
        for c in clients if c["_id"]
    ]

@api_router.get("/clients/{client_name}/history")
async def get_client_history(client_name: str):
    """Get all appointments for a specific client"""
    appointments = await db.appointments.find(
        {"client_name": client_name}, {"_id": 0}
    ).sort("date", -1).to_list(500)
    return [AppointmentResponse(**a) for a in appointments]


# ============================================================
# CLIENTS DATABASE (Phase 1) — Dedicated clients collection
# ============================================================

class ClientCreate(BaseModel):
    name: str
    email: Optional[str] = ""
    phone: Optional[str] = ""
    address: Optional[str] = ""
    notes: Optional[str] = ""
    tags: Optional[List[str]] = []

class ClientUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None
    tags: Optional[List[str]] = None

class ClientResponse(BaseModel):
    id: str
    name: str
    email: str
    phone: str
    address: str
    notes: str
    tags: List[str]
    created_at: str
    updated_at: str

class ClientMatchRequest(BaseModel):
    email: Optional[str] = ""
    phone: Optional[str] = ""
    name: Optional[str] = ""


def _normalize_email(email: Optional[str]) -> str:
    return (email or "").strip().lower()


def _normalize_phone(phone: Optional[str]) -> str:
    """Keep only digits for matching"""
    import re
    return re.sub(r"\D", "", phone or "")


def _client_to_response(c: dict) -> dict:
    return {
        "id": c.get("id", ""),
        "name": c.get("name", ""),
        "email": c.get("email", ""),
        "phone": c.get("phone", ""),
        "address": c.get("address", ""),
        "notes": c.get("notes", ""),
        "tags": c.get("tags", []) or [],
        "created_at": c.get("created_at", ""),
        "updated_at": c.get("updated_at", ""),
    }


@api_router.post("/clients-db", response_model=ClientResponse)
async def create_client_db(data: ClientCreate):
    """Create a new stored client"""
    email_norm = _normalize_email(data.email)
    phone_norm = _normalize_phone(data.phone)

    # Check for duplicates by email (if present) or phone
    if email_norm:
        existing = await db.clients.find_one({"email_norm": email_norm})
        if existing:
            raise HTTPException(status_code=409, detail=f"Client avec ce courriel existe déjà: {existing.get('name')}")
    elif phone_norm:
        existing = await db.clients.find_one({"phone_norm": phone_norm})
        if existing:
            raise HTTPException(status_code=409, detail=f"Client avec ce téléphone existe déjà: {existing.get('name')}")

    now = datetime.now(timezone.utc).isoformat()
    client_doc = {
        "id": str(uuid.uuid4()),
        "name": (data.name or "").strip(),
        "email": (data.email or "").strip(),
        "email_norm": email_norm,
        "phone": (data.phone or "").strip(),
        "phone_norm": phone_norm,
        "address": (data.address or "").strip(),
        "notes": (data.notes or "").strip(),
        "tags": data.tags or [],
        "created_at": now,
        "updated_at": now,
    }
    await db.clients.insert_one(client_doc)
    return ClientResponse(**_client_to_response(client_doc))


@api_router.get("/clients-db", response_model=List[ClientResponse])
async def list_clients_db(search: Optional[str] = None, limit: int = 500, archived: bool = False):
    """List all stored clients with optional search by name/email/phone.
    By default only active (non-archived) clients are returned.
    Use ?archived=true to list archived (soft-deleted) clients."""
    query: dict = {}
    if archived:
        query["archived"] = True
    else:
        query["$or"] = [{"archived": {"$exists": False}}, {"archived": False}]
    if search:
        search_trim = search.strip()
        search_filter = {
            "$or": [
                {"name": {"$regex": search_trim, "$options": "i"}},
                {"email": {"$regex": search_trim, "$options": "i"}},
                {"phone": {"$regex": search_trim, "$options": "i"}},
                {"address": {"$regex": search_trim, "$options": "i"}},
            ]
        }
        # Combine with archive filter using $and
        query = {"$and": [query, search_filter]} if query else search_filter
    cursor = db.clients.find(query, {"_id": 0}).sort("name", 1).limit(max(1, min(limit, 2000)))
    items = await cursor.to_list(2000)
    return [ClientResponse(**_client_to_response(c)) for c in items]


# --- BULK ARCHIVE / RESTORE / DELETE endpoints ---------------------------

class ClientBulkIds(BaseModel):
    ids: List[str]


@api_router.post("/clients-db/archive-bulk")
async def archive_clients_bulk(payload: ClientBulkIds):
    """Soft-delete multiple clients — moves them to archive (archived=true)."""
    if not payload.ids:
        return {"archived": 0}
    now = datetime.now(timezone.utc).isoformat()
    res = await db.clients.update_many(
        {"id": {"$in": payload.ids}},
        {"$set": {"archived": True, "archived_at": now, "updated_at": now}},
    )
    return {"archived": res.modified_count}


@api_router.post("/clients-db/restore-bulk")
async def restore_clients_bulk(payload: ClientBulkIds):
    """Restore multiple archived clients back to active list."""
    if not payload.ids:
        return {"restored": 0}
    now = datetime.now(timezone.utc).isoformat()
    res = await db.clients.update_many(
        {"id": {"$in": payload.ids}},
        {"$set": {"archived": False, "updated_at": now}, "$unset": {"archived_at": ""}},
    )
    return {"restored": res.modified_count}


@api_router.post("/clients-db/delete-permanent-bulk")
async def delete_clients_permanent_bulk(payload: ClientBulkIds):
    """Permanently delete multiple clients from DB (cannot be undone)."""
    if not payload.ids:
        return {"deleted": 0}
    res = await db.clients.delete_many({"id": {"$in": payload.ids}})
    # Tombstone all deleted client IDs so prod_sync doesn't re-import them
    for cid in payload.ids:
        await prod_sync_module.add_tombstone(db, "client", cid)
    return {"deleted": res.deleted_count}


@api_router.get("/clients-db/archive/count")
async def count_archived_clients_db():
    """Count how many clients are currently in the archive."""
    count = await db.clients.count_documents({"archived": True})
    return {"count": count}


@api_router.get("/clients-db/count")
async def count_clients_db():
    """Total count of stored clients"""
    count = await db.clients.count_documents({})
    return {"count": count}


@api_router.post("/clients-db/match")
async def match_client_db(data: ClientMatchRequest):
    """Find an existing client by email, phone, or name. Priority: email > phone > name."""
    email_norm = _normalize_email(data.email)
    phone_norm = _normalize_phone(data.phone)

    if email_norm:
        c = await db.clients.find_one({"email_norm": email_norm}, {"_id": 0})
        if c:
            return {"matched": True, "by": "email", "client": _client_to_response(c)}

    if phone_norm:
        c = await db.clients.find_one({"phone_norm": phone_norm}, {"_id": 0})
        if c:
            return {"matched": True, "by": "phone", "client": _client_to_response(c)}

    if data.name:
        c = await db.clients.find_one(
            {"name": {"$regex": f"^{data.name.strip()}$", "$options": "i"}}, {"_id": 0}
        )
        if c:
            return {"matched": True, "by": "name", "client": _client_to_response(c)}

    return {"matched": False, "by": None, "client": None}


@api_router.get("/clients-db/{client_id}", response_model=ClientResponse)
async def get_client_db(client_id: str):
    """Get a single stored client by id"""
    c = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Client introuvable")
    return ClientResponse(**_client_to_response(c))


@api_router.put("/clients-db/{client_id}", response_model=ClientResponse)
async def update_client_db(client_id: str, data: ClientUpdate):
    """Update a stored client"""
    c = await db.clients.find_one({"id": client_id})
    if not c:
        raise HTTPException(status_code=404, detail="Client introuvable")

    update_fields = {}
    for field in ["name", "email", "phone", "address", "notes", "tags"]:
        val = getattr(data, field, None)
        if val is not None:
            update_fields[field] = val

    if "email" in update_fields:
        update_fields["email_norm"] = _normalize_email(update_fields["email"])
    if "phone" in update_fields:
        update_fields["phone_norm"] = _normalize_phone(update_fields["phone"])

    update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()

    await db.clients.update_one({"id": client_id}, {"$set": update_fields})
    updated = await db.clients.find_one({"id": client_id}, {"_id": 0})
    return ClientResponse(**_client_to_response(updated))


@api_router.delete("/clients-db/{client_id}")
async def delete_client_db(client_id: str):
    """Delete a stored client"""
    res = await db.clients.delete_one({"id": client_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Client introuvable")
    return {"message": "Client supprimé", "id": client_id}


@api_router.get("/clients-db/{client_id}/history")
async def get_client_db_history(client_id: str):
    """Get all appointments and requests linked to this stored client"""
    c = await db.clients.find_one({"id": client_id}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Client introuvable")

    # Primary match: direct client_id link (set by Phase 4 auto-linking)
    conditions = [{"client_id": client_id}]

    # Fallback matches (for legacy appointments created before auto-linking)
    if c.get("email"):
        conditions.append({"client_email": {"$regex": f"^{c['email']}$", "$options": "i"}})
    if c.get("phone"):
        phone_digits = c.get("phone_norm", "")
        if phone_digits and len(phone_digits) >= 7:
            conditions.append({"client_phone": {"$regex": phone_digits[-7:]}})
    if c.get("name"):
        conditions.append({"client_name": {"$regex": f"^{c['name']}$", "$options": "i"}})

    query = {"$or": conditions}
    appointments = await db.appointments.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    # Dedup in case multiple conditions match the same doc
    seen_ids = set()
    unique = []
    for a in appointments:
        if a.get("id") not in seen_ids:
            unique.append(a)
            seen_ids.add(a.get("id"))

    return {
        "client": _client_to_response(c),
        "appointments": [AppointmentResponse(**a) for a in unique],
        "total": len(unique),
    }

# ============================================================
# CLIENTS: Public Subscribe + CSV/XLSX Import
# ============================================================

class ClientSubscribe(BaseModel):
    name: str
    email: Optional[str] = ""
    phone: Optional[str] = ""
    address: Optional[str] = ""
    notes: Optional[str] = ""
    source: Optional[str] = "web"


@api_router.post("/clients-db/subscribe")
async def subscribe_client(data: ClientSubscribe):
    """Public endpoint used by the 'Add me to the client list' form on the website.
    Creates new client OR updates existing one if email/phone matches. Always returns 200."""
    email_norm = _normalize_email(data.email)
    phone_norm = _normalize_phone(data.phone)

    # Try to find existing
    existing = None
    if email_norm:
        existing = await db.clients.find_one({"email_norm": email_norm})
    if not existing and phone_norm:
        existing = await db.clients.find_one({"phone_norm": phone_norm})

    now = datetime.now(timezone.utc).isoformat()

    if existing:
        # Update non-empty fields only
        update_fields = {"updated_at": now}
        for src_k, dst_k in [("name", "name"), ("email", "email"), ("phone", "phone"),
                             ("address", "address"), ("notes", "notes")]:
            val = getattr(data, src_k, None)
            if val and val.strip():
                update_fields[dst_k] = val.strip()
        if email_norm:
            update_fields["email_norm"] = email_norm
        if phone_norm:
            update_fields["phone_norm"] = phone_norm

        # Append source tag
        tags = existing.get("tags", []) or []
        if data.source and data.source not in tags:
            tags.append(data.source)
            update_fields["tags"] = tags

        await db.clients.update_one({"id": existing["id"]}, {"$set": update_fields})
        updated = await db.clients.find_one({"id": existing["id"]}, {"_id": 0})
        logger.info(f"Subscribe: updated existing client {updated.get('name')} ({updated.get('email')})")
        return {"status": "updated", "client": _client_to_response(updated)}

    # Create new
    client_doc = {
        "id": str(uuid.uuid4()),
        "name": (data.name or "").strip(),
        "email": (data.email or "").strip(),
        "email_norm": email_norm,
        "phone": (data.phone or "").strip(),
        "phone_norm": phone_norm,
        "address": (data.address or "").strip(),
        "notes": (data.notes or "").strip(),
        "tags": [data.source] if data.source else [],
        "created_at": now,
        "updated_at": now,
    }
    await db.clients.insert_one(client_doc)
    logger.info(f"Subscribe: new client {client_doc['name']} ({client_doc['email']}) via {data.source}")
    return {"status": "created", "client": _client_to_response(client_doc)}


@api_router.post("/clients-db/import")
async def import_clients(file: UploadFile = File(...)):
    """Bulk import clients from CSV or XLSX. Auto-detects columns in FR or EN.
    Dedupes by email (preferred) or phone. Returns a detailed report."""
    import pandas as pd
    import io
    import re as _re

    try:
        content = await file.read()
        filename = (file.filename or "").lower()

        # Parse according to extension
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            df = pd.read_excel(io.BytesIO(content))
        else:
            # Try utf-8, fallback to latin-1
            try:
                df = pd.read_csv(io.BytesIO(content), encoding="utf-8")
            except Exception:
                df = pd.read_csv(io.BytesIO(content), encoding="latin-1")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Fichier illisible: {str(e)}")

    if df.empty:
        raise HTTPException(status_code=400, detail="Fichier vide")

    # Normalize column names
    def norm_col(c):
        return _re.sub(r"[^a-z0-9]", "", str(c).lower().strip())

    cols = {norm_col(c): c for c in df.columns}

    # Column mappings (priority order)
    col_name = next((cols[k] for k in ["nom", "name", "nomcomplet", "fullname", "clientname"] if k in cols), None)
    col_first = next((cols[k] for k in ["prenom", "firstname", "namefirst", "first"] if k in cols), None)
    col_last = next((cols[k] for k in ["nomfamille", "lastname", "namelast", "last", "surname"] if k in cols), None)
    col_email = next((cols[k] for k in ["courriel", "email", "mail", "courrier"] if k in cols), None)
    col_phone = next((cols[k] for k in ["telephone", "phone", "tel", "mobile", "cellulaire"] if k in cols), None)
    col_address = next((cols[k] for k in ["adresse", "address", "addr", "rue"] if k in cols), None)
    col_notes = next((cols[k] for k in ["notes", "note", "remarques", "comment"] if k in cols), None)

    if not (col_name or col_first or col_last or col_email):
        detected = list(cols.values())
        raise HTTPException(
            status_code=400,
            detail=f"Colonnes requises introuvables. Colonnes détectées: {detected}. "
                   f"Requis: au moins une colonne Nom, Prénom, Nom de famille, ou Courriel."
        )

    created = 0
    updated = 0
    skipped = 0
    errors: List[str] = []
    seen_emails = set()
    seen_phones = set()

    now = datetime.now(timezone.utc).isoformat()

    for idx, row in df.iterrows():
        try:
            # Build name
            if col_name:
                name = str(row[col_name]) if pd.notna(row[col_name]) else ""
            else:
                first = str(row[col_first]) if col_first and pd.notna(row[col_first]) else ""
                last = str(row[col_last]) if col_last and pd.notna(row[col_last]) else ""
                name = f"{first.strip()} {last.strip()}".strip()

            email = str(row[col_email]) if col_email and pd.notna(row[col_email]) else ""
            phone = str(row[col_phone]) if col_phone and pd.notna(row[col_phone]) else ""
            address = str(row[col_address]) if col_address and pd.notna(row[col_address]) else ""
            notes = str(row[col_notes]) if col_notes and pd.notna(row[col_notes]) else ""

            # Clean up "nan" strings from pandas
            def clean(s):
                if not s or str(s).strip().lower() in ("nan", "none", "null"):
                    return ""
                return str(s).strip()

            name = clean(name)
            email = clean(email)
            phone = clean(phone)
            address = clean(address)
            notes = clean(notes)

            if not name and not email and not phone:
                skipped += 1
                continue

            email_norm = _normalize_email(email)
            phone_norm = _normalize_phone(phone)

            # Dedup within import batch
            if email_norm and email_norm in seen_emails:
                skipped += 1
                continue
            if phone_norm and not email_norm and phone_norm in seen_phones:
                skipped += 1
                continue
            if email_norm:
                seen_emails.add(email_norm)
            if phone_norm:
                seen_phones.add(phone_norm)

            # Dedup against DB
            existing = None
            if email_norm:
                existing = await db.clients.find_one({"email_norm": email_norm})
            if not existing and phone_norm:
                existing = await db.clients.find_one({"phone_norm": phone_norm})

            if existing:
                # Merge non-empty fields into existing
                update_fields = {"updated_at": now}
                for field, val in [("name", name), ("email", email), ("phone", phone),
                                   ("address", address), ("notes", notes)]:
                    if val and not existing.get(field):
                        update_fields[field] = val
                if email_norm and not existing.get("email_norm"):
                    update_fields["email_norm"] = email_norm
                if phone_norm and not existing.get("phone_norm"):
                    update_fields["phone_norm"] = phone_norm

                tags = existing.get("tags", []) or []
                if "import" not in tags:
                    tags.append("import")
                    update_fields["tags"] = tags

                await db.clients.update_one({"id": existing["id"]}, {"$set": update_fields})
                updated += 1
            else:
                client_doc = {
                    "id": str(uuid.uuid4()),
                    "name": name,
                    "email": email,
                    "email_norm": email_norm,
                    "phone": phone,
                    "phone_norm": phone_norm,
                    "address": address,
                    "notes": notes,
                    "tags": ["import"],
                    "created_at": now,
                    "updated_at": now,
                }
                await db.clients.insert_one(client_doc)
                created += 1

        except Exception as e:
            errors.append(f"Ligne {idx + 2}: {str(e)}")

    total = len(df)
    logger.info(f"Import: {created} créés, {updated} mis à jour, {skipped} ignorés, {len(errors)} erreurs (sur {total})")

    return {
        "total_rows": total,
        "created": created,
        "updated": updated,
        "skipped_duplicates": skipped,
        "errors_count": len(errors),
        "errors": errors[:20],  # limit sample
        "message": f"Import terminé: {created} créés, {updated} mis à jour, {skipped} doublons ignorés."
    }


@api_router.get("/clients-db/export/csv")
async def export_clients_csv():
    """Export all clients as CSV"""
    from fastapi.responses import StreamingResponse
    import csv
    import io

    clients_list = await db.clients.find({}, {"_id": 0}).sort("name", 1).to_list(5000)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Nom", "Courriel", "Téléphone", "Adresse", "Notes", "Tags", "Date création"])
    for c in clients_list:
        writer.writerow([
            c.get("name", ""),
            c.get("email", ""),
            c.get("phone", ""),
            c.get("address", ""),
            c.get("notes", ""),
            ", ".join(c.get("tags", []) or []),
            c.get("created_at", ""),
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=clients_brightcalendar.csv"}
    )





@api_router.post("/requests/seed")



# --- Employees ---

class EmployeeCreate(BaseModel):
    name: str
    phone: Optional[str] = ""
    email: Optional[str] = ""
    color: Optional[str] = "#0891B2"

class EmployeeResponse(BaseModel):
    id: str
    name: str
    phone: str
    email: str
    color: str
    active: bool
    created_at: str

@api_router.post("/employees", response_model=EmployeeResponse)
async def create_employee(data: EmployeeCreate):
    emp = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "phone": data.phone or "",
        "email": data.email or "",
        "color": data.color or "#0891B2",
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.employees.insert_one(emp)
    return EmployeeResponse(**{k: v for k, v in emp.items() if k != "_id"})

@api_router.get("/employees", response_model=List[EmployeeResponse])
async def get_employees():
    emps = await db.employees.find({"active": True}, {"_id": 0}).sort("name", 1).to_list(100)
    return [EmployeeResponse(**e) for e in emps]

@api_router.delete("/employees/{employee_id}")
async def delete_employee(employee_id: str):
    await db.employees.update_one({"id": employee_id}, {"$set": {"active": False}})
    return {"message": "Employé désactivé"}

@api_router.put("/appointments/{appointment_id}/assign")
async def assign_employee(appointment_id: str, employee_id: str):
    """Assign an employee to an appointment (or pass 'none' to unassign)"""
    if employee_id == "none":
        result = await db.appointments.update_one(
            {"id": appointment_id},
            {"$unset": {"assigned_to": "", "assigned_id": "", "assigned_color": ""}}
        )
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="RDV non trouvé")
        appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
        return AppointmentResponse(**appt)
    emp = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employé non trouvé")
    result = await db.appointments.update_one(
        {"id": appointment_id},
        {"$set": {"assigned_to": emp["name"], "assigned_id": employee_id, "assigned_color": emp.get("color", "#0891B2")}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="RDV non trouvé")
    appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    return AppointmentResponse(**appt)

@api_router.get("/employees/{employee_id}/schedule")
async def get_employee_schedule(employee_id: str, date: Optional[str] = None):
    """Get appointments assigned to an employee"""
    query = {"assigned_id": employee_id}
    if date:
        query["date"] = date
    appts = await db.appointments.find(query, {"_id": 0}).sort("date", 1).to_list(500)
    return [AppointmentResponse(**a) for a in appts]


async def _build_backup_data():
    """Return all DB data as a plain dict (used by export endpoint + scheduler)."""
    data = {
        "exported_at": datetime.utcnow().isoformat() + "Z",
        "version": 1,
        "appointments": [],
        "requests": [],
        "employees": [],
        "reviews": [],
    }
    async for d in db.appointments.find({}, {"_id": 0}):
        data["appointments"].append(d)
    async for d in db.appointment_requests.find({}, {"_id": 0}):
        data["requests"].append(d)
    async for d in db.employees.find({}, {"_id": 0}):
        data["employees"].append(d)
    async for d in db.reviews.find({}, {"_id": 0}):
        data["reviews"].append(d)
    return data


@api_router.get("/backup/export")
async def export_backup():
    """Export all data as JSON for backup"""
    return await _build_backup_data()


async def _create_auto_backup():
    """Creates a backup document and keeps only last 2."""
    try:
        data = await _build_backup_data()
        backup_doc = {
            "id": str(uuid.uuid4()),
            "created_at": datetime.utcnow().isoformat() + "Z",
            "type": "auto",
            "data": data,
            "stats": {
                "appointments": len(data.get("appointments", [])),
                "requests": len(data.get("requests", [])),
                "employees": len(data.get("employees", [])),
                "reviews": len(data.get("reviews", [])),
            },
        }
        await db.backups.insert_one(backup_doc)
        # Keep only 2 most recent automatic backups
        cursor = db.backups.find({"type": "auto"}, {"_id": 1, "created_at": 1}).sort("created_at", -1)
        all_backups = [b async for b in cursor]
        for old in all_backups[2:]:
            await db.backups.delete_one({"_id": old["_id"]})
        logger.info(f"Auto backup created. Kept {min(len(all_backups), 2)} backups.")
    except Exception as e:
        logger.error(f"Auto backup failed: {e}")


@api_router.get("/backup/list")
async def list_backups():
    """List stored automatic backups."""
    cursor = db.backups.find({}, {"_id": 0, "data": 0}).sort("created_at", -1)
    return [b async for b in cursor]


@api_router.post("/backup/run-now")
async def run_backup_now():
    """Trigger a backup immediately (manual trigger)."""
    await _create_auto_backup()
    return {"status": "ok", "message": "Backup created"}


@api_router.get("/backup/download/{backup_id}")
async def download_backup(backup_id: str):
    """Download a specific stored backup as JSON."""
    b = await db.backups.find_one({"id": backup_id}, {"_id": 0})
    if not b:
        raise HTTPException(status_code=404, detail="Backup non trouvé")
    return b.get("data", {})


@api_router.post("/backup/import")
async def import_backup(body: dict):
    """Import/restore data from a backup JSON. Upserts based on id field."""
    results = {"appointments": 0, "requests": 0, "employees": 0, "reviews": 0}
    for coll_key, coll_name in [
        ("appointments", "appointments"),
        ("requests", "appointment_requests"),
        ("employees", "employees"),
        ("reviews", "reviews"),
    ]:
        items = body.get(coll_key, [])
        if not isinstance(items, list):
            continue
        for item in items:
            if not isinstance(item, dict):
                continue
            item_id = item.get("id")
            if not item_id:
                continue
            await db[coll_name].update_one({"id": item_id}, {"$set": item}, upsert=True)
            results[coll_key] += 1
    return {"status": "ok", "imported": results}


@api_router.get("/clients/emails")
async def get_client_emails():
    """Get unique client emails for email campaigns.
    Pulls from BOTH the Clients database (CRM) AND past appointments, merged by email."""
    seen = {}

    # 1) Primary source: Clients CRM database (includes CSV imports, manually added clients)
    try:
        clients_cursor = db.clients.find(
            {
                "email": {"$nin": [None, ""]},
                # Exclude archived/deleted clients
                "$and": [
                    {"$or": [{"archived": {"$ne": True}}, {"archived": {"$exists": False}}]},
                    {"$or": [{"status": {"$ne": "archived"}}, {"status": {"$exists": False}}]},
                ],
            },
            {"_id": 0, "email": 1, "name": 1, "phone": 1, "created_at": 1, "updated_at": 1},
        )
        async for doc in clients_cursor:
            email_raw = (doc.get("email") or "").strip()
            if not email_raw or "@" not in email_raw:
                continue
            key = email_raw.lower()
            if key in seen:
                continue
            seen[key] = {
                "email": email_raw,
                "name": doc.get("name", "") or "",
                "phone": doc.get("phone", "") or "",
                "last_visit": doc.get("updated_at") or doc.get("created_at") or "",
            }
    except Exception as e:
        # If clients collection doesn't exist or query fails, fall through to appointments
        print(f"[clients/emails] Clients CRM query failed: {e}")

    # 2) Secondary source: Appointments (for legacy clients not yet in CRM)
    cursor = db.appointments.find(
        {"client_email": {"$nin": [None, ""]}},
        {"_id": 0, "client_email": 1, "client_name": 1, "client_phone": 1, "date": 1},
    )
    async for doc in cursor:
        email_raw = (doc.get("client_email") or "").strip()
        if not email_raw or "@" not in email_raw:
            continue
        key = email_raw.lower()
        if key in seen:
            # Update last_visit if appointment date is more recent
            appt_date = doc.get("date", "")
            if appt_date and appt_date > str(seen[key].get("last_visit", "")):
                seen[key]["last_visit"] = appt_date
            continue
        seen[key] = {
            "email": email_raw,
            "name": doc.get("client_name", "") or "",
            "phone": doc.get("client_phone", "") or "",
            "last_visit": doc.get("date", "") or "",
        }

    # Sort by name for stable ordering
    result = sorted(seen.values(), key=lambda c: (c.get("name") or "").lower())
    return result


# --- Invoice PDF ---

@api_router.get("/invoice/{appointment_id}")
async def generate_invoice(appointment_id: str):
    """Generate printable invoice for an appointment"""
    from fastapi.responses import HTMLResponse
    appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not appt:
        # Fallback: allow lookup by id prefix (so short URLs like /api/invoice/a44f7a6e work)
        if len(appointment_id) >= 6:
            appt = await db.appointments.find_one({"id": {"$regex": f"^{appointment_id}"}}, {"_id": 0})
    if not appt:
        raise HTTPException(status_code=404, detail="Appointment not found")
    appointment_id = appt.get("id", appointment_id)

    invoice_num = appointment_id[:8].upper()
    price = appt.get('price', 0)
    # Invoice-only logo (different from email branding)
    logo_url = invoice_logo.INVOICE_LOGO_DATA_URL or branding.LOGO_DATA_URL or os.environ.get('INVOICE_LOGO_URL', '')

    # Seasonal promo: 10% automne (sept-nov)
    from datetime import datetime
    try:
        appt_date = datetime.strptime(appt.get('date', ''), '%Y-%m-%d')
    except Exception:
        appt_date = datetime.now()
    is_autumn = 9 <= appt_date.month <= 11
    promo_banner_html = ""
    if is_autumn:
        promo_banner_html = """
<div style="background:linear-gradient(135deg,#F59E0B 0%,#DC2626 100%);color:#FFF;padding:14px 20px;border-radius:12px;margin:16px 0;text-align:center;box-shadow:0 4px 12px rgba(245,158,11,0.3);-webkit-print-color-adjust:exact;print-color-adjust:exact;">
  <div style="font-size:11px;letter-spacing:2px;font-weight:700;opacity:0.95;">🍂 PROMO AUTOMNE 🍂</div>
  <div style="font-size:22px;font-weight:800;margin-top:4px;">-10% sur votre prochain lavage</div>
  <div style="font-size:11px;margin-top:6px;opacity:0.95;">Valide jusqu'au 30 novembre • Mentionnez ce code: AUTOMNE10</div>
</div>
"""

    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>Facture {invoice_num}</title>
<style>
@page{{size:letter;margin:0;}}
*{{box-sizing:border-box;}}
html,body{{margin:0;padding:0;}}
body{{font-family:-apple-system,'Segoe UI',Roboto,sans-serif;max-width:780px;margin:0 auto;padding:22px 28px;color:#1F2937;font-size:12px;background:#FFFFFF;}}
.top-banner{{background:linear-gradient(135deg,#0891B2 0%,#06B6D4 100%);height:6px;border-radius:3px;margin-bottom:14px;}}
.header{{display:flex;justify-content:space-between;align-items:center;gap:14px;margin-bottom:16px;padding-bottom:14px;border-bottom:2px solid #F3F4F6;}}
.brand{{display:flex;align-items:center;gap:14px;flex:1;min-width:0;}}
.brand img{{width:120px;height:auto;max-height:120px;border-radius:8px;object-fit:contain;background:#FFFFFF;flex-shrink:0;}}
.brand-info{{display:flex;flex-direction:column;gap:4px;min-width:0;}}
.company-contact{{font-size:11px;color:#374151;line-height:1.7;}}
.company-contact .row{{display:flex;align-items:center;gap:6px;}}
.company-contact .icon{{color:#0891B2;font-weight:700;flex-shrink:0;}}
.company-contact a{{color:#374151;text-decoration:none;word-break:break-word;}}
.invoice-block{{text-align:right;flex-shrink:0;min-width:140px;}}
.invoice-title{{font-size:26px;font-weight:800;color:#0891B2;letter-spacing:-1px;line-height:1;}}
.invoice-meta{{font-size:11px;color:#6B7280;margin-top:6px;line-height:1.5;}}
.invoice-meta strong{{color:#111827;font-weight:600;}}
@media (max-width:480px){{
  .header{{flex-wrap:wrap;gap:10px;}}
  .brand{{flex:1 1 100%;}}
  .brand img{{width:90px;max-height:90px;}}
  .invoice-block{{flex:1 1 100%;text-align:left;border-top:1px solid #F3F4F6;padding-top:8px;}}
  .invoice-title{{font-size:22px;}}
}}
.greeting{{font-size:12px;color:#374151;margin-bottom:12px;line-height:1.5;}}
.greeting .name{{font-weight:700;color:#0891B2;}}
.card{{background:#F9FAFB;border-radius:10px;padding:12px 16px;margin-bottom:12px;border-left:4px solid #0891B2;}}
.card-title{{font-size:10px;font-weight:700;color:#6B7280;text-transform:uppercase;letter-spacing:1px;margin-bottom:8px;}}
.client-grid{{display:grid;grid-template-columns:auto 1fr;gap:4px 16px;}}
.client-grid .lbl{{color:#9CA3AF;font-size:11px;}}
.client-grid .val{{color:#111827;font-weight:500;font-size:12px;}}
.service-table{{width:100%;border-collapse:collapse;border-radius:10px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,0.04);margin-bottom:12px;}}
.service-table thead th{{background:#0891B2;color:#FFF;padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:1px;}}
.service-table thead th:last-child{{text-align:right;}}
.service-table tbody td{{padding:12px 14px;background:#FFFFFF;border-bottom:1px solid #F3F4F6;font-size:12px;}}
.service-table tbody td:last-child{{text-align:right;font-weight:600;}}
.total-row td{{background:#F0F9FF!important;padding:14px!important;font-size:17px!important;font-weight:800;color:#0891B2;border-bottom:none!important;}}
.notes-card{{background:#FFFBEB;border-left-color:#F59E0B;padding:10px 16px;}}
.notes-card .card-title{{color:#92400E;}}
.notes-card p{{color:#78350F;line-height:1.4;margin:0;font-size:11px;}}
.thankyou{{text-align:center;padding:14px 16px;margin-top:12px;background:#F0F9FF;border-radius:12px;}}
.thankyou h3{{color:#0891B2;font-size:14px;margin:0 0 4px 0;font-weight:700;}}
.thankyou p{{color:#374151;font-size:11px;margin:0;line-height:1.4;}}
.footer{{margin-top:14px;padding-top:10px;border-top:1px solid #F3F4F6;text-align:center;color:#9CA3AF;font-size:10px;line-height:1.5;}}
.footer strong{{color:#0891B2;}}
@media print{{
  body{{padding:20px 28px;-webkit-print-color-adjust:exact;print-color-adjust:exact;}}
  .top-banner,.service-table thead th,.total-row td,.card,.thankyou{{-webkit-print-color-adjust:exact;print-color-adjust:exact;}}
  .page{{page-break-after:avoid;page-break-inside:avoid;}}
  .actions-bar{{display:none!important;}}
}}
.actions-bar{{position:sticky;top:0;left:0;right:0;display:flex;justify-content:space-between;align-items:center;gap:8px;padding:10px 14px;background:rgba(255,255,255,0.95);backdrop-filter:blur(8px);-webkit-backdrop-filter:blur(8px);border-bottom:1px solid #E5E7EB;margin:-22px -28px 14px -28px;z-index:10;}}
.btn-back,.btn-print{{display:inline-flex;align-items:center;gap:6px;padding:9px 14px;border-radius:8px;font-size:13px;font-weight:600;border:none;cursor:pointer;font-family:inherit;}}
.btn-back{{background:#F3F4F6;color:#111827;}}
.btn-back:active{{background:#E5E7EB;}}
.btn-print{{background:#0891B2;color:#FFF;}}
.btn-print:active{{background:#0E7490;}}
</style></head><body>

<div class="page">
<div class="actions-bar">
  <button class="btn-back" onclick="(function(){{ if(window.history.length>1){{ window.history.back(); }} else if(window.opener){{ window.close(); }} else {{ window.location.href='/'; }} }})()">← Retour</button>
  <button class="btn-print" onclick="window.print()">🖨️ Imprimer / PDF</button>
</div>
<div class="top-banner"></div>

<div class="header">
  <div class="brand">
    <img src="{logo_url}" alt="Logo" />
    <div class="brand-info">
      <div class="company-contact">
        <div class="row"><span class="icon">☎</span><span>Tel.: 514-570-9802</span></div>
        <div class="row"><span class="icon">✉</span><a href="mailto:lavagedevitreboisfranc@live.com">lavagedevitreboisfranc@live.com</a></div>
        <div class="row"><span class="icon">🌐</span><a href="https://lavagedevitre.org">Lavagedevitre.org</a></div>
      </div>
    </div>
  </div>
  <div class="invoice-block">
    <div class="invoice-title">FACTURE</div>
    <div class="invoice-meta">
      <strong>N°</strong> {invoice_num}<br>
      <strong>Date:</strong> {appt.get('date','')}
    </div>
  </div>
</div>

<div class="greeting">
  Bonjour <span class="name">{appt.get('client_name','')}</span>, voici le détail de votre service. Merci de nous faire confiance! ✨
</div>

<div class="card">
  <div class="card-title">👤 Informations client</div>
  <div class="client-grid">
    <span class="lbl">Nom</span><span class="val">{appt.get('client_name','')}</span>
    <span class="lbl">Courriel</span><span class="val">{appt.get('client_email','') or '—'}</span>
    <span class="lbl">Téléphone</span><span class="val">{appt.get('client_phone','') or '—'}</span>
    <span class="lbl">Adresse</span><span class="val">{appt.get('client_address','') or '—'}</span>
  </div>
</div>

<table class="service-table">
  <thead><tr><th>Description du service</th><th>Date</th><th>Prix</th></tr></thead>
  <tbody>
    <tr>
      <td><strong>{appt.get('title','Service')}</strong></td>
      <td>{appt.get('date','')}</td>
      <td>{price:.2f} $</td>
    </tr>
    <tr class="total-row">
      <td colspan="2" style="text-align:right;">TOTAL À PAYER</td>
      <td>{price:.2f} $</td>
    </tr>
  </tbody>
</table>

{"<div class='notes-card card'><div class='card-title'>📝 Notes</div><p>" + appt.get('notes','') + "</p></div>" if appt.get('notes') else ""}

{promo_banner_html}

<div class="thankyou">
  <h3>Merci pour votre confiance! 💙</h3>
  <p>Nous espérons vous revoir bientôt. N'hésitez pas à nous contacter pour toute question.</p>
</div>

<div class="footer">
  <strong>Lavage de Vitres Bois-Franc</strong> · 514-570-9802 · Lavagedevitre.org<br>
  Service de lavage de vitres résidentiel et commercial
</div>
</div>

<script>window.onload=function(){{window.print();}}</script>
</body></html>"""
    return HTMLResponse(content=html)


# --- iCalendar (.ics) feed for Apple/Google/Outlook calendar subscription ---

@api_router.get("/calendar/{token}.ics")
async def calendar_ics_feed(token: str):
    """Public iCalendar (.ics) feed of all upcoming/recent appointments.

    The user adds this URL to Apple Calendar (Settings > Calendar > Accounts >
    Add Account > Other > Add Subscribed Calendar) and Apple Calendar will
    auto-refresh every ~15 min, showing all RDV on iPhone / Mac / Apple Watch.

    Token must match CALENDAR_TOKEN env var to prevent random scraping.
    """
    from fastapi.responses import Response
    expected = os.environ.get("CALENDAR_TOKEN", "").strip()
    if not expected or token != expected:
        raise HTTPException(status_code=404, detail="Not found")
    # Pull all non-archived appointments
    appts: list[dict] = []
    async for a in db.appointments.find({"status": {"$ne": "archived"}}, {"_id": 0}):
        appts.append(a)

    # Geocode all unique addresses (cached) so Apple Calendar gets tappable Maps links
    geocode_map: dict[str, tuple[float, float]] = {}
    unique_addrs = set()
    for a in appts:
        addr = (a.get("client_address") or "").strip()
        if addr:
            unique_addrs.add(addr)
    # Only geocode up to 30 addresses per request to avoid blocking too long;
    # the cache fills up over multiple refreshes (Apple polls every 15 min)
    for addr in list(unique_addrs)[:30]:
        try:
            coords = await geocoder_module.geocode(db, addr)
            if coords:
                geocode_map[addr] = coords
        except Exception:
            pass

    body = calendar_feed_module.build_ics_feed(appts, geocode_map=geocode_map)
    return Response(
        content=body,
        media_type="text/calendar; charset=utf-8",
        headers={
            "Content-Disposition": 'inline; filename="gexia360.ics"',
            "Cache-Control": "no-cache, must-revalidate",
        },
    )


# === URL Shortener (for clean SMS/email links) ===

import secrets as _secrets
import string as _string

_SHORT_ALPHABET = _string.ascii_letters + _string.digits


def _gen_short_code(length: int = 6) -> str:
    return ''.join(_secrets.choice(_SHORT_ALPHABET) for _ in range(length))


async def _make_short_url(target_path: str, base_url: str | None = None, request: "Request | None" = None) -> str:
    """Create or reuse a short URL like https://host/api/s/abc123 -> target_path.

    If a Request is provided, use its host headers (x-forwarded-host / host)
    instead of APP_URL — this guarantees the URL always matches the host the
    user is currently on (preview vs production vs custom domain).
    """
    if not base_url:
        if request is not None:
            try:
                host = request.headers.get("x-forwarded-host") or request.headers.get("host") or ""
                proto = request.headers.get("x-forwarded-proto") or "https"
                host = (host or "").split(",")[0].strip()
                if host:
                    base_url = f"{proto}://{host}".rstrip("/")
            except Exception:
                base_url = None
        if not base_url:
            base_url = (os.environ.get("APP_URL") or "").rstrip("/")
    # Try up to 5 times to avoid collisions
    for _ in range(5):
        code = _gen_short_code(6)
        existing = await db.short_links.find_one({"code": code})
        if existing:
            continue
        await db.short_links.insert_one({
            "code": code,
            "target": target_path,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        return f"{base_url}/api/s/{code}"
    # Fallback: long URL
    return f"{base_url}{target_path}"


@app.get("/api/s/{code}")
async def short_link_resolver(code: str, request: Request):
    """Redirect a 6-char short code to its target URL.

    Builds redirect using the request host (so works on any preview/production URL).
    """
    from fastapi.responses import RedirectResponse, HTMLResponse
    row = await db.short_links.find_one({"code": code})
    if not row or not row.get("target"):
        return HTMLResponse("<h1>Lien expiré ou invalide</h1>", status_code=404)
    target = row["target"]
    if target.startswith("http"):
        return RedirectResponse(url=target, status_code=302)
    # Build base from request headers (matches caller's host)
    try:
        host = request.headers.get("x-forwarded-host") or request.headers.get("host") or ""
        proto = request.headers.get("x-forwarded-proto") or "https"
        host = (host or "").split(",")[0].strip()
        base_url = f"{proto}://{host}".rstrip("/") if host else (os.environ.get("APP_URL") or "").rstrip("/")
    except Exception:
        base_url = (os.environ.get("APP_URL") or "").rstrip("/")
    return RedirectResponse(url=f"{base_url}{target}", status_code=302)



@api_router.post("/appointments/{appointment_id}/accept-alternative")
async def accept_client_alternative(appointment_id: str):
    """Accept the client's proposed alternative date/time.

    Reads client_suggested_date + client_suggested_time from the appointment,
    moves the appointment to that slot, marks client_confirmed=true, and clears
    client_requested_alternative. Returns the full updated appointment.
    """
    appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not appt:
        raise HTTPException(status_code=404, detail="Appointment not found")

    new_date = appt.get("client_suggested_date")
    new_time = appt.get("client_suggested_time")
    if not new_date or not new_time:
        raise HTTPException(status_code=400, detail="No client suggestion to accept")

    now_iso = datetime.now(timezone.utc).isoformat()
    await db.appointments.update_one(
        {"id": appointment_id},
        {"$set": {
            "date": new_date,
            "time_slot": new_time[:5],
            "client_confirmed": True,
            "client_confirmed_at": now_iso,
            "client_requested_alternative": False,
            # keep client_suggested_* for history; UI hides the card after accept
        }},
    )
    fresh = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    return fresh



class ShortenReq(BaseModel):
    target: str


@api_router.post("/shorten")
async def shorten_url(payload: ShortenReq, request: Request):
    """Create a 6-char short URL for any internal /api/... path or full URL.
    Used by the frontend to make SMS/email links compact and clickable.
    Uses the actual request host so URLs always match what the user sees.
    """
    target = (payload.target or "").strip()
    if not target:
        raise HTTPException(status_code=400, detail="target is required")
    short_url = await _make_short_url(target, request=request)
    return {"short_url": short_url}


# === Client appointment confirmation flow ===

@api_router.post("/appointments/{appointment_id}/send-client-confirmation")
async def send_client_confirmation(appointment_id: str, request: Request):
    """Send a confirmation email to the CLIENT with a single short link to a
    customer portal showing 3 buttons: Réservé / Modifier / Facture.

    Returns a ready-made SMS body the owner can use to send via iMessage too.
    """
    appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not appt:
        raise HTTPException(status_code=404, detail="Appointment not found")

    name = appt.get("client_name") or "Client"
    email = (appt.get("client_email") or "").strip()
    date = appt.get("date") or ""
    time_slot = appt.get("time_slot") or ""
    duration = appt.get("duration_minutes") or 60
    addr = appt.get("client_address") or ""
    price = appt.get("price")

    # Format the date nicely in French
    try:
        y, m, d = [int(x) for x in date.split('-')]
        months = ['janvier', 'février', 'mars', 'avril', 'mai', 'juin',
                  'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre']
        date_pretty = f"{d} {months[m-1]} {y}"
    except Exception:
        date_pretty = date

    # ONE short URL → portal page with 3 action buttons (Réservé / Modifier / Facture)
    portal_target = f"/api/appointments/{appointment_id}/client-confirm"
    portal_url = await _make_short_url(portal_target, request=request)

    sms_body = (
        f"Bonjour {name},\n\n"
        f"Votre rendez-vous Lavage de Vitres Bois-Franc:\n"
        f"📅 {date_pretty} à {time_slot}\n"
        f"📍 {addr}\n\n"
        f"👉 Voir / Confirmer / Modifier / Facture:\n{portal_url}\n\n"
        f"Merci!"
    )

    sent_email = False
    if email and resend.api_key:
        html = f"""
<div style="font-family:Arial,sans-serif;max-width:560px;margin:0 auto;color:#111;">
  <div style="background:#0B5394;color:#FFFFFF;padding:18px;border-radius:8px 8px 0 0;text-align:center;">
    <h2 style="margin:0;font-size:18px;">📅 Confirmation de rendez-vous</h2>
  </div>
  <div style="background:#FFFFFF;border:1px solid #E5E7EB;border-top:none;padding:22px;border-radius:0 0 8px 8px;">
    <p>Bonjour <strong>{name}</strong>,</p>
    <p>Nous avons planifié votre rendez-vous pour le service de lavage de vitres :</p>
    <div style="background:#F1F5F9;border-left:4px solid #0B5394;padding:12px 14px;border-radius:6px;margin:12px 0;">
      <p style="margin:0 0 4px;"><strong>📅 Date :</strong> {date_pretty}</p>
      <p style="margin:0 0 4px;"><strong>🕐 Heure :</strong> {time_slot} ({duration} min)</p>
      <p style="margin:0 0 4px;"><strong>📍 Adresse :</strong> {addr}</p>
      {f'<p style="margin:0;"><strong>💰 Prix estimé :</strong> {float(price):.2f} $</p>' if price else ''}
    </div>
    <p style="text-align:center;margin:18px 0;">
      <a href="{portal_url}" style="display:inline-block;padding:14px 28px;background:#0B5394;color:#FFFFFF;text-decoration:none;border-radius:8px;font-weight:700;font-size:16px;">📋 Voir mon rendez-vous</a>
    </p>
    <p style="font-size:13px;color:#64748B;text-align:center;">Confirmez, modifiez ou consultez votre facture en un seul clic.</p>
    <p style="font-size:12px;color:#94A3B8;text-align:center;margin-top:16px;">
      Si le bouton ne fonctionne pas, copiez ce lien :<br>
      <span style="word-break:break-all;">{portal_url}</span>
    </p>
  </div>
  <p style="text-align:center;color:#9CA3AF;font-size:11px;margin-top:8px;">Lavage de Vitres Bois-Franc &mdash; Gexia360</p>
</div>
"""
        try:
            resend.Emails.send({
                "from": os.environ.get("RESEND_FROM") or "onboarding@resend.dev",
                "to": [email],
                "subject": f"Confirmation de rendez-vous — {date_pretty}",
                "html": html,
            })
            sent_email = True
        except Exception as e:
            logger.warning(f"send_client_confirmation email failed: {e}")

    return {
        "email_sent": sent_email,
        "client_email": email,
        "client_phone": appt.get("client_phone") or "",
        "sms_body": sms_body,
        "portal_url": portal_url,
        # Backward-compat aliases (old frontends may still read these)
        "confirm_url": portal_url,
        "alternative_url": portal_url,
    }


@app.get("/api/appointments/{appointment_id}/client-confirm")
async def client_confirm_page(appointment_id: str, request: Request, action: str = ""):
    """Public landing page where the client sees their appointment with 3 buttons:
    Réservé (confirm) / Modifier (alternative) / Facture (invoice).

    If `?action=confirm` or `?action=alternative` is set, performs the action and
    shows a success page. Without action → shows the portal with 3 buttons.
    """
    from fastapi.responses import HTMLResponse, RedirectResponse
    appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not appt:
        return HTMLResponse(
            "<h1 style='font-family:sans-serif;text-align:center;margin-top:80px;'>Rendez-vous introuvable</h1>"
            "<p style='text-align:center;'>Veuillez contacter Lavage de Vitres Bois-Franc.</p>",
            status_code=404,
        )

    name = appt.get("client_name") or "Client"
    date = appt.get("date") or ""
    time_slot = appt.get("time_slot") or ""
    duration = appt.get("duration_minutes") or 60
    addr = appt.get("client_address") or ""
    price = appt.get("price") or 0
    status = appt.get("status") or "upcoming"
    is_paid = (status == "paid")
    confirmed = bool(appt.get("client_confirmed"))

    # Pretty French date
    try:
        y, m, d = [int(x) for x in date.split('-')]
        months = ['janvier', 'février', 'mars', 'avril', 'mai', 'juin',
                  'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre']
        date_pretty = f"{d} {months[m-1]} {y}"
    except Exception:
        date_pretty = date

    # Build URLs based on the actual host the client is on (NEVER from APP_URL env var)
    try:
        host = request.headers.get("x-forwarded-host") or request.headers.get("host") or ""
        proto = request.headers.get("x-forwarded-proto") or "https"
        host = (host or "").split(",")[0].strip()
        host_base = f"{proto}://{host}".rstrip("/") if host else (os.environ.get("APP_URL") or "").rstrip("/")
    except Exception:
        host_base = (os.environ.get("APP_URL") or "").rstrip("/")

    confirm_action_url = f"{host_base}/api/appointments/{appointment_id}/client-confirm?action=confirm"
    alternative_action_url = f"{host_base}/api/appointments/{appointment_id}/client-confirm?action=alternative"
    invoice_url = f"{host_base}/api/invoice/{appointment_id}"

    # === ACTION = CONFIRM ===
    if action == "confirm":
        await db.appointments.update_one(
            {"id": appointment_id},
            {"$set": {"client_confirmed": True, "client_confirmed_at": datetime.now(timezone.utc).isoformat()}},
        )
        return HTMLResponse(_render_status_page(
            title="✅ Rendez-vous confirmé!",
            color="#10B981",
            message=f"Merci {name}! Votre rendez-vous du <strong>{date_pretty} à {time_slot}</strong> est confirmé.",
            back_url=f"{host_base}/api/appointments/{appointment_id}/client-confirm",
        ))

    # === ACTION = ALTERNATIVE — show form to let client pick a new date/time ===
    if action == "alternative":
        # Pre-fill with client's previous suggestion (if any)
        prev_date = appt.get("client_suggested_date") or ""
        prev_time = appt.get("client_suggested_time") or ""
        prev_note = appt.get("client_suggested_note") or ""
        submit_url = f"{host_base}/api/appointments/{appointment_id}/client-suggest-alternative"
        back_url = f"{host_base}/api/appointments/{appointment_id}/client-confirm"

        form_html = f"""<!doctype html>
<html lang="fr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Proposer un autre moment</title>
<style>
*{{box-sizing:border-box;}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;background:#F1F5F9;margin:0;padding:0;min-height:100vh;color:#0F172A;}}
.wrap{{max-width:480px;margin:0 auto;padding:18px;}}
.brand{{text-align:center;padding:14px 0 18px;color:#475569;font-size:13px;font-weight:700;letter-spacing:0.5px;text-transform:uppercase;}}
.card{{background:#FFFFFF;border-radius:16px;padding:22px;box-shadow:0 4px 14px rgba(15,23,42,0.06);margin-bottom:14px;}}
h1{{font-size:22px;margin:0 0 6px;color:#F59E0B;}}
.sub{{color:#64748B;font-size:14px;margin:0 0 16px;line-height:1.5;}}
.lbl{{display:block;font-size:11px;font-weight:700;color:#64748B;letter-spacing:0.5px;margin:14px 0 6px;text-transform:uppercase;}}
.fld{{width:100%;padding:14px;border:1px solid #CBD5E1;border-radius:10px;font-size:16px;background:#FFFFFF;color:#0F172A;font-family:inherit;}}
.fld:focus{{outline:none;border-color:#F59E0B;box-shadow:0 0 0 3px rgba(245,158,11,0.15);}}
textarea.fld{{min-height:80px;resize:vertical;}}
.btns{{display:flex;flex-direction:column;gap:10px;margin-top:18px;}}
.btn{{display:block;padding:16px;border-radius:12px;text-decoration:none;font-weight:700;font-size:16px;text-align:center;border:none;cursor:pointer;font-family:inherit;}}
.btn-primary{{background:#F59E0B;color:#FFFFFF;}}
.btn-secondary{{background:#FFFFFF;color:#475569;border:1px solid #CBD5E1;}}
.current{{background:#FEF3C7;border:1px solid #FCD34D;border-radius:10px;padding:12px;margin-bottom:14px;font-size:13px;color:#78350F;}}
.brand-foot{{text-align:center;color:#94A3B8;font-size:11px;margin-top:14px;}}
</style></head>
<body>
<div class="wrap">
  <div class="brand">Lavage de Vitres Bois-Franc</div>
  <div class="card">
    <h1>🔄 Proposer un autre moment</h1>
    <p class="sub">Bonjour {name}, sélectionnez votre date et heure préférées. Nous confirmerons rapidement.</p>
    <div class="current">RDV actuel : <strong>{date_pretty} à {time_slot}</strong></div>
    <form method="POST" action="{submit_url}">
      <label class="lbl" for="alt_date">Nouvelle date</label>
      <input class="fld" type="date" id="alt_date" name="date" value="{prev_date or date}" required>

      <label class="lbl" for="alt_time">Nouvelle heure</label>
      <input class="fld" type="time" id="alt_time" name="time" value="{prev_time or time_slot}" required>

      <label class="lbl" for="alt_note">Note (optionnel)</label>
      <textarea class="fld" id="alt_note" name="note" placeholder="Ex: Préférablement en après-midi...">{prev_note}</textarea>

      <div class="btns">
        <button type="submit" class="btn btn-primary">📤 Envoyer ma proposition</button>
        <a href="{back_url}" class="btn btn-secondary">← Annuler</a>
      </div>
    </form>
  </div>
  <div class="brand-foot">Lavage de Vitres Bois-Franc &mdash; Gexia360</div>
</div>
</body></html>"""
        return HTMLResponse(content=form_html)

    # === DEFAULT — Portal page with 3 buttons (Réservé / Modifier / Facture) ===
    confirmed_badge = (
        '<div style="display:inline-block;padding:6px 12px;background:#D1FAE5;color:#065F46;border-radius:99px;font-size:13px;font-weight:700;margin-bottom:14px;">✓ Déjà confirmé</div>'
        if confirmed else ""
    )
    paid_badge = (
        '<div style="display:inline-block;padding:6px 12px;background:#CFFAFE;color:#155E75;border-radius:99px;font-size:13px;font-weight:700;margin-bottom:14px;">💰 Payé</div>'
        if is_paid else ""
    )
    price_row = (
        f'<div style="display:flex;justify-content:space-between;padding:10px 0;border-top:1px solid #E5E7EB;"><span style="color:#64748B;">Prix</span><strong>{float(price):.2f} $</strong></div>'
        if price and float(price) > 0 else ""
    )

    html = f"""<!doctype html>
<html lang="fr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Mon rendez-vous — Lavage de Vitres Bois-Franc</title>
<style>
*{{box-sizing:border-box;}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;background:#F1F5F9;margin:0;padding:0;min-height:100vh;color:#0F172A;}}
.wrap{{max-width:480px;margin:0 auto;padding:18px;}}
.brand{{text-align:center;padding:14px 0 18px;color:#475569;font-size:13px;font-weight:700;letter-spacing:0.5px;text-transform:uppercase;}}
.card{{background:#FFFFFF;border-radius:16px;padding:22px;box-shadow:0 4px 14px rgba(15,23,42,0.06);margin-bottom:14px;}}
h1{{font-size:22px;margin:0 0 6px;color:#0F172A;}}
.sub{{color:#64748B;font-size:14px;margin:0 0 16px;}}
.detail-row{{display:flex;justify-content:space-between;padding:10px 0;border-top:1px solid #E5E7EB;}}
.detail-row:first-of-type{{border-top:none;}}
.detail-row .label{{color:#64748B;}}
.detail-row strong{{color:#0F172A;}}
.btns{{display:flex;flex-direction:column;gap:10px;margin-top:6px;}}
.btn{{display:block;padding:16px;border-radius:12px;text-decoration:none;font-weight:700;font-size:16px;text-align:center;transition:transform .08s ease;}}
.btn:active{{transform:scale(0.98);}}
.btn-primary{{background:#10B981;color:#FFFFFF;}}
.btn-warn{{background:#F59E0B;color:#FFFFFF;}}
.btn-info{{background:#0B5394;color:#FFFFFF;}}
.brand-foot{{text-align:center;color:#94A3B8;font-size:11px;margin-top:14px;}}
</style></head>
<body>
<div class="wrap">
  <div class="brand">Lavage de Vitres Bois-Franc</div>
  <div class="card">
    {confirmed_badge}{paid_badge}
    <h1>Bonjour {name}!</h1>
    <p class="sub">Voici les détails de votre rendez-vous.</p>
    <div class="detail-row"><span class="label">Date</span><strong>{date_pretty}</strong></div>
    <div class="detail-row"><span class="label">Heure</span><strong>{time_slot}</strong></div>
    <div class="detail-row"><span class="label">Durée</span><strong>{duration} min</strong></div>
    <div class="detail-row"><span class="label">Adresse</span><strong style="text-align:right;">{addr}</strong></div>
    {price_row}
  </div>

  <div class="card">
    <div class="btns">
      <a href="{confirm_action_url}" class="btn btn-primary">✅ Réservé (je confirme)</a>
      <a href="{alternative_action_url}" class="btn btn-warn">🔄 Modifier (autre moment)</a>
    </div>
  </div>

  <div class="brand-foot">Lavage de Vitres Bois-Franc &mdash; Gexia360</div>
</div>
</body></html>"""
    return HTMLResponse(content=html)


@app.post("/api/appointments/{appointment_id}/client-suggest-alternative")
async def client_suggest_alternative(appointment_id: str, request: Request):
    """Receive the alternative date/time the CLIENT proposes from the public portal.

    Stores client_suggested_date / client_suggested_time / client_suggested_note +
    client_requested_alternative=true on the appointment, then renders a thank-you page.
    Also notifies the owner via email (if Resend configured + NOTIFY_EMAIL set).
    """
    from fastapi.responses import HTMLResponse
    appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not appt:
        return HTMLResponse(
            "<h1 style='font-family:sans-serif;text-align:center;margin-top:80px;'>Rendez-vous introuvable</h1>",
            status_code=404,
        )
    try:
        form = await request.form()
    except Exception:
        form = {}
    new_date = (form.get("date") or "").strip()
    new_time = (form.get("time") or "").strip()
    note = (form.get("note") or "").strip()

    # Basic validation: YYYY-MM-DD + HH:MM
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", new_date) or not re.match(r"^\d{2}:\d{2}$", new_time):
        return HTMLResponse(
            "<h1 style='font-family:sans-serif;text-align:center;margin-top:80px;color:#DC2626;'>Date ou heure invalide</h1>"
            "<p style='text-align:center;'>Veuillez retourner et réessayer.</p>",
            status_code=400,
        )

    now_iso = datetime.now(timezone.utc).isoformat()
    await db.appointments.update_one(
        {"id": appointment_id},
        {"$set": {
            "client_requested_alternative": True,
            "client_alt_requested_at": now_iso,
            "client_suggested_date": new_date,
            "client_suggested_time": new_time,
            "client_suggested_note": note,
        }},
    )

    # Build the host base for the back link
    try:
        host = request.headers.get("x-forwarded-host") or request.headers.get("host") or ""
        proto = request.headers.get("x-forwarded-proto") or "https"
        host = (host or "").split(",")[0].strip()
        host_base = f"{proto}://{host}".rstrip("/") if host else (os.environ.get("APP_URL") or "").rstrip("/")
    except Exception:
        host_base = (os.environ.get("APP_URL") or "").rstrip("/")

    # Pretty French date
    try:
        y, m, d = [int(x) for x in new_date.split('-')]
        months = ['janvier', 'février', 'mars', 'avril', 'mai', 'juin',
                  'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre']
        date_pretty = f"{d} {months[m-1]} {y}"
    except Exception:
        date_pretty = new_date

    # === Notify the owner via email ===
    owner_email = (NOTIFY_EMAIL or "").strip()
    if resend.api_key and owner_email:
        try:
            client_name = appt.get("client_name") or "Client"
            client_phone = appt.get("client_phone") or ""
            old_date = appt.get("date") or ""
            old_time = appt.get("time_slot") or ""
            note_block = f'<p style="margin:8px 0;color:#475569;font-size:14px;"><strong>Note du client :</strong> {note}</p>' if note else ""
            html = f"""
<div style="font-family:Arial,sans-serif;max-width:560px;margin:0 auto;color:#0F172A;">
  <div style="background:#F59E0B;color:#FFFFFF;padding:16px 18px;border-radius:8px 8px 0 0;">
    <h2 style="margin:0;font-size:18px;">🔄 Le client propose un autre moment</h2>
  </div>
  <div style="background:#FFFFFF;border:1px solid #E5E7EB;border-top:none;padding:20px;border-radius:0 0 8px 8px;">
    <p style="margin:0 0 10px 0;"><strong>{client_name}</strong>{f' — 📞 {client_phone}' if client_phone else ''}</p>
    <p style="margin:0 0 6px 0;color:#64748B;font-size:13px;">RDV initial :</p>
    <p style="margin:0 0 14px 0;font-size:15px;"><s>{old_date} à {old_time}</s></p>
    <p style="margin:0 0 6px 0;color:#64748B;font-size:13px;">Proposition du client :</p>
    <p style="margin:0 0 14px 0;font-size:18px;font-weight:700;color:#92400E;">{date_pretty} à {new_time}</p>
    {note_block}
    <p style="font-size:13px;color:#475569;margin-top:14px;">Connectez-vous à Gexia360 pour accepter, modifier ou contacter le client.</p>
  </div>
</div>"""
            await asyncio.to_thread(resend.Emails.send, {
                "from": os.environ.get("RESEND_FROM") or "onboarding@resend.dev",
                "to": [owner_email],
                "subject": f"🔄 {appt.get('client_name','Client')} propose : {date_pretty} à {new_time}",
                "html": html,
            })
            logger.info(f"Client alternative proposal email sent to owner ({owner_email}) for appt {appointment_id}")
        except Exception as e:
            logger.warning(f"client-suggest-alternative owner notify failed: {e}")

    return HTMLResponse(_render_status_page(
        title="🔄 Proposition envoyée!",
        color="#F59E0B",
        message=(
            f"Merci! Votre proposition de <strong>{date_pretty} à {new_time}</strong> a bien été reçue.<br><br>"
            "Nous vous contacterons rapidement pour confirmer."
        ),
        back_url=f"{host_base}/api/appointments/{appointment_id}/client-confirm",
    ))




def _render_status_page(title: str, color: str, message: str, back_url: str = "") -> str:
    """Small helper to render a centered status confirmation page."""
    back_btn = (
        f'<p style="text-align:center;margin-top:18px;"><a href="{back_url}" style="color:#0B5394;text-decoration:none;font-weight:600;">← Retour à mon rendez-vous</a></p>'
        if back_url else ""
    )
    return f"""<!doctype html>
<html lang="fr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title}</title>
<style>
body{{font-family:-apple-system,Arial,sans-serif;background:#F8FAFC;margin:0;padding:0;min-height:100vh;display:flex;align-items:center;justify-content:center;}}
.card{{background:#FFFFFF;border-radius:14px;padding:28px;max-width:420px;width:90%;box-shadow:0 6px 20px rgba(0,0,0,0.08);text-align:center;}}
h1{{color:{color};font-size:26px;margin:8px 0 14px;}}
p{{color:#374151;font-size:15px;line-height:1.6;}}
.brand{{color:#94A3B8;font-size:12px;margin-top:18px;}}
</style></head>
<body><div class="card">
<h1>{title}</h1>
<p>{message}</p>
{back_btn}
<p class="brand">Lavage de Vitres Bois-Franc &mdash; Gexia360</p>
</div></body></html>"""


@api_router.get("/calendar/info")
async def calendar_info(request: Request):
    """Return the calendar subscription URL (with token) for the frontend page."""
    token = os.environ.get("CALENDAR_TOKEN", "").strip()
    if not token:
        return {"enabled": False, "message": "CALENDAR_TOKEN non configuré"}
    # Build URL from incoming request host so it always matches what the user
    # is actually using (preview, production, custom domain, etc.)
    try:
        host = request.headers.get("x-forwarded-host") or request.headers.get("host") or request.url.hostname or "localhost"
        proto = request.headers.get("x-forwarded-proto") or ("https" if request.url.scheme == "https" else "https")
        # Strip port for clean URL
        host = host.split(",")[0].strip()
        base = f"{proto}://{host}"
    except Exception:
        base = (os.environ.get("APP_URL") or "").strip().rstrip("/")
    url = f"{base}/api/calendar/{token}.ics"
    # Webcal (one-tap subscribe in Apple Calendar) version
    webcal = url.replace("https://", "webcal://").replace("http://", "webcal://")
    return {
        "enabled": True,
        "url": url,
        "webcal_url": webcal,
        "refresh_minutes": 15,
    }


@api_router.post("/invoice/{appointment_id}/send")
async def send_invoice_email(appointment_id: str):
    """Email the invoice HTML to the client + BCC the business owner.

    - Reuses the same HTML template as GET /api/invoice/{id}
    - Strips the print/action UI for clean email rendering
    - BCC: NOTIFY_EMAIL (so the owner keeps a copy of every sent invoice)
    - Returns {sent: true, to, bcc} on success or 502 on Resend failure
    """
    appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not appt:
        raise HTTPException(status_code=404, detail="Rendez-vous introuvable")
    client_email = (appt.get("client_email") or "").strip()
    if not client_email:
        raise HTTPException(status_code=400, detail="Ce client n'a pas de courriel.")
    if not resend.api_key:
        raise HTTPException(status_code=503, detail="Service email non configuré.")

    # Re-use the existing invoice HTML by calling the GET endpoint internally
    inner = await generate_invoice(appointment_id)
    html = inner.body.decode("utf-8") if hasattr(inner, "body") else str(inner)
    # Strip the on-screen actions bar and the auto-print script (not wanted in emails)
    html = re.sub(r'<div class="actions-bar">[\s\S]*?</div>', '', html, count=1)
    html = re.sub(r'<script>\s*window\.onload\s*=\s*function\(\)\s*\{[^<]*\}\s*</script>', '', html, count=1)

    invoice_num = appointment_id[:8].upper()
    subject = f"Votre facture #{invoice_num} — Lavage de Vitres Bois-Franc"
    from_addr = os.environ.get("RESEND_FROM") or "onboarding@resend.dev"

    payload = {
        "from": from_addr,
        "to": [client_email],
        "subject": subject,
        "html": html,
    }
    bcc_addr = (NOTIFY_EMAIL or "").strip()
    if bcc_addr and bcc_addr.lower() != client_email.lower():
        payload["bcc"] = [bcc_addr]

    try:
        result = await asyncio.to_thread(resend.Emails.send, payload)
        sent_id = (result or {}).get("id", "")
        logger.info(
            f"Invoice #{invoice_num} sent to {client_email}"
            + (f" (BCC: {bcc_addr})" if bcc_addr else "")
            + (f" - resend_id={sent_id}" if sent_id else "")
        )
        return {
            "sent": True,
            "to": client_email,
            "bcc": bcc_addr or None,
            "subject": subject,
            "invoice_num": invoice_num,
            "resend_id": sent_id,
        }
    except Exception as e:
        logger.error(f"Failed to send invoice #{invoice_num} to {client_email}: {e}")
        raise HTTPException(status_code=502, detail=f"Échec d'envoi: {e}")


# --- 24-hour reminders ---

@api_router.get("/reminders/tomorrow")
async def get_tomorrow_reminders():
    """List all upcoming appointments scheduled for tomorrow (Eastern time).

    Returns: { date, date_label, count, appointments: [...] }.
    Each appointment includes its reminder_email_sent_at / reminder_sms_sent_at fields.
    """
    return await reminders_module.get_tomorrow_appointments(db)


@api_router.post("/reminders/run-now")
async def run_reminders_now():
    """Manually trigger the 24h reminder job (sends emails to clients + summary to owner).

    Useful for testing or for the user to re-trigger if the 9 AM job missed.
    """
    return await reminders_module.send_24h_reminders_for_tomorrow(db)


@api_router.post("/reminders/{appointment_id}/mark-sms-sent")
async def mark_reminder_sms_sent(appointment_id: str):
    """Mark that the user manually sent the SMS reminder via their phone.

    Called from the frontend after the iPhone SMS sheet is opened so the
    /reminders screen can display a green checkmark.
    """
    appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not appt:
        raise HTTPException(status_code=404, detail="Rendez-vous introuvable")
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.appointments.update_one(
        {"id": appointment_id},
        {"$set": {"reminder_sms_sent_at": now_iso}},
    )
    return {"id": appointment_id, "reminder_sms_sent_at": now_iso}


@api_router.post("/reminders/{appointment_id}/mark-sms-unsent")
async def mark_reminder_sms_unsent(appointment_id: str):
    """Reset the SMS-sent marker (in case the user tapped by mistake)."""
    appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not appt:
        raise HTTPException(status_code=404, detail="Rendez-vous introuvable")
    await db.appointments.update_one(
        {"id": appointment_id},
        {"$unset": {"reminder_sms_sent_at": ""}},
    )
    return {"id": appointment_id, "reminder_sms_sent_at": None}


# --- Proposed alternatives (3 tentative slots while waiting for client reply) ---

@api_router.put("/appointments/{appointment_id}/proposed-alternatives")
async def set_proposed_alternatives(appointment_id: str, data: ProposedAlternativesUpdate):
    """Save up to 3 tentatively-proposed alternative slots on an appointment.

    Each item must have date (YYYY-MM-DD), time_slot (HH:MM) and duration_minutes (int).
    These are shown in YELLOW on the calendar so the user doesn't double-book those
    slots while waiting for the client's confirmation.
    """
    appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not appt:
        raise HTTPException(status_code=404, detail="Rendez-vous introuvable")
    cleaned = []
    for item in (data.alternatives or [])[:3]:
        d = (item.get("date") or "").strip()
        t = (item.get("time_slot") or "")[:5]
        dur = int(item.get("duration_minutes") or 60)
        if d and t:
            cleaned.append({"date": d, "time_slot": t, "duration_minutes": dur})
    await db.appointments.update_one(
        {"id": appointment_id},
        {"$set": {"proposed_alternatives": cleaned}},
    )
    return {"id": appointment_id, "proposed_alternatives": cleaned}


@api_router.delete("/appointments/{appointment_id}/proposed-alternatives")
async def clear_proposed_alternatives(appointment_id: str):
    """Clear all tentatively-proposed alternative slots on an appointment."""
    appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not appt:
        raise HTTPException(status_code=404, detail="Rendez-vous introuvable")
    await db.appointments.update_one(
        {"id": appointment_id},
        {"$unset": {"proposed_alternatives": ""}},
    )
    return {"id": appointment_id, "proposed_alternatives": []}


@api_router.post("/appointments/{appointment_id}/confirm-alternative")
async def confirm_alternative(appointment_id: str, payload: Dict[str, Any] = Body(...)):
    """Confirm one of the proposed alternatives → updates date/time and clears the rest.

    Body: { "date": "YYYY-MM-DD", "time_slot": "HH:MM", "notify_client": bool }
    """
    appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not appt:
        raise HTTPException(status_code=404, detail="Rendez-vous introuvable")
    new_date = (payload.get("date") or "").strip()
    new_time = (payload.get("time_slot") or "")[:5]
    if not new_date or not new_time:
        raise HTTPException(status_code=400, detail="date et time_slot requis")
    update = AppointmentUpdate(
        date=new_date,
        time_slot=new_time,
        notify_client=bool(payload.get("notify_client", False)),
    )
    # Reuse update_appointment to also send the reschedule email (if requested)
    response = await update_appointment(appointment_id, update)
    # Then clear proposed_alternatives
    await db.appointments.update_one(
        {"id": appointment_id},
        {"$unset": {"proposed_alternatives": ""}},
    )
    return response


# --- Monthly Report ---

@api_router.get("/report/monthly")
async def monthly_report(month: Optional[str] = None):
    """Generate monthly report"""
    from fastapi.responses import HTMLResponse
    now = datetime.now(timezone.utc)
    if not month:
        month = now.strftime("%Y-%m")
    month_start = f"{month}-01"
    month_end = f"{month}-31"

    appts = await db.appointments.find({"date": {"$gte": month_start, "$lte": month_end}}, {"_id": 0}).sort("date", 1).to_list(10000)
    reqs = await db.appointment_requests.find({"preferred_date": {"$gte": month_start, "$lte": month_end}}, {"_id": 0}).to_list(10000)

    total_rev = sum(a.get('price', 0) for a in appts)
    completed = len([a for a in appts if a.get('status') == 'completed'])
    upcoming = len([a for a in appts if a.get('status') == 'upcoming'])
    cancelled = len([a for a in appts if a.get('status') == 'cancelled'])
    accepted_reqs = len([r for r in reqs if r.get('status') == 'accepted'])
    pending_reqs = len([r for r in reqs if r.get('status') == 'pending'])

    # Top clients
    client_stats = {}
    for a in appts:
        name = a.get('client_name', '')
        if name not in client_stats:
            client_stats[name] = {'count': 0, 'revenue': 0}
        client_stats[name]['count'] += 1
        client_stats[name]['revenue'] += a.get('price', 0)
    top = sorted(client_stats.items(), key=lambda x: x[1]['revenue'], reverse=True)[:10]

    appt_rows = ""
    for a in appts:
        appt_rows += f"<tr><td>{a.get('date','')}</td><td>{a.get('time_slot','')}</td><td>{a.get('client_name','')}</td><td>{a.get('title','')}</td><td>{a.get('price',0):.2f}$</td><td>{a.get('status','')}</td></tr>"

    client_rows = ""
    for name, stats in top:
        client_rows += f"<tr><td>{name}</td><td>{stats['count']}</td><td>{stats['revenue']:.2f}$</td></tr>"

    display_month = datetime.strptime(month + "-01", "%Y-%m-%d").strftime("%B %Y")

    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>Rapport {display_month}</title>
<style>
body{{font-family:-apple-system,sans-serif;max-width:800px;margin:30px auto;padding:20px;color:#0A0A0A;font-size:13px;}}
h1{{font-size:24px;font-weight:800;}}.brand{{color:#0891B2;font-size:14px;margin-bottom:24px;}}
h2{{font-size:16px;margin-top:24px;border-bottom:2px solid #0891B2;padding-bottom:6px;}}
.stats{{display:flex;gap:12px;margin:16px 0;}}
.stat{{flex:1;background:#F5F5F5;border-radius:8px;padding:14px;text-align:center;}}
.stat-val{{font-size:22px;font-weight:800;color:#0891B2;}}.stat-label{{font-size:11px;color:#737373;text-transform:uppercase;margin-top:4px;}}
table{{width:100%;border-collapse:collapse;margin-top:10px;}}
th{{background:#0891B2;color:white;padding:8px;text-align:left;font-size:11px;text-transform:uppercase;}}
td{{padding:7px 8px;border-bottom:1px solid #E5E5E5;}}
@media print{{body{{margin:0;}}}}
</style></head><body>
<h1>Rapport mensuel</h1>
<div class="brand">{display_month} — Gexia360</div>
<div class="stats">
<div class="stat"><div class="stat-val">{total_rev:.2f}$</div><div class="stat-label">Revenu</div></div>
<div class="stat"><div class="stat-val">{len(appts)}</div><div class="stat-label">RDV total</div></div>
<div class="stat"><div class="stat-val">{completed}</div><div class="stat-label">Complétés</div></div>
<div class="stat"><div class="stat-val">{cancelled}</div><div class="stat-label">Annulés</div></div>
<div class="stat"><div class="stat-val">{len(reqs)}</div><div class="stat-label">Demandes</div></div>
</div>
<h2>Meilleurs clients</h2>
<table><tr><th>Client</th><th>RDV</th><th>Revenu</th></tr>{client_rows}</table>
<h2>Tous les rendez-vous ({len(appts)})</h2>
<table><tr><th>Date</th><th>Heure</th><th>Client</th><th>Service</th><th>Prix</th><th>Statut</th></tr>{appt_rows}</table>
<script>window.onload=function(){{window.print();}}</script>
</body></html>"""
    return HTMLResponse(content=html)

# --- Client Reviews ---

class ReviewCreate(BaseModel):
    appointment_id: str
    client_name: str
    rating: int  # 1-5
    comment: Optional[str] = ""

class ReviewResponse(BaseModel):
    id: str
    appointment_id: str
    client_name: str
    rating: int
    comment: str
    created_at: str

@api_router.post("/reviews", response_model=ReviewResponse)
async def create_review(data: ReviewCreate):
    review = {
        "id": str(uuid.uuid4()),
        "appointment_id": data.appointment_id,
        "client_name": data.client_name,
        "rating": data.rating,
        "comment": data.comment or "",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.reviews.insert_one(review)
    return ReviewResponse(**{k: v for k, v in review.items() if k != "_id"})

@api_router.get("/reviews", response_model=List[ReviewResponse])
async def get_reviews():
    reviews = await db.reviews.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [ReviewResponse(**r) for r in reviews]

@api_router.post("/reviews/send-request/{appointment_id}")
async def send_review_request(appointment_id: str):
    """Send review request email to client after service"""
    appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not appt:
        raise HTTPException(status_code=404, detail="Appointment not found")

    client_email = appt.get("client_email", "")
    if not client_email:
        raise HTTPException(status_code=400, detail="Client n'a pas de courriel")

    app_url = os.environ.get("APP_URL", "").rstrip("/")
    review_url = f"{app_url}/api/review-page/{appointment_id}"
    html = f"""<div style="font-family:sans-serif;max-width:500px;margin:0 auto;padding:20px;">
    <h2 style="color:#0891B2;">Comment était votre expérience?</h2>
    <p>Bonjour {appt.get('client_name','')},</p>
    <p>Merci d'avoir fait appel à nos services! Nous aimerions avoir votre avis.</p>
    <a href="{review_url}" style="display:inline-block;background:#0891B2;color:white;padding:12px 24px;border-radius:4px;text-decoration:none;font-weight:600;margin:16px 0;">Laisser un avis</a>
    <p style="color:#A3A3A3;font-size:13px;">— Gexia360</p>
    </div>"""

    if resend.api_key:
        try:
            await asyncio.to_thread(resend.Emails.send, {
                "from": os.environ.get("RESEND_FROM") or "onboarding@resend.dev",
                "to": [client_email],
                "subject": "Comment était votre expérience? — Gexia360",
                "html": inject_branding(html),
            })
            return {"message": f"Demande d'avis envoyée à {client_email}"}
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
    raise HTTPException(status_code=400, detail="Email non configuré")

@api_router.get("/review-page/{appointment_id}")
async def review_page(appointment_id: str):
    """Public page for client to leave a review"""
    from fastapi.responses import HTMLResponse
    appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    name = appt.get('client_name', 'Client') if appt else 'Client'

    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Votre avis</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box;}}body{{font-family:-apple-system,sans-serif;background:#FAFAFA;display:flex;justify-content:center;padding:40px 20px;}}
.c{{max-width:480px;width:100%;}}h1{{font-size:24px;font-weight:800;margin-bottom:8px;}}
.sub{{color:#737373;margin-bottom:24px;}}.stars{{display:flex;gap:8px;margin:16px 0;}}
.star{{font-size:36px;cursor:pointer;color:#E5E5E5;transition:color 0.2s;}}.star.active{{color:#F59E0B;}}
textarea{{width:100%;border:none;border-bottom:1px solid #E5E5E5;padding:12px 0;font-size:16px;font-family:inherit;resize:none;min-height:80px;outline:none;}}
.btn{{width:100%;padding:16px;background:#0891B2;color:white;border:none;border-radius:4px;font-size:16px;font-weight:600;cursor:pointer;margin-top:16px;}}
.success{{display:none;text-align:center;padding:40px;}}.success h2{{font-size:22px;margin-bottom:8px;}}.success p{{color:#737373;}}
</style></head><body><div class="c">
<div id="form"><h1>Votre avis compte!</h1><p class="sub">Merci {name}, comment était votre expérience?</p>
<div class="stars" id="stars"></div>
<textarea id="comment" placeholder="Commentaire (optionnel)..."></textarea>
<button class="btn" onclick="submit()">Envoyer mon avis</button></div>
<div class="success" id="success"><h2>Merci!</h2><p>Votre avis a été enregistré.</p></div>
<script>
let rating=0;const stars=document.getElementById('stars');
for(let i=1;i<=5;i++){{const s=document.createElement('span');s.className='star';s.textContent='★';s.onclick=()=>{{rating=i;document.querySelectorAll('.star').forEach((el,idx)=>el.className=idx<i?'star active':'star');}};stars.appendChild(s);}}
async function submit(){{if(!rating){{alert('Choisissez une note');return;}}
const res=await fetch('/api/reviews',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{appointment_id:'{appointment_id}',client_name:'{name}',rating,comment:document.getElementById('comment').value}})}});
if(res.ok){{document.getElementById('form').style.display='none';document.getElementById('success').style.display='block';}}}}
</script></div></body></html>"""
    return HTMLResponse(content=html)


# --- Backup & Export ---

@api_router.get("/backup/export")
async def export_backup():
    """Export all data as readable HTML page"""
    from fastapi.responses import HTMLResponse
    appointments = await db.appointments.find({}, {"_id": 0}).sort("date", -1).to_list(10000)
    requests = await db.appointment_requests.find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)

    appt_rows = ""
    for a in appointments:
        appt_rows += f"""<tr>
            <td>{a.get('date','')}</td><td>{a.get('time_slot','')}</td>
            <td>{a.get('client_name','')}</td><td>{a.get('client_phone','')}</td>
            <td>{a.get('client_email','')}</td><td>{a.get('client_address','')}</td>
            <td>{a.get('duration_minutes','')}m</td><td>{a.get('price',0):.2f} $</td>
            <td>{a.get('status','')}</td><td>{a.get('notes','')}</td>
        </tr>"""

    req_rows = ""
    for r in requests:
        req_rows += f"""<tr>
            <td>{r.get('preferred_date','')}</td><td>{r.get('preferred_time','')}</td>
            <td>{r.get('customer_name','')}</td><td>{r.get('customer_phone','')}</td>
            <td>{r.get('customer_email','')}</td><td>{r.get('customer_address','')}</td>
            <td>{r.get('status','')}</td><td>{r.get('message','')}</td>
        </tr>"""

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")

    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>Backup Gexia360 - {now}</title>
<style>
body{{font-family:-apple-system,sans-serif;max-width:1200px;margin:20px auto;padding:20px;color:#0A0A0A;font-size:13px;}}
h1{{font-size:24px;margin-bottom:2px;}}
.brand{{color:#0891B2;margin-bottom:20px;}}
h2{{font-size:18px;margin-top:30px;border-bottom:2px solid #0891B2;padding-bottom:6px;}}
table{{width:100%;border-collapse:collapse;margin-top:10px;}}
th{{background:#0891B2;color:white;padding:8px 6px;text-align:left;font-size:12px;text-transform:uppercase;letter-spacing:0.3px;}}
td{{padding:7px 6px;border-bottom:1px solid #E5E5E5;}}
tr:hover{{background:#F5F5F5;}}
.count{{color:#737373;font-size:14px;}}
@media print{{body{{margin:0;font-size:11px;}} th{{background:#333;}} }}
</style></head><body>
<h1>Backup Gexia360</h1>
<div class="brand">{now}</div>

<h2>Rendez-vous <span class="count">({len(appointments)})</span></h2>
<table>
<tr><th>Date</th><th>Heure</th><th>Client</th><th>Tél.</th><th>Courriel</th><th>Adresse</th><th>Durée</th><th>Prix</th><th>Statut</th><th>Notes</th></tr>
{appt_rows}
</table>

<h2>Demandes <span class="count">({len(requests)})</span></h2>
<table>
<tr><th>Date</th><th>Heure</th><th>Client</th><th>Tél.</th><th>Courriel</th><th>Adresse</th><th>Statut</th><th>Message</th></tr>
{req_rows}
</table>

<script>window.onload=function(){{window.print();}}</script>
</body></html>"""
    return HTMLResponse(content=html)

@api_router.get("/backup/clients-csv")
async def export_clients_csv():
    """Export all client data as CSV text"""
    from fastapi.responses import PlainTextResponse
    pipeline = [
        {"$group": {
            "_id": "$client_name",
            "count": {"$sum": 1},
            "total_spent": {"$sum": "$price"},
            "last_visit": {"$max": "$date"},
            "email": {"$first": "$client_email"},
            "phone": {"$first": "$client_phone"},
            "address": {"$first": "$client_address"},
        }},
        {"$sort": {"last_visit": -1}},
    ]
    clients = await db.appointments.aggregate(pipeline).to_list(500)

    lines = ["Nom,Courriel,Téléphone,Adresse,Nombre RDV,Total dépensé,Dernière visite"]
    for c in clients:
        if not c["_id"]:
            continue
        name = (c["_id"] or "").replace(",", " ")
        email = (c.get("email") or "").replace(",", " ")
        phone = (c.get("phone") or "").replace(",", " ")
        address = (c.get("address") or "").replace(",", " ")
        lines.append(f'{name},{email},{phone},{address},{c["count"]},{c.get("total_spent", 0):.2f},{c.get("last_visit", "")}')

    csv_text = "\n".join(lines)
    return PlainTextResponse(content=csv_text, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=clients_brightcalendar.csv"})

@api_router.get("/print/appointment/{appointment_id}")
async def print_appointment(appointment_id: str):
    """Generate printable HTML for an appointment"""
    from fastapi.responses import HTMLResponse
    appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not appt:
        raise HTTPException(status_code=404, detail="Appointment not found")

    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>RDV - {appt.get('client_name','')}</title>
<style>
body{{font-family:-apple-system,sans-serif;max-width:600px;margin:40px auto;padding:20px;color:#0A0A0A;}}
h1{{font-size:24px;margin-bottom:4px;}}
.brand{{color:#0891B2;font-size:14px;margin-bottom:24px;}}
table{{width:100%;border-collapse:collapse;margin-top:16px;}}
td{{padding:10px 0;vertical-align:top;}}
.label{{font-size:12px;font-weight:600;color:#737373;text-transform:uppercase;letter-spacing:0.5px;width:120px;}}
.value{{font-size:15px;}}
.divider{{border-top:1px solid #E5E5E5;margin:16px 0;}}
.price{{font-size:20px;font-weight:800;color:#0891B2;}}
@media print{{body{{margin:0;}}}}
</style></head><body>
<h1>{appt.get('title','Rendez-vous')}</h1>
<div class="brand">Gexia360</div>
<div class="divider"></div>
<table>
<tr><td class="label">Client</td><td class="value">{appt.get('client_name','')}</td></tr>
<tr><td class="label">Courriel</td><td class="value">{appt.get('client_email','')}</td></tr>
<tr><td class="label">Téléphone</td><td class="value">{appt.get('client_phone','')}</td></tr>
<tr><td class="label">Adresse</td><td class="value">{appt.get('client_address','')}</td></tr>
</table>
<div class="divider"></div>
<table>
<tr><td class="label">Date</td><td class="value">{appt.get('date','')}</td></tr>
<tr><td class="label">Heure</td><td class="value">{appt.get('time_slot','')}</td></tr>
<tr><td class="label">Durée</td><td class="value">{appt.get('duration_minutes','')} minutes</td></tr>
</table>
<div class="divider"></div>
<table>
<tr><td class="label">Prix</td><td class="price">{appt.get('price',0):.2f} $</td></tr>
</table>
{"<div class='divider'></div><table><tr><td class='label'>Notes</td><td class='value'>" + appt.get('notes','') + "</td></tr></table>" if appt.get('notes') else ""}
<div class="divider"></div>
<p style="font-size:12px;color:#A3A3A3;margin-top:24px;">Imprimé depuis Gexia360</p>
<script>window.onload=function(){{window.print();}}</script>
</body></html>"""
    return HTMLResponse(content=html)


# --- Backup by Email ---

@api_router.post("/backup/email")
async def backup_by_email():
    """Send full backup to owner's email — with JSON file attached + HTML preview."""
    appointments = await db.appointments.find({}, {"_id": 0}).sort("date", -1).to_list(10000)
    requests_data = await db.appointment_requests.find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    clients_data = await db.clients.find({}, {"_id": 0}).to_list(10000)
    revenues_data = await db.revenues.find({}, {"_id": 0}).to_list(10000)
    expenses_data = await db.expenses.find({}, {"_id": 0}).to_list(10000)

    now_full = datetime.now(timezone.utc)
    now = now_full.strftime("%Y-%m-%d %H:%M")
    today_iso = now_full.strftime("%Y-%m-%d")

    # Build the full JSON payload (same shape as /api/backup/export)
    payload = {
        "exported_at": now_full.isoformat(),
        "version": 1,
        "appointments": appointments,
        "appointment_requests": requests_data,
        "clients": clients_data,
        "revenues": revenues_data,
        "expenses": expenses_data,
    }
    json_bytes = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    filename = f"gexia360-sauvegarde-{today_iso}.json"
    json_b64 = base64.b64encode(json_bytes).decode("ascii")

    # Build HTML preview
    appt_rows = ""
    for a in appointments[:50]:  # limit preview to 50 rows for performance
        appt_rows += f"<tr><td>{a.get('date','')}</td><td>{a.get('time_slot','')}</td><td>{a.get('client_name','')}</td><td>{a.get('client_phone','')}</td><td>{a.get('client_email','')}</td><td>{a.get('client_address','')[:30]}</td><td>{a.get('price',0):.2f}$</td><td>{a.get('status','')}</td></tr>"
    req_rows = ""
    for r in requests_data[:30]:
        req_rows += f"<tr><td>{r.get('preferred_date','')}</td><td>{r.get('preferred_time','')}</td><td>{r.get('customer_name','')}</td><td>{r.get('customer_phone','')}</td><td>{r.get('customer_email','')}</td><td>{r.get('status','')}</td></tr>"

    size_kb = round(len(json_bytes) / 1024, 1)
    html = f"""<div style="font-family:sans-serif;max-width:800px;margin:0 auto;font-size:13px;">
    <h1 style="color:#0891B2;margin:0 0 8px 0;">📦 Sauvegarde Gexia360</h1>
    <p style="color:#475569;margin:0 0 16px 0;">{now} — Le fichier complet est <strong>en pièce jointe</strong> ({filename}, {size_kb} Ko)</p>
    <div style="background:#F0F9FF;border-left:4px solid #0891B2;padding:12px 14px;border-radius:8px;margin:0 0 16px 0;">
      <strong>📊 Contenu de la sauvegarde:</strong><br>
      📅 Rendez-vous: <strong>{len(appointments)}</strong><br>
      📨 Demandes: <strong>{len(requests_data)}</strong><br>
      👥 Clients: <strong>{len(clients_data)}</strong><br>
      💰 Revenus: <strong>{len(revenues_data)}</strong><br>
      💸 Dépenses: <strong>{len(expenses_data)}</strong>
    </div>
    <p style="color:#64748B;font-size:12px;font-style:italic;margin:16px 0;">
      💾 Pour restaurer: enregistrez la pièce jointe sur votre iPhone (Fichiers → iCloud Drive),
      puis dans l'app Gexia360 → Sauvegarde → "Restaurer un fichier JSON".
    </p>
    <h2 style="color:#0F172A;margin:20px 0 8px 0;">Rendez-vous (aperçu — {min(50, len(appointments))}/{len(appointments)})</h2>
    <table style="width:100%;border-collapse:collapse;font-size:11px;"><tr style="background:#0891B2;color:white;"><th style="padding:6px;">Date</th><th style="padding:6px;">Heure</th><th style="padding:6px;">Client</th><th style="padding:6px;">Tél</th><th style="padding:6px;">Email</th><th style="padding:6px;">Adresse</th><th style="padding:6px;">Prix</th><th style="padding:6px;">Statut</th></tr>{appt_rows}</table>
    <h2 style="color:#0F172A;margin:20px 0 8px 0;">Demandes (aperçu — {min(30, len(requests_data))}/{len(requests_data)})</h2>
    <table style="width:100%;border-collapse:collapse;font-size:11px;"><tr style="background:#0891B2;color:white;"><th style="padding:6px;">Date</th><th style="padding:6px;">Heure</th><th style="padding:6px;">Client</th><th style="padding:6px;">Tél</th><th style="padding:6px;">Email</th><th style="padding:6px;">Statut</th></tr>{req_rows}</table>
    </div>"""

    if NOTIFY_EMAIL and resend.api_key:
        try:
            await asyncio.to_thread(resend.Emails.send, {
                "from": os.environ.get("RESEND_FROM") or "onboarding@resend.dev",
                "to": [NOTIFY_EMAIL],
                "subject": f"📦 Sauvegarde Gexia360 — {now} ({size_kb} Ko)",
                "html": inject_branding(html),
                "attachments": [
                    {
                        "filename": filename,
                        "content": json_b64,
                        "content_type": "application/json",
                    }
                ],
            })
            logger.info(f"Backup sent by email to {NOTIFY_EMAIL} ({size_kb} Ko, {len(appointments)} appts)")
            return {
                "message": f"Backup envoyé à {NOTIFY_EMAIL}",
                "filename": filename,
                "size_kb": size_kb,
                "to": NOTIFY_EMAIL,
                "counts": {
                    "appointments": len(appointments),
                    "requests": len(requests_data),
                    "clients": len(clients_data),
                    "revenues": len(revenues_data),
                    "expenses": len(expenses_data),
                },
            }
        except Exception as e:
            logger.error(f"Backup email failed: {e}")
            raise HTTPException(status_code=500, detail=f"Erreur envoi: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="Email non configuré")

# --- Price Estimation ---

class PriceEstimate(BaseModel):
    num_windows: int = 0
    window_type: str = "standard"  # standard, large, skylight

@api_router.post("/estimate")
async def estimate_price(data: PriceEstimate):
    """Calculate price estimate based on number of windows"""
    rates = {
        "standard": 15.0,
        "standard_coulissante": 20.0,
        "standard_double_coulissante": 40.0,
        "large": 20.0,
        "skylight": 30.0,
        "patio_simple": 40.0,
        "patio_double": 60.0,
    }
    rate = rates.get(data.window_type, 8.0)
    total = data.num_windows * rate
    return {
        "num_windows": data.num_windows,
        "window_type": data.window_type,
        "rate_per_window": rate,
        "estimated_total": round(total, 2),
    }


class EstimateItem(BaseModel):
    label: str
    qty: int = 0
    unit_price: float = 0.0


class EstimateSendRequest(BaseModel):
    client_name: str = ""
    client_email: str = ""
    items: List[EstimateItem] = []
    fixed_price: Optional[float] = None
    discount_percent: float = 0.0
    total: float = 0.0
    notes: str = ""
    valid_until: Optional[str] = ""  # ISO date
    detailed: bool = True  # If False, hide the item breakdown and show only the total


@api_router.post("/estimate/send")
async def send_estimate_email(data: EstimateSendRequest):
    """Send a beautifully branded HTML estimation email to the client,
    with the Lavage de Vitres Bois-Franc logo and contact info automatically
    injected. This is used by the in-app Estimation builder (/estimate)."""
    if not resend.api_key:
        raise HTTPException(status_code=500, detail="Email service not configured")
    if not data.client_email or "@" not in data.client_email:
        raise HTTPException(status_code=400, detail="Valid client email required")

    price_str = f"{data.total:,.2f} $".replace(",", " ")
    client_first = (data.client_name or "").split()[0] if data.client_name else "cher client"

    # Build line items table
    items_rows = ""
    has_items = False
    for it in (data.items or []):
        if it.qty <= 0:
            continue
        has_items = True
        line_total = it.qty * it.unit_price
        items_rows += f"""
        <tr>
          <td style="padding:10px 8px;border-bottom:1px solid #F3F4F6;color:#374151;font-size:13px">{it.label}</td>
          <td style="padding:10px 8px;border-bottom:1px solid #F3F4F6;color:#6B7280;font-size:13px;text-align:center">{it.qty}</td>
          <td style="padding:10px 8px;border-bottom:1px solid #F3F4F6;color:#6B7280;font-size:13px;text-align:right">{it.unit_price:.2f} $</td>
          <td style="padding:10px 8px;border-bottom:1px solid #F3F4F6;color:#111827;font-size:13px;font-weight:600;text-align:right">{line_total:.2f} $</td>
        </tr>
        """
    items_table = ""
    if has_items and data.detailed:
        items_table = f"""
        <table style="width:100%;border-collapse:collapse;margin:16px 0;background:#FAFAFA;border-radius:8px;overflow:hidden">
          <thead>
            <tr style="background:#1E5BA8;color:#fff">
              <th style="padding:10px 8px;text-align:left;font-size:12px;text-transform:uppercase;letter-spacing:0.5px;font-weight:700">Description</th>
              <th style="padding:10px 8px;text-align:center;font-size:12px;text-transform:uppercase;letter-spacing:0.5px;font-weight:700">Qté</th>
              <th style="padding:10px 8px;text-align:right;font-size:12px;text-transform:uppercase;letter-spacing:0.5px;font-weight:700">Prix unit.</th>
              <th style="padding:10px 8px;text-align:right;font-size:12px;text-transform:uppercase;letter-spacing:0.5px;font-weight:700">Total</th>
            </tr>
          </thead>
          <tbody>
            {items_rows}
          </tbody>
        </table>
        """
    elif not data.detailed:
        # Non-detailed mode: show a simple service description instead of the breakdown
        items_table = f"""
        <div style="background:#F9FAFB;border:1px solid #E5E7EB;border-radius:8px;padding:14px 16px;margin:16px 0">
          <p style="margin:0;color:#374151;font-size:15px;font-weight:600;text-align:center">🪟 Service de lavage de vitres</p>
        </div>
        """

    discount_block = ""
    if data.discount_percent and data.discount_percent > 0 and data.detailed:
        discount_block = f"""
        <p style="margin:4px 0;text-align:right;color:#059669;font-weight:700;font-size:14px">
          Rabais: -{data.discount_percent:.0f}%
        </p>
        """

    notes_block = ""
    if data.notes and data.notes.strip():
        safe_note = data.notes.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('\n', '<br>')
        notes_block = f"""
        <div style="background:#FFF7ED;border-left:4px solid #FB923C;padding:14px 16px;border-radius:8px;margin:20px 0">
          <p style="margin:0 0 6px 0;color:#9A3412;font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:0.5px">📝 Notes / Informations</p>
          <p style="margin:0;color:#7C2D12;font-size:14px;line-height:1.6">{safe_note}</p>
        </div>
        """

    valid_block = ""
    if data.valid_until:
        valid_block = f"""
        <p style="margin:8px 0 0 0;color:#6B7280;font-size:12px;text-align:center">
          Valide jusqu'au {data.valid_until}
        </p>
        """

    body_html = f"""
    <div style="text-align:center;padding:8px 0 16px 0">
      <p style="margin:0;font-size:32px">💰</p>
      <h1 style="margin:8px 0 4px 0;color:#111827;font-size:22px">Votre estimation</h1>
      <p style="margin:0;color:#6B7280;font-size:14px">Bonjour {client_first} !</p>
    </div>

    <p style="margin:0 0 14px 0;color:#374151;font-size:15px;line-height:1.6">
      Merci pour votre intérêt pour nos services de lavage de vitres. Voici l'estimation détaillée :
    </p>

    {items_table}
    {discount_block}

    <div style="background:linear-gradient(135deg,#ECFDF5,#D1FAE5);border:2px solid #10B981;border-radius:12px;padding:20px;text-align:center;margin:16px 0">
      <p style="margin:0 0 6px 0;color:#047857;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:1px">Total</p>
      <p style="margin:0;color:#065F46;font-size:34px;font-weight:800">{price_str}</p>
      {valid_block}
    </div>

    {notes_block}

    <p style="margin:20px 0 8px 0;color:#4B5563;font-size:14px;line-height:1.6">
      Cette estimation vous convient-elle ? Répondez-moi simplement ou appelez-moi pour réserver votre rendez-vous.
    </p>

    <p style="margin:16px 0 0 0;color:#6B7280;font-size:13px;font-style:italic;line-height:1.5">
      Merci de votre confiance,<br>
      <strong style="color:#1E5BA8">Louis-Philippe Fournier</strong>
    </p>
    """

    full_html = branding.wrap_email(body_html, subtitle=f"Estimation pour {data.client_name}" if data.client_name else "Estimation",
                                     unsubscribe_url=(f"{os.environ.get('APP_URL', '').rstrip('/')}/api/unsubscribe?email={data.client_email}" if os.environ.get('APP_URL') else ""))

    from_addr = os.environ.get("RESEND_FROM") or "onboarding@resend.dev"
    try:
        await asyncio.to_thread(
            resend.Emails.send,
            {
                "from": from_addr,
                "to": [data.client_email],
                "reply_to": NOTIFY_EMAIL if NOTIFY_EMAIL else None,
                "subject": f"Votre estimation — Lavage de Vitres Bois-Franc — {price_str}",
                "html": inject_branding(full_html),
            },
        )
        # Also notify the business owner
        if NOTIFY_EMAIL:
            try:
                await asyncio.to_thread(
                    resend.Emails.send,
                    {
                        "from": os.environ.get("RESEND_FROM") or "onboarding@resend.dev",
                        "to": [NOTIFY_EMAIL],
                        "subject": f"📤 Estimation envoyée à {data.client_name or data.client_email} — {price_str}",
                        "html": inject_branding(f"<div><p><strong>Estimation envoyée</strong></p><p>Client: {data.client_name or '—'}<br>Courriel: {data.client_email}<br>Montant: <strong>{price_str}</strong></p>{notes_block}</div>"),
                    },
                )
            except Exception:
                pass
        return {"ok": True, "sent_to": data.client_email}
    except Exception as e:
        logging.exception("Failed to send estimate email")
        raise HTTPException(status_code=500, detail=f"Email send failed: {e}")


# --- Share Appointment ---

@api_router.get("/share/appointment/{appointment_id}")
async def share_appointment(appointment_id: str):
    """Generate shareable text for an appointment"""
    appt = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    if not appt:
        raise HTTPException(status_code=404, detail="Appointment not found")

    text = f"""Rendez-vous confirmé — Gexia360

Client: {appt.get('client_name','')}
Date: {appt.get('date','')}
Heure: {appt.get('time_slot','')}
Durée: {appt.get('duration_minutes','')} min
Adresse: {appt.get('client_address','')}
"""
    if appt.get('price', 0) > 0:
        text += f"Prix: {appt['price']:.2f} $\n"

    return {"text": text.strip()}

# --- Recurrence ---

class RecurrenceCreate(BaseModel):
    appointment_id: str
    interval_months: int = 3  # every X months
    occurrences: int = 4  # how many times

@api_router.post("/appointments/recurrence")
async def create_recurring(data: RecurrenceCreate):
    """Create recurring appointments from an existing appointment"""
    original = await db.appointments.find_one({"id": data.appointment_id}, {"_id": 0})
    if not original:
        raise HTTPException(status_code=404, detail="Appointment not found")

    created = []
    base_date = datetime.strptime(original["date"], "%Y-%m-%d")

    for i in range(1, data.occurrences + 1):
        new_date = base_date
        # Add months
        month = base_date.month + (data.interval_months * i)
        year = base_date.year + (month - 1) // 12
        month = ((month - 1) % 12) + 1
        day = min(base_date.day, 28)  # safe day
        new_date = datetime(year, month, day)

        new_appt = {
            "id": str(uuid.uuid4()),
            "title": original["title"],
            "client_name": original["client_name"],
            "client_email": original.get("client_email", ""),
            "client_phone": original.get("client_phone", ""),
            "client_address": original.get("client_address", ""),
            "date": new_date.strftime("%Y-%m-%d"),
            "time_slot": original["time_slot"],
            "duration_minutes": original["duration_minutes"],
            "price": original.get("price", 0),
            "notes": f"Récurrence #{i} — {original.get('notes', '')}",
            "status": "upcoming",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.appointments.insert_one(new_appt)
        created.append({"id": new_appt["id"], "date": new_appt["date"]})

    return {"message": f"{len(created)} rendez-vous créés", "appointments": created}



async def seed_requests():
    """Seed sample requests for testing"""
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")

    samples = [
        {
            "id": str(uuid.uuid4()),
            "customer_name": "Emma Johnson",
            "customer_email": "emma@example.com",
            "customer_phone": "514-555-1234",
            "customer_address": "123 Rue Principale, Bois-Franc",
            "preferred_date": today,
            "preferred_time": "10:00",
            "message": "I'd like to discuss a new marketing strategy for Q3.",
            "status": "pending",
            "suggested_date": None,
            "suggested_time": None,
            "suggested_note": None,
            "created_at": now.isoformat(),
        },
        {
            "id": str(uuid.uuid4()),
            "customer_name": "Frank Miller",
            "customer_email": "frank@company.org",
            "customer_phone": "514-555-5678",
            "customer_address": "456 Boulevard des Sources",
            "preferred_date": today,
            "preferred_time": "13:00",
            "message": "Need help with tax planning for the new fiscal year.",
            "status": "pending",
            "suggested_date": None,
            "suggested_time": None,
            "suggested_note": None,
            "created_at": now.isoformat(),
        },
        {
            "id": str(uuid.uuid4()),
            "customer_name": "Grace Lee",
            "customer_email": "grace.lee@startup.io",
            "customer_phone": "438-555-9012",
            "customer_address": "789 Avenue Sainte-Croix",
            "preferred_date": today,
            "preferred_time": "15:30",
            "message": "Looking for consulting on our product launch timeline.",
            "status": "pending",
            "suggested_date": None,
            "suggested_time": None,
            "suggested_note": None,
            "created_at": now.isoformat(),
        },
    ]

    await db.appointment_requests.delete_many({})
    await db.appointment_requests.insert_many(samples)
    return {"message": f"Seeded {len(samples)} requests for {today}"}

@api_router.post("/appointments/seed")
async def seed_appointments():
    """Seed sample appointments for testing"""
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")
    
    samples = [
        {
            "id": str(uuid.uuid4()),
            "title": "Strategy Review",
            "client_name": "Alice Martin",
            "client_email": "alice@example.com",
            "client_phone": "514-555-0001",
            "client_address": "100 Rue Principale, Bois-Franc",
            "date": today,
            "time_slot": "09:00",
            "duration_minutes": 60,
            "notes": "Quarterly strategy review meeting",
            "status": "upcoming",
            "created_at": now.isoformat(),
        },
        {
            "id": str(uuid.uuid4()),
            "title": "Project Kickoff",
            "client_name": "Bob Chen",
            "client_email": "bob@company.com",
            "client_phone": "514-555-0002",
            "client_address": "200 Boulevard des Sources",
            "date": today,
            "time_slot": "11:00",
            "duration_minutes": 45,
            "notes": "New website redesign project",
            "status": "upcoming",
            "created_at": now.isoformat(),
        },
        {
            "id": str(uuid.uuid4()),
            "title": "Budget Planning",
            "client_name": "Carol Davis",
            "client_email": "carol@firm.ca",
            "client_phone": "438-555-0003",
            "client_address": "300 Avenue Sainte-Croix",
            "date": today,
            "time_slot": "14:00",
            "duration_minutes": 30,
            "notes": "Annual budget discussion",
            "status": "upcoming",
            "created_at": now.isoformat(),
        },
    ]
    
    await db.appointments.delete_many({})
    await db.appointments.insert_many(samples)
    return {"message": f"Seeded {len(samples)} appointments for {today}"}

# NOTE: app.include_router(api_router) has been moved to the end of the file
# (after SCHEDULED CAMPAIGNS endpoints) so new routes are registered.

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    try:
        scheduler.shutdown(wait=False)
    except Exception:
        pass
    client.close()


# ============================================================
# EXPENSES (Dépenses avec catégories et photos de reçus)
# ============================================================
VALID_EXPENSE_CATEGORIES = [
    "gas", "resto", "resin", "equipement", "reparation", "communication", "publicite"
]


class ExpenseCreate(BaseModel):
    amount: float
    category: str
    date: str  # ISO date (YYYY-MM-DD)
    description: Optional[str] = ""
    vendor: Optional[str] = ""
    receipt_photo: Optional[str] = None  # base64 string (data URL or raw)
    receipt_pdf: Optional[str] = None  # base64 PDF (data URL or raw) — multi-page scan


class ExpenseUpdate(BaseModel):
    amount: Optional[float] = None
    category: Optional[str] = None
    date: Optional[str] = None
    description: Optional[str] = None
    vendor: Optional[str] = None
    receipt_photo: Optional[str] = None
    receipt_pdf: Optional[str] = None


class ExpenseResponse(BaseModel):
    id: str
    amount: float
    category: str
    date: str
    description: str
    vendor: str
    receipt_photo: Optional[str] = None
    receipt_pdf: Optional[str] = None
    created_at: str
    updated_at: str


def _expense_doc_to_response(doc: dict) -> dict:
    return {
        "id": doc.get("id"),
        "amount": float(doc.get("amount", 0)),
        "category": doc.get("category", ""),
        "date": doc.get("date", ""),
        "description": doc.get("description", "") or "",
        "vendor": doc.get("vendor", "") or "",
        "receipt_photo": doc.get("receipt_photo"),
        "receipt_pdf": doc.get("receipt_pdf"),
        "created_at": doc.get("created_at", ""),
        "updated_at": doc.get("updated_at", ""),
    }


@api_router.post("/expenses", response_model=ExpenseResponse)
async def create_expense(payload: ExpenseCreate):
    if payload.amount <= 0:
        raise HTTPException(status_code=400, detail="Le montant doit être positif")
    if payload.category not in VALID_EXPENSE_CATEGORIES:
        raise HTTPException(status_code=400, detail=f"Catégorie invalide. Valides: {VALID_EXPENSE_CATEGORIES}")
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "amount": float(payload.amount),
        "category": payload.category,
        "date": payload.date,
        "description": payload.description or "",
        "vendor": payload.vendor or "",
        "receipt_photo": payload.receipt_photo,
        "receipt_pdf": payload.receipt_pdf,
        "created_at": now,
        "updated_at": now,
    }
    await db.expenses.insert_one(doc)
    return ExpenseResponse(**_expense_doc_to_response(doc))


@api_router.get("/expenses", response_model=List[ExpenseResponse])
async def list_expenses(category: Optional[str] = None, start_date: Optional[str] = None, end_date: Optional[str] = None, limit: int = 500):
    query: dict = {}
    if category:
        query["category"] = category
    if start_date or end_date:
        date_q = {}
        if start_date:
            date_q["$gte"] = start_date
        if end_date:
            date_q["$lte"] = end_date
        query["date"] = date_q
    cursor = db.expenses.find(query, {"_id": 0}).sort("date", -1).limit(max(1, min(limit, 2000)))
    items = await cursor.to_list(2000)
    return [ExpenseResponse(**_expense_doc_to_response(it)) for it in items]


@api_router.get("/expenses/stats")
async def expenses_stats(start_date: Optional[str] = None, end_date: Optional[str] = None):
    """Returns total amount per category and grand total."""
    match_q: dict = {}
    if start_date or end_date:
        date_q = {}
        if start_date:
            date_q["$gte"] = start_date
        if end_date:
            date_q["$lte"] = end_date
        match_q["date"] = date_q

    pipeline = [
        {"$match": match_q} if match_q else {"$match": {}},
        {"$group": {"_id": "$category", "total": {"$sum": "$amount"}, "count": {"$sum": 1}}},
    ]
    cursor = db.expenses.aggregate(pipeline)
    by_category = {}
    grand_total = 0.0
    async for row in cursor:
        cat = row.get("_id") or "autre"
        total = float(row.get("total", 0))
        by_category[cat] = {"total": total, "count": row.get("count", 0)}
        grand_total += total

    # Fill missing categories with 0
    for c in VALID_EXPENSE_CATEGORIES:
        if c not in by_category:
            by_category[c] = {"total": 0.0, "count": 0}

    return {"by_category": by_category, "grand_total": grand_total}


@api_router.get("/expenses/{expense_id}", response_model=ExpenseResponse)
async def get_expense(expense_id: str):
    doc = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Dépense introuvable")
    return ExpenseResponse(**_expense_doc_to_response(doc))


@api_router.put("/expenses/{expense_id}", response_model=ExpenseResponse)
async def update_expense(expense_id: str, payload: ExpenseUpdate):
    # Allow explicit null for receipt fields (so user can delete an attached receipt)
    ALLOW_NULL_FIELDS = {"receipt_photo", "receipt_pdf", "description", "vendor"}
    raw = payload.dict(exclude_unset=True)
    update = {}
    for k, v in raw.items():
        if v is None and k not in ALLOW_NULL_FIELDS:
            continue
        update[k] = v
    if "category" in update and update["category"] not in VALID_EXPENSE_CATEGORIES:
        raise HTTPException(status_code=400, detail="Catégorie invalide")
    if not update:
        raise HTTPException(status_code=400, detail="Aucune modification")
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    res = await db.expenses.update_one({"id": expense_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dépense introuvable")
    updated = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    return ExpenseResponse(**_expense_doc_to_response(updated))


@api_router.delete("/expenses/{expense_id}/receipt")
async def delete_expense_receipt(expense_id: str, type: str = "all"):
    """Delete the attached receipt from an expense without deleting the expense itself.
    Query param `type`: 'photo' | 'pdf' | 'all' (default 'all').
    """
    if type not in {"photo", "pdf", "all"}:
        raise HTTPException(status_code=400, detail="Type invalide (photo|pdf|all)")
    doc = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Dépense introuvable")
    unset = {}
    if type in ("photo", "all"):
        unset["receipt_photo"] = None
    if type in ("pdf", "all"):
        unset["receipt_pdf"] = None
    unset["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.expenses.update_one({"id": expense_id}, {"$set": unset})
    updated = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    return {
        "deleted": type,
        "expense": _expense_doc_to_response(updated),
    }


@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str):
    res = await db.expenses.delete_one({"id": expense_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dépense introuvable")
    return {"deleted": 1}


class ImagesToPdfRequest(BaseModel):
    images: List[str]  # list of base64 image data URIs or raw base64


@api_router.post("/expenses/images-to-pdf")
async def convert_images_to_pdf(payload: ImagesToPdfRequest):
    """Convert one or multiple base64 images into a single PDF.
    Each image becomes a page in the PDF. Returns the PDF as base64 data URL.
    Used for scanning multi-page receipts/invoices.
    """
    from PIL import Image
    from io import BytesIO

    if not payload.images or len(payload.images) == 0:
        raise HTTPException(status_code=400, detail="Aucune image fournie")
    if len(payload.images) > 20:
        raise HTTPException(status_code=400, detail="Maximum 20 pages par PDF")

    pil_images = []
    try:
        for idx, img_str in enumerate(payload.images):
            # Handle data URL prefix
            if img_str.startswith("data:"):
                img_str = img_str.split(",", 1)[1] if "," in img_str else img_str
            try:
                raw = base64.b64decode(img_str)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Image {idx+1} invalide: {e}")
            img = Image.open(BytesIO(raw))
            # Convert to RGB (PDF doesn't support RGBA/transparency)
            if img.mode in ("RGBA", "LA", "P"):
                bg = Image.new("RGB", img.size, (255, 255, 255))
                if img.mode == "P":
                    img = img.convert("RGBA")
                bg.paste(img, mask=img.split()[-1] if img.mode in ("RGBA", "LA") else None)
                img = bg
            elif img.mode != "RGB":
                img = img.convert("RGB")

            # Resize very large images to fit within A4-ish dimensions (keeps file size reasonable)
            max_dim = 2200
            if img.width > max_dim or img.height > max_dim:
                ratio = min(max_dim / img.width, max_dim / img.height)
                new_size = (int(img.width * ratio), int(img.height * ratio))
                img = img.resize(new_size, Image.LANCZOS)

            pil_images.append(img)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Image processing error: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur traitement image: {e}")

    # Build PDF in memory
    buf = BytesIO()
    try:
        first = pil_images[0]
        rest = pil_images[1:] if len(pil_images) > 1 else []
        first.save(buf, format="PDF", save_all=True, append_images=rest, resolution=150.0)
        pdf_bytes = buf.getvalue()
    except Exception as e:
        logger.error(f"PDF generation error: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur génération PDF: {e}")
    finally:
        buf.close()

    pdf_b64 = base64.b64encode(pdf_bytes).decode("ascii")
    pdf_data_url = f"data:application/pdf;base64,{pdf_b64}"
    return {
        "pdf_base64": pdf_data_url,
        "pages": len(pil_images),
        "size_kb": round(len(pdf_bytes) / 1024, 1),
    }


@api_router.get("/expenses/{expense_id}/receipt-pdf")
async def download_expense_receipt_pdf(expense_id: str):
    """Download the PDF receipt/invoice attached to an expense (inline view)."""
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    doc = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Dépense introuvable")
    pdf_str = doc.get("receipt_pdf")
    if not pdf_str:
        raise HTTPException(status_code=404, detail="Aucun PDF attaché")
    # Strip data URL prefix if present
    if pdf_str.startswith("data:"):
        pdf_str = pdf_str.split(",", 1)[1] if "," in pdf_str else pdf_str
    try:
        raw = base64.b64decode(pdf_str)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF corrompu: {e}")
    vendor = (doc.get("vendor") or "recu").replace(" ", "_")[:40]
    date_str = doc.get("date", "")
    filename = f"Recu_{vendor}_{date_str}.pdf"
    return StreamingResponse(
        BytesIO(raw),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


class OcrReceiptRequest(BaseModel):
    images: List[str]  # list of base64 images (data URL or raw), max 10


@api_router.post("/expenses/ocr-receipt")
async def ocr_receipt(payload: OcrReceiptRequest):
    """Extract structured data from scanned receipt/invoice images using Gemini Vision.
    Returns: { amount, vendor, date, description, raw_text, confidence }
    Uses Emergent LLM Key + Gemini 2.5 Flash (fast + accurate for receipts).
    """
    if not payload.images or len(payload.images) == 0:
        raise HTTPException(status_code=400, detail="Aucune image fournie")
    if len(payload.images) > 10:
        raise HTTPException(status_code=400, detail="Maximum 10 pages par OCR (limite LLM)")

    # Lazy imports
    from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent

    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise HTTPException(
            status_code=500,
            detail="Clé LLM non configurée (EMERGENT_LLM_KEY)",
        )

    # Normalize images to raw base64 (strip data URL prefix if present)
    raw_images: List[str] = []
    for img_str in payload.images:
        if img_str.startswith("data:"):
            if "," in img_str:
                img_str = img_str.split(",", 1)[1]
        raw_images.append(img_str)

    system_prompt = (
        "Tu es un expert en extraction de données depuis des reçus et factures. "
        "Analyse l'image fournie (ou les images si plusieurs pages) et retourne UNIQUEMENT un JSON valide avec ces clés EXACTES:\n"
        "{\n"
        '  "amount": <float|null>,   // Montant TOTAL final (TTC), en dollars. Null si illisible.\n'
        '  "vendor": <string|null>,  // Nom du commerce/fournisseur (ex: "Canadian Tire", "Costco"). Null si non visible.\n'
        '  "date": <string|null>,    // Date au format YYYY-MM-DD. Null si illisible.\n'
        '  "description": <string|null>,  // Résumé court des articles principaux (1-2 phrases).\n'
        '  "raw_text": <string>,     // Transcription complète du texte visible sur le(s) reçu(s), ligne par ligne.\n'
        '  "confidence": <float>     // 0.0 à 1.0 — ton niveau de confiance global dans l extraction.\n'
        "}\n\n"
        "RÈGLES STRICTES:\n"
        "- Retourne UNIQUEMENT le JSON, aucun texte avant ou après, pas de markdown ``` ni explication.\n"
        "- Pour le montant, cherche TOTAL, GRAND TOTAL, MONTANT DÛ — pas le sous-total.\n"
        "- Gère FR et EN indifféremment (reçus du Québec typiquement bilingues).\n"
        "- Date: convertis TOUT format (DD/MM/YYYY, MM-DD-YY, 23 avr. 2026, etc.) en YYYY-MM-DD.\n"
        "- Si plusieurs images: combine l information (page 1 + page 2 = même reçu).\n"
        "- Si illisible/pas un reçu: mets null partout sauf raw_text et confidence=0.0.\n"
    )

    chat = LlmChat(
        api_key=api_key,
        session_id=f"ocr-receipt-{uuid.uuid4().hex[:8]}",
        system_message=system_prompt,
    ).with_model("gemini", "gemini-2.5-flash")

    # Build message with all images
    image_contents = [ImageContent(image_base64=img) for img in raw_images]
    user_msg = UserMessage(
        text=(
            "Extrais les informations structurées de ce reçu/facture. "
            "Retourne uniquement le JSON demandé (voir instructions système)."
        ),
        file_contents=image_contents,
    )

    try:
        response_text = await chat.send_message(user_msg)
    except Exception as e:
        logger.error(f"Gemini OCR error: {e}")
        raise HTTPException(status_code=502, detail=f"Erreur OCR (LLM): {str(e)[:200]}")

    # Parse JSON response (handle markdown code fences just in case)
    import json as _json
    import re as _re
    cleaned = response_text.strip()
    # Strip markdown code fences if model added them despite instructions
    fence_match = _re.search(r"```(?:json)?\s*(\{.*?\})\s*```", cleaned, _re.DOTALL)
    if fence_match:
        cleaned = fence_match.group(1)
    # Extract first JSON object if response has extra text
    obj_match = _re.search(r"\{.*\}", cleaned, _re.DOTALL)
    if obj_match:
        cleaned = obj_match.group(0)

    try:
        parsed = _json.loads(cleaned)
    except Exception as e:
        logger.warning(f"OCR JSON parse error. Raw response: {response_text[:500]}")
        # Fallback: return raw text only
        return {
            "amount": None,
            "vendor": None,
            "date": None,
            "description": None,
            "raw_text": response_text[:4000],
            "confidence": 0.0,
            "parse_error": str(e)[:200],
        }

    # Normalize output
    def _safe_float(v):
        if v is None:
            return None
        try:
            return float(v)
        except Exception:
            return None

    def _safe_str(v):
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    return {
        "amount": _safe_float(parsed.get("amount")),
        "vendor": _safe_str(parsed.get("vendor")),
        "date": _safe_str(parsed.get("date")),
        "description": _safe_str(parsed.get("description")),
        "raw_text": _safe_str(parsed.get("raw_text")) or "",
        "confidence": _safe_float(parsed.get("confidence")) or 0.0,
    }


@api_router.get("/expenses/export/excel")
async def export_expenses_excel(start_date: Optional[str] = None, end_date: Optional[str] = None, category: Optional[str] = None):
    """Generates an Excel (.xlsx) export of all expenses with summary sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    # Build query
    query: dict = {}
    if category:
        query["category"] = category
    if start_date or end_date:
        d = {}
        if start_date: d["$gte"] = start_date
        if end_date: d["$lte"] = end_date
        query["date"] = d

    cursor = db.expenses.find(query, {"_id": 0}).sort("date", -1).allow_disk_use(True)
    items = await cursor.to_list(10000)

    CATEGORY_LABELS = {
        "gas": "⛽ Essence",
        "resto": "🍽️ Resto",
        "resin": "🧪 Résine",
        "equipement": "🔧 Équipement",
        "reparation": "🛠️ Réparation",
        "communication": "📞 Communication",
        "publicite": "📢 Publicité",
    }

    wb = Workbook()
    # ---- Sheet 1: Détails ----
    ws = wb.active
    ws.title = "Dépenses"

    header = ["Date", "Catégorie", "Montant ($)", "Commerce", "Description"]
    ws.append(header)

    header_fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    total = 0.0
    totals_by_cat = {k: 0.0 for k in CATEGORY_LABELS.keys()}
    for exp in items:
        row = [
            exp.get("date", ""),
            CATEGORY_LABELS.get(exp.get("category", ""), exp.get("category", "")),
            float(exp.get("amount", 0)),
            exp.get("vendor", "") or "",
            exp.get("description", "") or "",
        ]
        ws.append(row)
        amt = float(exp.get("amount", 0))
        total += amt
        cat = exp.get("category", "")
        if cat in totals_by_cat:
            totals_by_cat[cat] += amt

    # Format currency column
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).number_format = '"$"#,##0.00'
        for col in range(1, len(header) + 1):
            ws.cell(row=row_idx, column=col).border = border

    # Total row
    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True, size=14)
    total_cell = ws.cell(row=total_row, column=3, value=total)
    total_cell.font = Font(bold=True, color="10B981", size=14)
    total_cell.number_format = '"$"#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 40

    # ---- Sheet 2: Résumé par catégorie ----
    ws2 = wb.create_sheet("Résumé")
    ws2.append(["Catégorie", "Total ($)", "Nombre"])
    for col in range(1, 4):
        c = ws2.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
        c.border = border

    for cat_key, label in CATEGORY_LABELS.items():
        amt = totals_by_cat.get(cat_key, 0)
        count = sum(1 for e in items if e.get("category") == cat_key)
        ws2.append([label, amt, count])

    # Format
    for row_idx in range(2, ws2.max_row + 1):
        ws2.cell(row=row_idx, column=2).number_format = '"$"#,##0.00'
        for col in range(1, 4):
            ws2.cell(row=row_idx, column=col).border = border

    # Grand total row on Résumé
    ws2.append([])
    summary_total_row = ws2.max_row + 1
    ws2.cell(row=summary_total_row, column=1, value="GRAND TOTAL").font = Font(bold=True, size=14)
    gt = ws2.cell(row=summary_total_row, column=2, value=total)
    gt.font = Font(bold=True, color="10B981", size=14)
    gt.number_format = '"$"#,##0.00'

    ws2.column_dimensions['A'].width = 26
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 12

    # Output
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"depenses_crystaltask_{today}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============================================================
# REVENUES (Revenus / Encaissements)
# ============================================================
VALID_REVENUE_CATEGORIES = [
    "printemps", "automne"
]
VALID_PAYMENT_METHODS = ["etransfert", "cash", "cheque", "credit"]


class RevenueCreate(BaseModel):
    amount: float
    category: str
    date: str  # YYYY-MM-DD
    description: Optional[str] = ""
    client_name: Optional[str] = ""
    payment_method: Optional[str] = "cash"
    appointment_id: Optional[str] = None


class RevenueUpdate(BaseModel):
    amount: Optional[float] = None
    category: Optional[str] = None
    date: Optional[str] = None
    description: Optional[str] = None
    client_name: Optional[str] = None
    payment_method: Optional[str] = None
    appointment_id: Optional[str] = None


class RevenueResponse(BaseModel):
    id: str
    amount: float
    category: str
    date: str
    description: str
    client_name: str
    payment_method: str
    appointment_id: Optional[str] = None
    created_at: str
    updated_at: str


def _revenue_doc_to_response(doc: dict) -> dict:
    return {
        "id": doc.get("id"),
        "amount": float(doc.get("amount", 0)),
        "category": doc.get("category", ""),
        "date": doc.get("date", ""),
        "description": doc.get("description", "") or "",
        "client_name": doc.get("client_name", "") or "",
        "payment_method": doc.get("payment_method", "cash") or "cash",
        "appointment_id": doc.get("appointment_id"),
        "created_at": doc.get("created_at", ""),
        "updated_at": doc.get("updated_at", ""),
    }


@api_router.post("/revenues", response_model=RevenueResponse)
async def create_revenue(payload: RevenueCreate):
    if payload.amount <= 0:
        raise HTTPException(status_code=400, detail="Le montant doit être positif")
    if payload.category not in VALID_REVENUE_CATEGORIES:
        raise HTTPException(status_code=400, detail=f"Catégorie invalide. Valides: {VALID_REVENUE_CATEGORIES}")
    pm = payload.payment_method or "cash"
    if pm not in VALID_PAYMENT_METHODS:
        raise HTTPException(status_code=400, detail=f"Mode de paiement invalide. Valides: {VALID_PAYMENT_METHODS}")
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "amount": float(payload.amount),
        "category": payload.category,
        "date": payload.date,
        "description": payload.description or "",
        "client_name": payload.client_name or "",
        "payment_method": pm,
        "appointment_id": payload.appointment_id,
        "created_at": now,
        "updated_at": now,
    }
    await db.revenues.insert_one(doc)
    return RevenueResponse(**_revenue_doc_to_response(doc))


@api_router.get("/revenues", response_model=List[RevenueResponse])
async def list_revenues(category: Optional[str] = None, start_date: Optional[str] = None, end_date: Optional[str] = None, limit: int = 500):
    query: dict = {}
    if category:
        query["category"] = category
    if start_date or end_date:
        date_q = {}
        if start_date: date_q["$gte"] = start_date
        if end_date: date_q["$lte"] = end_date
        query["date"] = date_q
    cursor = db.revenues.find(query, {"_id": 0}).sort("date", -1).limit(max(1, min(limit, 2000)))
    items = await cursor.to_list(2000)
    return [RevenueResponse(**_revenue_doc_to_response(it)) for it in items]


@api_router.get("/revenues/stats")
async def revenues_stats(start_date: Optional[str] = None, end_date: Optional[str] = None):
    match_q: dict = {}
    if start_date or end_date:
        date_q = {}
        if start_date: date_q["$gte"] = start_date
        if end_date: date_q["$lte"] = end_date
        match_q["date"] = date_q
    pipeline = [
        {"$match": match_q} if match_q else {"$match": {}},
        {"$group": {"_id": "$category", "total": {"$sum": "$amount"}, "count": {"$sum": 1}}},
    ]
    cursor = db.revenues.aggregate(pipeline)
    by_category = {}
    grand_total = 0.0
    async for row in cursor:
        cat = row.get("_id") or "autre"
        total = float(row.get("total", 0))
        by_category[cat] = {"total": total, "count": row.get("count", 0)}
        grand_total += total
    for c in VALID_REVENUE_CATEGORIES:
        if c not in by_category:
            by_category[c] = {"total": 0.0, "count": 0}
    # Also return payment-method breakdown
    pm_pipeline = [
        {"$match": match_q} if match_q else {"$match": {}},
        {"$group": {"_id": "$payment_method", "total": {"$sum": "$amount"}, "count": {"$sum": 1}}},
    ]
    pm_cursor = db.revenues.aggregate(pm_pipeline)
    by_payment = {}
    async for row in pm_cursor:
        pm = row.get("_id") or "cash"
        by_payment[pm] = {"total": float(row.get("total", 0)), "count": row.get("count", 0)}
    for pm in VALID_PAYMENT_METHODS:
        if pm not in by_payment:
            by_payment[pm] = {"total": 0.0, "count": 0}
    return {"by_category": by_category, "by_payment": by_payment, "grand_total": grand_total}


@api_router.get("/revenues/{revenue_id}", response_model=RevenueResponse)
async def get_revenue(revenue_id: str):
    doc = await db.revenues.find_one({"id": revenue_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Revenu introuvable")
    return RevenueResponse(**_revenue_doc_to_response(doc))


@api_router.put("/revenues/{revenue_id}", response_model=RevenueResponse)
async def update_revenue(revenue_id: str, payload: RevenueUpdate):
    update = {k: v for k, v in payload.dict(exclude_unset=True).items() if v is not None}
    if "category" in update and update["category"] not in VALID_REVENUE_CATEGORIES:
        raise HTTPException(status_code=400, detail="Catégorie invalide")
    if "payment_method" in update and update["payment_method"] not in VALID_PAYMENT_METHODS:
        raise HTTPException(status_code=400, detail="Mode de paiement invalide")
    if not update:
        raise HTTPException(status_code=400, detail="Aucune modification")
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    res = await db.revenues.update_one({"id": revenue_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Revenu introuvable")
    updated = await db.revenues.find_one({"id": revenue_id}, {"_id": 0})
    return RevenueResponse(**_revenue_doc_to_response(updated))


@api_router.delete("/revenues/{revenue_id}")
async def delete_revenue(revenue_id: str):
    res = await db.revenues.delete_one({"id": revenue_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Revenu introuvable")
    return {"deleted": 1}


# --- Encaisser (Collect payment for an appointment) ---
@api_router.post("/appointments/{appointment_id}/encaisser")
async def encaisser_appointment(appointment_id: str, payload: dict = Body(...)):
    """
    Collect payment for an appointment.
    Atomically:
      1. Creates a Revenue document linked to this appointment
      2. Marks the appointment as status='paid' with paid_at, paid_amount, paid_method
    Body:
      amount (float, required >0), payment_method ('cash'|'etransfert'),
      category ('printemps'|'automne'), date ('YYYY-MM-DD'),
      description (str, optional)
    """
    appt = await db.appointments.find_one({"id": appointment_id})
    if not appt:
        raise HTTPException(status_code=404, detail="Appointment not found")

    try:
        amount = float(payload.get("amount", 0) or 0)
    except Exception:
        amount = 0.0
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Le montant doit être positif")

    pm = (payload.get("payment_method") or "cash").lower()
    if pm not in ("cash", "etransfert"):
        raise HTTPException(status_code=400, detail="Mode de paiement: cash ou etransfert")

    category = (payload.get("category") or "printemps").lower()
    if category not in VALID_REVENUE_CATEGORIES:
        raise HTTPException(status_code=400, detail=f"Catégorie invalide. Valides: {VALID_REVENUE_CATEGORIES}")

    date_str = payload.get("date") or datetime.now(timezone.utc).date().isoformat()
    address = appt.get("client_address", "") or ""
    default_desc = f"Lavage de vitres - {address}".strip(" -")
    description = (payload.get("description") or default_desc).strip()

    now = datetime.now(timezone.utc).isoformat()

    # Create revenue document
    rev_doc = {
        "id": str(uuid.uuid4()),
        "amount": amount,
        "category": category,
        "date": date_str,
        "description": description,
        "client_name": appt.get("client_name", "") or "",
        "payment_method": pm,
        "appointment_id": appointment_id,
        "created_at": now,
        "updated_at": now,
    }
    await db.revenues.insert_one(rev_doc)

    # Mark appointment as paid
    await db.appointments.update_one(
        {"id": appointment_id},
        {"$set": {
            "status": "paid",
            "paid_at": now,
            "paid_amount": amount,
            "paid_method": pm,
            "revenue_id": rev_doc["id"],
        }}
    )

    return {
        "ok": True,
        "revenue": _revenue_doc_to_response(rev_doc),
        "appointment_id": appointment_id,
        "status": "paid",
        "paid_at": now,
        "paid_amount": amount,
        "paid_method": pm,
    }




@api_router.get("/revenues/export/excel")
async def export_revenues_excel(start_date: Optional[str] = None, end_date: Optional[str] = None, category: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    query: dict = {}
    if category:
        query["category"] = category
    if start_date or end_date:
        d = {}
        if start_date: d["$gte"] = start_date
        if end_date: d["$lte"] = end_date
        query["date"] = d

    cursor = db.revenues.find(query, {"_id": 0}).sort("date", -1)
    items = await cursor.to_list(10000)

    CATEGORY_LABELS = {
        "printemps": "🌸 Saison Printemps",
        "automne": "🍂 Saison Automne",
    }
    PAYMENT_LABELS = {
        "etransfert": "📱 E-transfert",
        "cash": "💵 Cash",
        "cheque": "📝 Chèque",
        "credit": "💳 Carte de crédit",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Revenus"

    header = ["Date", "Catégorie", "Montant ($)", "Client", "Paiement", "Description"]
    ws.append(header)
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'),
    )
    for col in range(1, len(header) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border

    total = 0.0
    totals_by_cat = {k: 0.0 for k in CATEGORY_LABELS}
    for rev in items:
        row = [
            rev.get("date", ""),
            CATEGORY_LABELS.get(rev.get("category", ""), rev.get("category", "")),
            float(rev.get("amount", 0)),
            rev.get("client_name", "") or "",
            PAYMENT_LABELS.get(rev.get("payment_method", "cash"), rev.get("payment_method", "")),
            rev.get("description", "") or "",
        ]
        ws.append(row)
        amt = float(rev.get("amount", 0))
        total += amt
        cat = rev.get("category", "")
        if cat in totals_by_cat:
            totals_by_cat[cat] += amt

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).number_format = '"$"#,##0.00'
        for col in range(1, len(header) + 1):
            ws.cell(row=row_idx, column=col).border = border

    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True, size=14)
    tc = ws.cell(row=total_row, column=3, value=total)
    tc.font = Font(bold=True, color="10B981", size=14)
    tc.number_format = '"$"#,##0.00'

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 26
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 40

    # Summary sheet
    ws2 = wb.create_sheet("Résumé")
    ws2.append(["Catégorie", "Total ($)", "Nombre"])
    for col in range(1, 4):
        c = ws2.cell(row=1, column=col)
        c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal="center"); c.border = border
    for key, label in CATEGORY_LABELS.items():
        amt = totals_by_cat.get(key, 0)
        count = sum(1 for r in items if r.get("category") == key)
        ws2.append([label, amt, count])
    for row_idx in range(2, ws2.max_row + 1):
        ws2.cell(row=row_idx, column=2).number_format = '"$"#,##0.00'
        for col in range(1, 4):
            ws2.cell(row=row_idx, column=col).border = border
    ws2.append([])
    stot_row = ws2.max_row + 1
    ws2.cell(row=stot_row, column=1, value="GRAND TOTAL").font = Font(bold=True, size=14)
    gt = ws2.cell(row=stot_row, column=2, value=total)
    gt.font = Font(bold=True, color="10B981", size=14)
    gt.number_format = '"$"#,##0.00'
    ws2.column_dimensions['A'].width = 26
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 12

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"revenus_crystaltask_{today}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============================================================
# BILAN FINANCIER (Finance balance = Revenus - Dépenses)
# ============================================================
@api_router.get("/finance/bilan")
async def finance_bilan(start_date: Optional[str] = None, end_date: Optional[str] = None):
    """Returns total revenues, total expenses, net profit, and breakdown for a period."""
    q_exp: dict = {}
    q_rev: dict = {}
    if start_date or end_date:
        d = {}
        if start_date: d["$gte"] = start_date
        if end_date: d["$lte"] = end_date
        q_exp["date"] = dict(d)
        q_rev["date"] = dict(d)

    # Expenses
    exp_pipeline = [
        {"$match": q_exp} if q_exp else {"$match": {}},
        {"$group": {"_id": "$category", "total": {"$sum": "$amount"}, "count": {"$sum": 1}}},
    ]
    rev_pipeline = [
        {"$match": q_rev} if q_rev else {"$match": {}},
        {"$group": {"_id": "$category", "total": {"$sum": "$amount"}, "count": {"$sum": 1}}},
    ]

    exp_by_cat = {}
    total_expenses = 0.0
    async for row in db.expenses.aggregate(exp_pipeline):
        cat = row.get("_id") or "autre"
        total = float(row.get("total", 0))
        exp_by_cat[cat] = {"total": total, "count": row.get("count", 0)}
        total_expenses += total

    rev_by_cat = {}
    total_revenues = 0.0
    async for row in db.revenues.aggregate(rev_pipeline):
        cat = row.get("_id") or "autre"
        total = float(row.get("total", 0))
        rev_by_cat[cat] = {"total": total, "count": row.get("count", 0)}
        total_revenues += total

    net_profit = total_revenues - total_expenses
    margin = (net_profit / total_revenues * 100.0) if total_revenues > 0 else 0.0

    return {
        "period": {"start_date": start_date, "end_date": end_date},
        "total_revenues": round(total_revenues, 2),
        "total_expenses": round(total_expenses, 2),
        "net_profit": round(net_profit, 2),
        "margin_pct": round(margin, 2),
        "revenues_by_category": rev_by_cat,
        "expenses_by_category": exp_by_cat,
    }


# ============================================================
# SCHEDULED CAMPAIGNS
# ============================================================
class ScheduledCampaignCreate(BaseModel):
    season: str  # spring | autumn | summer
    subject: str
    body: str
    recipients: List[str]  # list of emails (BCC)
    scheduled_at: str  # ISO datetime
    locale: Optional[str] = "fr"


class ScheduledCampaignResponse(BaseModel):
    id: str
    season: str
    subject: str
    body: str
    recipients: List[str]
    scheduled_at: str
    status: str  # pending | sent | ready | failed | cancelled
    error: Optional[str] = None
    sent_at: Optional[str] = None
    created_at: str


@api_router.post("/campaigns/preview-html", response_class=HTMLResponse)
async def preview_campaign_html(payload: dict):
    """Render the rich HTML campaign email (as it will appear in recipients' inbox).
    Accepts {body: str, subject: str}. Returns a full HTML page for preview."""
    body = payload.get("body", "") or ""
    subject = payload.get("subject", "Aperçu campagne") or "Aperçu campagne"
    html = _build_seasonal_campaign_html(body, subject)
    return HTMLResponse(content=html)


@api_router.post("/campaigns/send-now")
async def send_campaign_now(payload: dict = Body(...)):
    """Send a marketing campaign IMMEDIATELY via Resend, with rich HTML rendering
    (logo + QR code + clickable links + seasonal accents).

    Body: { subject: str, body: str, recipients: [emails] }

    Returns: { sent: bool, count: int, skipped: int, resend_id: str|None }
    """
    subject = (payload.get("subject") or "").strip()
    body = (payload.get("body") or "").strip()
    recipients_raw = payload.get("recipients") or []
    if not subject or not body:
        raise HTTPException(status_code=400, detail="Sujet et corps requis")
    if not recipients_raw:
        raise HTTPException(status_code=400, detail="Aucun destinataire")
    if not resend.api_key:
        raise HTTPException(status_code=503, detail="Service email non configuré")

    # CASL: filter unsubscribed
    unsubbed_set = set()
    async for u in db.unsubscribes.find({}, {"email": 1, "_id": 0}):
        unsubbed_set.add((u.get("email") or "").lower())
    recipients = [r for r in recipients_raw if r and r.lower() not in unsubbed_set]
    skipped = len(recipients_raw) - len(recipients)
    if not recipients:
        raise HTTPException(status_code=400, detail="Tous les destinataires sont désabonnés")

    # Build the rich HTML (logo + QR + season-themed)
    html_body = _build_seasonal_campaign_html(body, subject)
    html_final = inject_branding(html_body)
    from_addr = os.environ.get("RESEND_FROM") or "onboarding@resend.dev"

    try:
        result = await asyncio.to_thread(
            resend.Emails.send,
            {
                "from": from_addr,
                "to": ["onboarding@resend.dev"],  # placeholder (Resend requires 'to')
                "bcc": recipients,
                "subject": subject,
                "html": html_final,
            },
        )
        sent_id = (result or {}).get("id", "")
        logger.info(f"Campaign sent NOW to {len(recipients)} recipients (skipped {skipped} unsubbed) — resend_id={sent_id}")
        return {
            "sent": True,
            "count": len(recipients),
            "skipped": skipped,
            "resend_id": sent_id,
        }
    except Exception as e:
        err_msg = str(e)
        logger.error(f"Campaign send-now failed: {err_msg}")
        raise HTTPException(status_code=502, detail=f"Échec d'envoi: {err_msg[:200]}")


@api_router.post("/scheduled-campaigns", response_model=ScheduledCampaignResponse)
async def create_scheduled_campaign(payload: ScheduledCampaignCreate):
    """Plan a campaign to be sent at a future date/time."""
    if not payload.recipients:
        raise HTTPException(status_code=400, detail="Aucun destinataire")
    try:
        when = datetime.fromisoformat(payload.scheduled_at.replace('Z', '+00:00'))
    except Exception:
        raise HTTPException(status_code=400, detail="Format de date invalide (ISO 8601 attendu)")
    now = datetime.now(timezone.utc)
    if when.tzinfo is None:
        when = when.replace(tzinfo=timezone.utc)
    if when <= now:
        raise HTTPException(status_code=400, detail="La date doit être dans le futur")

    doc = {
        "id": str(uuid.uuid4()),
        "season": payload.season,
        "subject": payload.subject,
        "body": payload.body,
        "recipients": payload.recipients,
        "scheduled_at": when.isoformat(),
        "status": "pending",
        "locale": payload.locale or "fr",
        "created_at": now.isoformat(),
    }
    await db.scheduled_campaigns.insert_one(doc)
    doc.pop("_id", None)
    return ScheduledCampaignResponse(**doc)


@api_router.get("/scheduled-campaigns", response_model=List[ScheduledCampaignResponse])
async def list_scheduled_campaigns(status: Optional[str] = None):
    query = {}
    if status:
        query["status"] = status
    cursor = db.scheduled_campaigns.find(query, {"_id": 0}).sort("scheduled_at", 1).limit(500)
    items = await cursor.to_list(500)
    return [ScheduledCampaignResponse(**{k: v for k, v in c.items() if k != "locale"}) for c in items]


@api_router.delete("/scheduled-campaigns/{campaign_id}")
async def cancel_scheduled_campaign(campaign_id: str):
    res = await db.scheduled_campaigns.update_one(
        {"id": campaign_id, "status": {"$in": ["pending", "ready"]}},
        {"$set": {"status": "cancelled"}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Campagne introuvable ou déjà envoyée")
    return {"cancelled": 1}


@api_router.post("/scheduled-campaigns/{campaign_id}/mark-sent")
async def mark_campaign_sent_manually(campaign_id: str):
    """Called by the app when user has manually sent a 'ready' campaign via mailto:"""
    now = datetime.now(timezone.utc).isoformat()
    res = await db.scheduled_campaigns.update_one(
        {"id": campaign_id},
        {"$set": {"status": "sent", "sent_at": now}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Campagne introuvable")
    return {"ok": True}


def _load_asset_base64(filename: str, mime: str = "image/jpeg") -> str:
    """Load an asset file and return it as a base64 data URI for inline email embedding."""
    try:
        p = ROOT_DIR / "assets" / filename
        if not p.exists():
            return ""
        with open(p, "rb") as f:
            raw = f.read()
        b64 = base64.b64encode(raw).decode("ascii")
        return f"data:{mime};base64,{b64}"
    except Exception as e:
        logger.warning(f"Failed to load asset {filename}: {e}")
        return ""


def _build_seasonal_campaign_html(plain_body: str, subject: str) -> str:
    """Convert the plain text seasonal campaign body into a rich HTML email with:
    - Large watermark logo in the background
    - Clickable links (auto-linkify Lavagedevitre.org, phone numbers)
    - Embedded QR code for booking
    - Clean typography and spacing
    - Seasonal color accents (auto-detected from subject emojis)
    """
    logo_data_uri = _load_asset_base64("company-logo.jpeg", "image/jpeg")
    qr_data_uri = _load_asset_base64("booking_qr.jpeg", "image/jpeg")

    # Auto-detect season from subject or body emojis for themed colors
    text_for_detect = (subject or "") + " " + (plain_body or "")
    if "🌷" in text_for_detect or "printemps" in text_for_detect.lower():
        season = "spring"
        accent_gradient = "linear-gradient(90deg,#10B981,#34D399,#A7F3D0)"  # Vert printemps
        accent_solid = "#10B981"
        season_label = "🌷 Printemps"
    elif "☀️" in text_for_detect or "été" in text_for_detect.lower() or "summer" in text_for_detect.lower():
        season = "summer"
        accent_gradient = "linear-gradient(90deg,#F59E0B,#FCD34D,#FDE68A)"  # Jaune/doré été
        accent_solid = "#F59E0B"
        season_label = "☀️ Été"
    elif "🍂" in text_for_detect or "automne" in text_for_detect.lower():
        season = "autumn"
        accent_gradient = "linear-gradient(90deg,#D97706,#F97316,#FED7AA)"  # Orange/brun automne
        accent_solid = "#D97706"
        season_label = "🍂 Automne"
    else:
        season = "default"
        accent_gradient = "linear-gradient(90deg,#0891B2,#06B6D4,#22D3EE)"  # Cyan par défaut
        accent_solid = "#0891B2"
        season_label = ""

    # Escape HTML-sensitive chars first
    safe = plain_body.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    # Auto-linkify — use token placeholders to avoid double-wrapping
    import re as _re
    WEBSITE_LINK = f'<a href="https://Lavagedevitre.org" style="color:{accent_solid};text-decoration:underline;font-weight:700;">Lavagedevitre.org</a>'
    PHONE_LINK = f'<a href="tel:+15145709802" style="color:{accent_solid};text-decoration:none;font-weight:700;">514-570-9802</a>'

    # 1) Replace https://Lavagedevitre.org (with/without trailing slash) → token1
    safe = _re.sub(r'https?://Lavagedevitre\.org/?', '§§LINK_URL§§', safe, flags=_re.IGNORECASE)
    # 2) Replace bare Lavagedevitre.org → token2
    safe = _re.sub(r'\bLavagedevitre\.org\b', '§§LINK_BARE§§', safe, flags=_re.IGNORECASE)
    # 3) Replace phone → token
    safe = safe.replace("514-570-9802", '§§PHONE§§')

    # 4) Generic URL linkify (fallback for booking URLs etc.) - only those not already tokenized
    def _linkify_url(match):
        url = match.group(0)
        return f'<a href="{url}" style="color:{accent_solid};text-decoration:underline;font-weight:700;">{url}</a>'
    safe = _re.sub(r'https?://[^\s<>"\']+', _linkify_url, safe)

    # Preserve line breaks FIRST (before swapping tokens since tokens contain no newlines)
    safe = safe.replace("\n", "<br>")

    # Now swap tokens with final HTML links
    safe = safe.replace('§§LINK_URL§§', WEBSITE_LINK)
    safe = safe.replace('§§LINK_BARE§§', WEBSITE_LINK)
    safe = safe.replace('§§PHONE§§', PHONE_LINK)

    # QR section — only include if QR image loaded
    qr_section = ""
    if qr_data_uri:
        qr_section = f"""
        <div style="text-align:center;margin:24px 0 8px 0;">
          <div style="display:inline-block;padding:12px;background:#FFFFFF;border-radius:16px;box-shadow:0 4px 14px rgba(8,145,178,0.15);border:1px solid #E5E7EB;">
            <img src="{qr_data_uri}" alt="QR code pour prendre rendez-vous"
                 style="width:220px;height:220px;display:block;border-radius:8px;" />
          </div>
          <p style="margin:12px 0 0 0;font-size:13px;color:#6B7280;font-weight:600;">
            📱 Scannez pour prendre rendez-vous
          </p>
        </div>
        """

    # Seasonal ribbon (top of content)
    season_ribbon = ""
    if season_label:
        season_ribbon = f"""
        <div style="text-align:center;margin-bottom:14px;">
          <span style="display:inline-block;padding:6px 16px;background:{accent_solid};color:#FFFFFF;border-radius:999px;font-size:12px;font-weight:800;letter-spacing:0.5px;text-transform:uppercase;">
            {season_label}
          </span>
        </div>
        """

    # Logo watermark in background (enlarged, low opacity, centered)
    bg_style = ""
    if logo_data_uri:
        bg_style = (
            f"background-image:url('{logo_data_uri}');"
            f"background-repeat:no-repeat;"
            f"background-position:center center;"
            f"background-size:82% auto;"
        )

    # Full HTML email
    html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{subject}</title>
</head>
<body style="margin:0;padding:0;background:#F3F4F6;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#F3F4F6;padding:20px 0;">
    <tr>
      <td align="center">
        <table role="presentation" width="600" cellpadding="0" cellspacing="0" style="max-width:600px;width:100%;background:#FFFFFF;border-radius:16px;overflow:hidden;box-shadow:0 6px 28px rgba(0,0,0,0.08);">
          <!-- Top accent bar (seasonal) -->
          <tr>
            <td style="height:6px;background:{accent_gradient};"></td>
          </tr>

          <!-- Content with watermark logo -->
          <tr>
            <td style="padding:40px 32px;{bg_style}">
              <div style="background:rgba(255,255,255,0.82);border-radius:12px;padding:20px;">
                {season_ribbon}
                <div style="font-size:16px;line-height:1.75;color:#1F2937;">
                  {safe}
                </div>

                {qr_section}
              </div>
            </td>
          </tr>

          <!-- Footer -->
          <tr>
            <td style="padding:18px 32px;background:#0F172A;text-align:center;">
              <p style="margin:0;font-size:12px;color:#94A3B8;">
                Cet email a été envoyé par Lavage de Vitres Bois-Franc.
              </p>
              <p style="margin:4px 0 0 0;font-size:12px;color:#94A3B8;">
                🌐 <a href="https://Lavagedevitre.org" style="color:#22D3EE;text-decoration:none;">Lavagedevitre.org</a>
                &nbsp;·&nbsp;
                📞 <a href="tel:+15145709802" style="color:#22D3EE;text-decoration:none;">514-570-9802</a>
              </p>
            </td>
          </tr>
        </table>
      </td>
    </tr>
  </table>
</body>
</html>"""
    return html


async def _send_scheduled_campaign(doc: dict) -> bool:
    """Try to send a scheduled campaign via Resend.
    Returns True on success, False otherwise. Marks status accordingly."""
    cid = doc["id"]
    subject = doc["subject"]
    body = doc["body"]
    recipients_raw = doc.get("recipients") or []
    now_iso = datetime.now(timezone.utc).isoformat()

    # Filter out unsubscribed addresses (CASL compliance)
    unsubbed_set = set()
    async for u in db.unsubscribes.find({}, {"email": 1, "_id": 0}):
        unsubbed_set.add((u.get("email") or "").lower())
    recipients = [r for r in recipients_raw if r and r.lower() not in unsubbed_set]
    skipped = len(recipients_raw) - len(recipients)
    if skipped > 0:
        logger.info(f"Campaign {cid}: skipped {skipped} unsubscribed recipient(s)")

    if not recipients:
        await db.scheduled_campaigns.update_one(
            {"id": cid},
            {"$set": {"status": "failed", "error": "Aucun destinataire"}},
        )
        return False

    # Build a rich HTML email: clickable Lavagedevitre.org link, QR code, enlarged watermark logo
    html_body = _build_seasonal_campaign_html(body, subject)

    if not resend.api_key:
        # No Resend at all → mark as ready for manual sending
        await db.scheduled_campaigns.update_one(
            {"id": cid},
            {"$set": {"status": "ready", "error": "Resend non configuré — envoi manuel requis"}},
        )
        return False

    try:
        # Attempt server-side send via Resend (BCC)
        await asyncio.to_thread(
            resend.Emails.send,
            {
                "from": os.environ.get("RESEND_FROM") or "onboarding@resend.dev",
                "to": ["onboarding@resend.dev"],  # placeholder (Resend requires 'to')
                "bcc": recipients,
                "subject": subject,
                "html": inject_branding(html_body),
            },
        )
        await db.scheduled_campaigns.update_one(
            {"id": cid},
            {"$set": {"status": "sent", "sent_at": now_iso, "error": None}},
        )
        logger.info(f"Scheduled campaign {cid} sent to {len(recipients)} recipients via Resend")
        return True
    except Exception as e:
        err_msg = str(e)
        # Sandbox mode (non-verified domain) → mark 'ready' so user can send via mailto:
        if "only send testing emails" in err_msg.lower() or "verify a domain" in err_msg.lower():
            await db.scheduled_campaigns.update_one(
                {"id": cid},
                {"$set": {"status": "ready", "error": "Domaine non vérifié — envoi manuel via mailto requis"}},
            )
            logger.info(f"Scheduled campaign {cid} marked 'ready' (domain not verified)")
        else:
            await db.scheduled_campaigns.update_one(
                {"id": cid},
                {"$set": {"status": "failed", "error": err_msg[:300]}},
            )
            logger.error(f"Scheduled campaign {cid} failed: {err_msg}")
        return False


async def _process_due_campaigns():
    """Scheduler job: find pending campaigns whose time has come, try to send them."""
    try:
        now_iso = datetime.now(timezone.utc).isoformat()
        cursor = db.scheduled_campaigns.find({
            "status": "pending",
            "scheduled_at": {"$lte": now_iso},
        }, {"_id": 0})
        due_list = await cursor.to_list(100)
        if not due_list:
            return
        logger.info(f"Processing {len(due_list)} due scheduled campaign(s)")
        for doc in due_list:
            await _send_scheduled_campaign(doc)
    except Exception as e:
        logger.error(f"Scheduler _process_due_campaigns error: {e}")


# Scheduler: auto backup every day at midnight (00:00)
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from apscheduler.triggers.interval import IntervalTrigger

scheduler = AsyncIOScheduler(timezone="America/Toronto")


@app.on_event("startup")
async def start_scheduler():
    try:
        scheduler.add_job(_create_auto_backup, CronTrigger(hour=0, minute=0), id="daily_backup", replace_existing=True)
        scheduler.add_job(_process_due_campaigns, IntervalTrigger(minutes=1), id="process_campaigns", replace_existing=True)
        scheduler.add_job(
            _run_24h_reminders,
            CronTrigger(hour=9, minute=0),
            id="daily_24h_reminders",
            replace_existing=True,
        )
        scheduler.add_job(
            _run_prod_sync,
            IntervalTrigger(minutes=3),
            id="prod_sync",
            replace_existing=True,
        )
        scheduler.start()
        logger.info("Scheduler started: daily backup @ 00:00 + campaigns every minute + 24h reminders @ 09:00 ET + prod sync every 3 min")
    except Exception as e:
        logger.error(f"Scheduler failed to start: {e}")


async def _run_24h_reminders():
    """Daily 9 AM ET job: send 24h reminders to clients + summary to owner."""
    try:
        result = await reminders_module.send_24h_reminders_for_tomorrow(db)
        logger.info(f"24h reminder job done: {result}")
    except Exception as e:
        logger.error(f"24h reminder job error: {e}")


async def _run_prod_sync():
    """Every 3 min: pull new requests/appointments/clients from production DB to preview."""
    try:
        result = await prod_sync_module.run_sync(db)
        # Only log if something was added (silent otherwise)
        if any(result.values()):
            logger.info(f"prod_sync synced: {result}")
    except Exception as e:
        logger.error(f"prod_sync job error: {e}")


# Include router AT THE END so all route definitions are registered
app.include_router(api_router)
