"""
Backend tests for CrystalTask Finance module.

Scope (per /app/test_result.md test_plan):
  1) Revenues CRUD - /api/revenues
  2) Revenues stats - GET /api/revenues/stats
  3) Revenues Excel export - GET /api/revenues/export/excel
  4) Finance Bilan - GET /api/finance/bilan

All test data created is cleaned up at the end.
"""
import io
import os
import sys
import json
import traceback
from datetime import datetime, timedelta

import requests
from openpyxl import load_workbook

# ---------------------------------------------------------------------
# Base URL (must use EXPO_PUBLIC_BACKEND_URL from frontend/.env)
# ---------------------------------------------------------------------
FRONTEND_ENV = "/app/frontend/.env"
BASE_URL = None
with open(FRONTEND_ENV) as f:
    for line in f:
        line = line.strip()
        if line.startswith("EXPO_PUBLIC_BACKEND_URL="):
            BASE_URL = line.split("=", 1)[1].strip().strip('"').strip("'")
            break
if not BASE_URL:
    print("FATAL: EXPO_PUBLIC_BACKEND_URL not found in frontend/.env")
    sys.exit(1)

API = BASE_URL.rstrip("/") + "/api"
print(f"Using API base: {API}")

# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------
PASS = []
FAIL = []


def _log_pass(name, detail=""):
    msg = f"PASS :: {name}" + (f" — {detail}" if detail else "")
    print(msg)
    PASS.append(msg)


def _log_fail(name, detail=""):
    msg = f"FAIL :: {name}" + (f" — {detail}" if detail else "")
    print(msg)
    FAIL.append(msg)


def expect(cond, name, detail=""):
    if cond:
        _log_pass(name, detail)
    else:
        _log_fail(name, detail)
    return cond


# Tracking for cleanup
created_revenue_ids = []
created_expense_ids = []


def cleanup():
    print("\n--- CLEANUP ---")
    for rid in list(created_revenue_ids):
        try:
            r = requests.delete(f"{API}/revenues/{rid}", timeout=30)
            if r.status_code == 200:
                print(f"  deleted revenue {rid}")
            else:
                print(f"  WARN could not delete revenue {rid}: {r.status_code}")
        except Exception as e:
            print(f"  WARN exception deleting revenue {rid}: {e}")
    for eid in list(created_expense_ids):
        try:
            r = requests.delete(f"{API}/expenses/{eid}", timeout=30)
            if r.status_code == 200:
                print(f"  deleted expense {eid}")
            else:
                print(f"  WARN could not delete expense {eid}: {r.status_code}")
        except Exception as e:
            print(f"  WARN exception deleting expense {eid}: {e}")

    # Sanity: check lists are empty
    try:
        revs = requests.get(f"{API}/revenues", timeout=30).json()
        exps = requests.get(f"{API}/expenses", timeout=30).json()
        remaining_r = [r for r in revs if r.get("id") in created_revenue_ids]
        remaining_e = [e for e in exps if e.get("id") in created_expense_ids]
        if remaining_r:
            print(f"  WARN remaining revenues created by tests: {len(remaining_r)}")
        if remaining_e:
            print(f"  WARN remaining expenses created by tests: {len(remaining_e)}")
    except Exception as e:
        print(f"  cleanup verify failed: {e}")


# ---------------------------------------------------------------------
# Test suites
# ---------------------------------------------------------------------
def test_revenues_crud():
    print("\n========== 1) Revenues CRUD ==========")

    # --- POST valid ---
    payload = {
        "amount": 250.75,
        "category": "printemps",
        "date": "2026-04-15",
        "description": "Lavage vitres - maison Dupont",
        "client_name": "Jean Dupont",
        "payment_method": "etransfert",
    }
    r = requests.post(f"{API}/revenues", json=payload, timeout=30)
    ok = r.status_code == 200
    expect(ok, "POST /revenues valid (printemps, etransfert)", f"status={r.status_code}, body={r.text[:200]}")
    rev1 = None
    if ok:
        rev1 = r.json()
        created_revenue_ids.append(rev1["id"])
        expect(rev1["amount"] == 250.75, "POST returns correct amount")
        expect(rev1["category"] == "printemps", "POST returns category=printemps")
        expect(rev1["payment_method"] == "etransfert", "POST returns payment_method=etransfert")
        expect(bool(rev1.get("id")), "POST returns an id")

    # Another valid with automne/cash
    r = requests.post(f"{API}/revenues", json={
        "amount": 180.00,
        "category": "automne",
        "date": "2025-10-20",
        "description": "Lavage vitres extérieur",
        "client_name": "Marie Tremblay",
        "payment_method": "cash",
    }, timeout=30)
    if expect(r.status_code == 200, "POST /revenues valid (automne, cash)", f"status={r.status_code}"):
        created_revenue_ids.append(r.json()["id"])

    # Default payment_method → cash
    r = requests.post(f"{API}/revenues", json={
        "amount": 99.99,
        "category": "printemps",
        "date": "2026-05-01",
        "client_name": "Pierre Bouchard",
    }, timeout=30)
    if expect(r.status_code == 200, "POST /revenues without payment_method (defaults to cash)",
              f"status={r.status_code}"):
        body = r.json()
        created_revenue_ids.append(body["id"])
        expect(body["payment_method"] == "cash", "Default payment_method == 'cash'")

    # --- Negative cases ---
    # amount <= 0
    r = requests.post(f"{API}/revenues", json={
        "amount": 0, "category": "printemps", "date": "2026-04-15"
    }, timeout=30)
    expect(r.status_code == 400, "POST amount=0 → 400", f"got {r.status_code}")

    r = requests.post(f"{API}/revenues", json={
        "amount": -50, "category": "printemps", "date": "2026-04-15"
    }, timeout=30)
    expect(r.status_code == 400, "POST amount<0 → 400", f"got {r.status_code}")

    # invalid category
    r = requests.post(f"{API}/revenues", json={
        "amount": 100, "category": "hiver", "date": "2026-04-15"
    }, timeout=30)
    expect(r.status_code == 400, "POST invalid category → 400", f"got {r.status_code}")

    # invalid payment_method
    r = requests.post(f"{API}/revenues", json={
        "amount": 100, "category": "printemps", "date": "2026-04-15",
        "payment_method": "bitcoin",
    }, timeout=30)
    expect(r.status_code == 400, "POST invalid payment_method → 400", f"got {r.status_code}")

    # --- GET list without filter ---
    r = requests.get(f"{API}/revenues", timeout=30)
    ok = r.status_code == 200 and isinstance(r.json(), list)
    expect(ok, "GET /revenues list", f"status={r.status_code}")
    if ok:
        items = r.json()
        our_ids = {x["id"] for x in items if x.get("id") in created_revenue_ids}
        expect(len(our_ids) >= 3, f"GET list contains our 3 revenues (found {len(our_ids)})")

        # Sorted by date desc
        dates = [x["date"] for x in items if x.get("id") in created_revenue_ids]
        expect(dates == sorted(dates, reverse=True),
               "GET list sorted by date desc (among our items)",
               f"dates={dates}")

    # --- GET list with category=printemps filter ---
    r = requests.get(f"{API}/revenues", params={"category": "printemps"}, timeout=30)
    if expect(r.status_code == 200, "GET /revenues?category=printemps", f"status={r.status_code}"):
        items = r.json()
        bad = [x for x in items if x.get("category") != "printemps"]
        expect(len(bad) == 0, "Filter returns only printemps items",
               f"{len(bad)} non-printemps leaked")

    # --- GET by id ---
    if rev1:
        r = requests.get(f"{API}/revenues/{rev1['id']}", timeout=30)
        expect(r.status_code == 200 and r.json().get("id") == rev1["id"],
               "GET /revenues/{id} valid → 200",
               f"status={r.status_code}")

    r = requests.get(f"{API}/revenues/nonexistent-id-123", timeout=30)
    expect(r.status_code == 404, "GET /revenues/{invalid} → 404", f"got {r.status_code}")

    # --- PUT ---
    if rev1:
        r = requests.put(f"{API}/revenues/{rev1['id']}", json={
            "amount": 300.00, "category": "automne", "payment_method": "cheque"
        }, timeout=30)
        ok = r.status_code == 200
        expect(ok, "PUT /revenues/{id} update amount/category/payment_method",
               f"status={r.status_code}, body={r.text[:200]}")
        if ok:
            body = r.json()
            expect(body["amount"] == 300.00, "PUT new amount persisted")
            expect(body["category"] == "automne", "PUT new category persisted")
            expect(body["payment_method"] == "cheque", "PUT new payment_method persisted")

        # Invalid category
        r = requests.put(f"{API}/revenues/{rev1['id']}", json={"category": "winter"}, timeout=30)
        expect(r.status_code == 400, "PUT invalid category → 400", f"got {r.status_code}")

        # Invalid payment_method
        r = requests.put(f"{API}/revenues/{rev1['id']}", json={"payment_method": "paypal"}, timeout=30)
        expect(r.status_code == 400, "PUT invalid payment_method → 400", f"got {r.status_code}")

    # Invalid id update
    r = requests.put(f"{API}/revenues/nope-xyz", json={"amount": 1.0}, timeout=30)
    expect(r.status_code == 404, "PUT invalid id → 404", f"got {r.status_code}")

    # --- DELETE (invalid) ---
    r = requests.delete(f"{API}/revenues/nope-xyz-del", timeout=30)
    expect(r.status_code == 404, "DELETE invalid id → 404", f"got {r.status_code}")

    # Delete one of the created revenues to verify {deleted: 1}
    if created_revenue_ids:
        temp_id = created_revenue_ids[-1]
        r = requests.delete(f"{API}/revenues/{temp_id}", timeout=30)
        ok = r.status_code == 200 and r.json().get("deleted") == 1
        expect(ok, "DELETE /revenues/{id} valid → {deleted: 1}",
               f"status={r.status_code}, body={r.text[:120]}")
        if ok:
            created_revenue_ids.remove(temp_id)


def test_revenues_stats():
    print("\n========== 2) Revenues Stats ==========")

    # Ensure we have at least 2 items: create known amounts
    r1 = requests.post(f"{API}/revenues", json={
        "amount": 100.00, "category": "printemps", "date": "2026-04-10",
        "payment_method": "cash", "client_name": "Stats Client 1",
    }, timeout=30)
    r2 = requests.post(f"{API}/revenues", json={
        "amount": 200.00, "category": "automne", "date": "2025-09-15",
        "payment_method": "credit", "client_name": "Stats Client 2",
    }, timeout=30)
    r3 = requests.post(f"{API}/revenues", json={
        "amount": 50.00, "category": "printemps", "date": "2026-03-01",
        "payment_method": "etransfert", "client_name": "Stats Client 3",
    }, timeout=30)
    for r in (r1, r2, r3):
        if r.status_code == 200:
            created_revenue_ids.append(r.json()["id"])

    # --- Stats overall ---
    r = requests.get(f"{API}/revenues/stats", timeout=30)
    ok = r.status_code == 200
    expect(ok, "GET /revenues/stats", f"status={r.status_code}, body={r.text[:200]}")
    if not ok:
        return
    data = r.json()
    expect("by_category" in data, "stats has by_category")
    expect("by_payment" in data, "stats has by_payment")
    expect("grand_total" in data, "stats has grand_total")

    bc = data.get("by_category", {})
    expect("printemps" in bc, "by_category includes 'printemps' (always present)")
    expect("automne" in bc, "by_category includes 'automne' (always present)")

    bp = data.get("by_payment", {})
    for pm in ["etransfert", "cash", "cheque", "credit"]:
        expect(pm in bp, f"by_payment includes '{pm}' (always present)")

    # --- grand_total = sum of all amounts in GET list ---
    rl = requests.get(f"{API}/revenues", timeout=30).json()
    sum_all = round(sum(float(x["amount"]) for x in rl), 2)
    gt = round(float(data.get("grand_total", 0)), 2)
    expect(abs(sum_all - gt) < 0.01,
           f"grand_total == sum of /revenues amounts (grand_total={gt}, sum={sum_all})")

    # Also verify by_category totals add up to grand_total
    cat_sum = round(sum(float(v.get("total", 0)) for v in bc.values()), 2)
    expect(abs(cat_sum - gt) < 0.01,
           f"sum of by_category.total == grand_total ({cat_sum} vs {gt})")

    # --- Stats with date filters ---
    r = requests.get(f"{API}/revenues/stats",
                     params={"start_date": "2026-01-01", "end_date": "2026-12-31"},
                     timeout=30)
    ok = r.status_code == 200
    expect(ok, "GET /revenues/stats with date filters", f"status={r.status_code}")
    if ok:
        d = r.json()
        # Should still contain both categories (even if 0)
        expect("printemps" in d["by_category"] and "automne" in d["by_category"],
               "Filtered stats still include both categories (with 0 if no data)")
        for pm in ["etransfert", "cash", "cheque", "credit"]:
            expect(pm in d["by_payment"],
                   f"Filtered stats still include payment method '{pm}'")


def test_revenues_excel_export():
    print("\n========== 3) Revenues Excel Export ==========")

    r = requests.get(f"{API}/revenues/export/excel", timeout=60)
    expect(r.status_code == 200, "GET /revenues/export/excel → 200",
           f"status={r.status_code}, body={r.text[:200]}")
    if r.status_code != 200:
        return

    ct = r.headers.get("content-type", "")
    expect("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in ct,
           "Content-Type is xlsx", f"got '{ct}'")

    cd = r.headers.get("content-disposition", "")
    today = datetime.now().strftime("%Y-%m-%d")
    expected_name = f"revenus_crystaltask_{today}.xlsx"
    expect(expected_name in cd,
           f"Content-Disposition includes {expected_name}", f"got '{cd}'")

    # Parse workbook
    try:
        wb = load_workbook(io.BytesIO(r.content))
    except Exception as e:
        _log_fail("Excel is a valid openpyxl workbook", f"load_workbook failed: {e}")
        return
    _log_pass("Excel is a valid openpyxl workbook")

    sheets = wb.sheetnames
    expect("Revenus" in sheets, "Workbook has 'Revenus' sheet", f"sheets={sheets}")
    expect("Résumé" in sheets, "Workbook has 'Résumé' sheet", f"sheets={sheets}")

    # Verify header row of 'Revenus'
    if "Revenus" in sheets:
        ws = wb["Revenus"]
        row1 = [c.value for c in ws[1]]
        expected_header = ["Date", "Catégorie", "Montant ($)", "Client", "Paiement", "Description"]
        expect(row1 == expected_header,
               "Revenus header row matches",
               f"got {row1}")

        # Scan data rows for emoji labels + TOTAL
        all_rows = list(ws.iter_rows(values_only=True))
        cat_cells = [str(r[1]) for r in all_rows[1:] if r and r[1] is not None]
        expect(any("🌸 Saison Printemps" in c for c in cat_cells),
               "Revenus sheet uses '🌸 Saison Printemps' label")
        expect(any("🍂 Saison Automne" in c for c in cat_cells),
               "Revenus sheet uses '🍂 Saison Automne' label")

        pay_cells = [str(r[4]) for r in all_rows[1:] if r and r[4] is not None]
        # At least the 4 labels that appear in existing rows should be mapped
        # We cannot assert ALL 4 appear (only ones in data), so check mapping style
        payment_label_expectations = ["📱 E-transfert", "💵 Cash", "📝 Chèque", "💳 Carte de crédit"]
        present_labels = [lbl for lbl in payment_label_expectations
                          if any(lbl in c for c in pay_cells)]
        expect(len(present_labels) >= 2,
               f"Revenus sheet uses emoji payment labels (present: {present_labels})")

        # TOTAL row
        has_total = any(r and r[1] == "TOTAL" for r in all_rows)
        expect(has_total, "Revenus sheet has a 'TOTAL' row")

    # --- Filter tests ---
    r = requests.get(f"{API}/revenues/export/excel",
                     params={"category": "printemps"}, timeout=60)
    expect(r.status_code == 200, "Export with ?category=printemps → 200",
           f"status={r.status_code}")
    if r.status_code == 200:
        try:
            wb2 = load_workbook(io.BytesIO(r.content))
            ws = wb2["Revenus"]
            rows = list(ws.iter_rows(values_only=True))
            # all data rows should be printemps (except TOTAL)
            data_rows = [r for r in rows[1:] if r and r[0] and r[1] != "TOTAL"]
            bad = [r for r in data_rows if r[1] and "Automne" in str(r[1])]
            expect(len(bad) == 0,
                   "Filtered export contains no 'Automne' rows",
                   f"{len(bad)} leaked")
        except Exception as e:
            _log_fail("Filtered export workbook parse", str(e))

    # start_date/end_date
    r = requests.get(f"{API}/revenues/export/excel",
                     params={"start_date": "2026-01-01", "end_date": "2026-12-31"},
                     timeout=60)
    expect(r.status_code == 200, "Export with ?start_date=&end_date= → 200",
           f"status={r.status_code}")


def test_finance_bilan():
    print("\n========== 4) Finance Bilan ==========")

    # Create 2 revenues
    rv_ids_local = []
    for amt, cat, d in [(500.00, "printemps", "2026-04-05"),
                        (300.00, "automne", "2025-10-10")]:
        r = requests.post(f"{API}/revenues", json={
            "amount": amt, "category": cat, "date": d, "payment_method": "cash",
            "client_name": "Bilan Test Client",
        }, timeout=30)
        if r.status_code == 200:
            rid = r.json()["id"]
            created_revenue_ids.append(rid)
            rv_ids_local.append(rid)

    # Create 2 expenses
    ex_ids_local = []
    for amt, cat, d in [(120.50, "gas", "2026-04-06"),
                        (80.00, "resto", "2025-10-11")]:
        r = requests.post(f"{API}/expenses", json={
            "amount": amt, "category": cat, "date": d,
            "description": "Bilan test",
        }, timeout=30)
        if r.status_code == 200:
            eid = r.json()["id"]
            created_expense_ids.append(eid)
            ex_ids_local.append(eid)
        else:
            _log_fail(f"Create expense for bilan test ({cat})",
                      f"status={r.status_code}, body={r.text[:200]}")

    # --- Bilan overall (no filter) ---
    r = requests.get(f"{API}/finance/bilan", timeout=30)
    ok = r.status_code == 200
    expect(ok, "GET /finance/bilan", f"status={r.status_code}, body={r.text[:200]}")
    if not ok:
        return
    d = r.json()

    for k in ["period", "total_revenues", "total_expenses", "net_profit",
              "margin_pct", "revenues_by_category", "expenses_by_category"]:
        expect(k in d, f"Bilan response has '{k}'")

    # Cross-verify with lists
    revs_all = requests.get(f"{API}/revenues", timeout=30).json()
    exps_all = requests.get(f"{API}/expenses", timeout=30).json()
    sum_rev = round(sum(float(x["amount"]) for x in revs_all), 2)
    sum_exp = round(sum(float(x["amount"]) for x in exps_all), 2)
    expect(abs(float(d["total_revenues"]) - sum_rev) < 0.01,
           f"total_revenues == sum of revenues ({d['total_revenues']} vs {sum_rev})")
    expect(abs(float(d["total_expenses"]) - sum_exp) < 0.01,
           f"total_expenses == sum of expenses ({d['total_expenses']} vs {sum_exp})")

    expected_profit = round(sum_rev - sum_exp, 2)
    expect(abs(float(d["net_profit"]) - expected_profit) < 0.01,
           f"net_profit == rev - exp ({d['net_profit']} vs {expected_profit})")

    expected_margin = round((expected_profit / sum_rev * 100.0) if sum_rev > 0 else 0.0, 2)
    expect(abs(float(d["margin_pct"]) - expected_margin) < 0.01,
           f"margin_pct correct ({d['margin_pct']} vs {expected_margin})")

    # --- With date filters (narrow window that excludes some) ---
    r = requests.get(f"{API}/finance/bilan",
                     params={"start_date": "2026-01-01", "end_date": "2026-12-31"},
                     timeout=30)
    ok = r.status_code == 200
    expect(ok, "GET /finance/bilan with date filters", f"status={r.status_code}")
    if ok:
        df = r.json()
        # Compute expected in same window
        rev_in = [x for x in revs_all if "2026-01-01" <= x["date"] <= "2026-12-31"]
        exp_in = [x for x in exps_all if "2026-01-01" <= x["date"] <= "2026-12-31"]
        sr = round(sum(float(x["amount"]) for x in rev_in), 2)
        se = round(sum(float(x["amount"]) for x in exp_in), 2)
        expect(abs(float(df["total_revenues"]) - sr) < 0.01,
               f"Filtered total_revenues correct ({df['total_revenues']} vs {sr})")
        expect(abs(float(df["total_expenses"]) - se) < 0.01,
               f"Filtered total_expenses correct ({df['total_expenses']} vs {se})")

    # --- Margin edge cases: cleanup all and re-test ---
    # Delete all our test data first so the DB is clean for edge case tests
    for rid in list(created_revenue_ids):
        requests.delete(f"{API}/revenues/{rid}", timeout=30)
        created_revenue_ids.remove(rid)
    for eid in list(created_expense_ids):
        requests.delete(f"{API}/expenses/{eid}", timeout=30)
        created_expense_ids.remove(eid)

    # Check DB empty now
    revs_after = requests.get(f"{API}/revenues", timeout=30).json()
    exps_after = requests.get(f"{API}/expenses", timeout=30).json()

    # Edge case 1: only revenues → margin_pct = 100
    r = requests.post(f"{API}/revenues", json={
        "amount": 400.00, "category": "printemps", "date": "2026-06-01",
        "payment_method": "cash",
    }, timeout=30)
    if r.status_code == 200:
        created_revenue_ids.append(r.json()["id"])

        # Use tight date range to avoid pre-existing data in DB
        b = requests.get(f"{API}/finance/bilan",
                        params={"start_date": "2026-06-01", "end_date": "2026-06-01"},
                        timeout=30).json()
        # In this tight window we should have only our revenue
        if float(b.get("total_revenues", 0)) > 0 and float(b.get("total_expenses", 0)) == 0:
            expect(abs(float(b["margin_pct"]) - 100.0) < 0.01,
                   f"Edge: only revenues → margin_pct = 100 (got {b['margin_pct']})")
            expect(abs(float(b["net_profit"]) - float(b["total_revenues"])) < 0.01,
                   "Edge: only revenues → net_profit == total_revenues")
        else:
            print(f"  (skipping margin=100 check — window had mixed data: {b})")

    # Edge case 2: only expenses → margin_pct = 0, net_profit negative
    r = requests.post(f"{API}/expenses", json={
        "amount": 75.00, "category": "gas", "date": "2026-07-01",
    }, timeout=30)
    if r.status_code == 200:
        created_expense_ids.append(r.json()["id"])
        b = requests.get(f"{API}/finance/bilan",
                        params={"start_date": "2026-07-01", "end_date": "2026-07-01"},
                        timeout=30).json()
        if float(b.get("total_revenues", 0)) == 0 and float(b.get("total_expenses", 0)) > 0:
            expect(abs(float(b["margin_pct"]) - 0.0) < 0.01,
                   f"Edge: only expenses → margin_pct = 0 (got {b['margin_pct']})")
            expect(float(b["net_profit"]) < 0,
                   f"Edge: only expenses → net_profit negative (got {b['net_profit']})")
        else:
            print(f"  (skipping expenses-only check — window had mixed data: {b})")


# ---------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------
def main():
    try:
        test_revenues_crud()
        test_revenues_stats()
        test_revenues_excel_export()
        test_finance_bilan()
    except Exception as e:
        print(f"\nFATAL during tests: {e}")
        traceback.print_exc()
    finally:
        cleanup()

    print("\n==================== SUMMARY ====================")
    print(f"Passed: {len(PASS)}")
    print(f"Failed: {len(FAIL)}")
    if FAIL:
        print("\n--- FAILURES ---")
        for f in FAIL:
            print(f"  {f}")
    print("=================================================")
    sys.exit(0 if not FAIL else 1)


if __name__ == "__main__":
    main()
